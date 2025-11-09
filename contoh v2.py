
#APP.py

# Core Flask and Database
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify, abort
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import (
    text, create_engine, func, case, or_, and_, extract, 
    literal_column, inspect, union_all, MetaData
)
from sqlalchemy.sql import text as sql_text
from sqlalchemy.exc import OperationalError
from sqlalchemy.dialects.mysql import insert
import pymysql

# Data Processing
import pandas as pd
import numpy as np

# Date and Time
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import time

# File Handling
from io import BytesIO
import io
import os
import tempfile

# Excel/PDF Processing
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import inch, cm, mm
from xhtml2pdf import pisa

# Plotting
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

# Security and Utilities
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from functools import wraps
from html import escape
from markupsafe import escape
from bs4 import BeautifulSoup
from urllib.parse import quote_plus

# Other
import base64
import threading
import traceback
import logging
import locale
import sys
import re
import requests

# Set locale untuk format mata uang Indonesia
try:
    locale.setlocale(locale.LC_ALL, 'id_ID.UTF-8')
except locale.Error:
    locale.setlocale(locale.LC_ALL, 'Indonesian_indonesia.1252')


# Set locale to Indonesian for currency formatting
try:
    locale.setlocale(locale.LC_ALL, 'id_ID.UTF-8')
except locale.Error:
    locale.setlocale(locale.LC_ALL, 'Indonesian_indonesia.1252')


# --- HELPER FUNCTION FOR CURRENCY ---
def format_rp(value):
    if value is None or value == 0:
        return "Rp. - "
    return locale.format_string("Rp. %d", value, grouping=True)

# matplotlib.use('Agg')  # Gunakan non-GUI backend!
app = Flask(__name__)
# def get_db_connection2():
#     connection = pymysql.connect(
#         host=db_config['host'],
#         user=db_config['user'],
#         password=db_config['password'],
#         database=db_config['database'],
#         charset='utf8mb4',
#         cursorclass=pymysql.cursors.DictCursor
#     )
#     return connection
# # Middleware untuk proteksi login

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get('loggedin'):
            flash('Anda harus login terlebih dahulu.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return wrapper


# def get_engine(db_name):
#     db_config = {
#         'host': 'localhost',
#         'user': 'root',
#         'password': '',  # Leave empty if no password
#         'database': 'billing_v2'
#     }
#     return create_engine(
#         # f"mysql+pymysql://{db_config['user']}:{db_config['password']}@{db_config['host']}/{db_config['database']}"
#          f"mysql+pymysql://{db_config['user']}:{db_config['password']}@{db_config['host']}/{db_config['database']}",
#         connect_args={"connect_timeout": 300},
#         pool_pre_ping=True
#     )

# db_config = {
#     'host': 'localhost',
#     'user': 'root',  # ganti dengan username database Anda
#     'password': '',  # ganti dengan password database Anda
#     'database': 'billing_v2'  # ganti dengan nama database Anda
# }

@app.route('/hapus_user', methods=['POST'])
def hapus_user():
    data = request.get_json()
    try:
        user = TbUserPksr.query.get(data['id'])
        if not user:
            return jsonify({'status': 'error', 'message': 'User tidak ditemukan'})

        db.session.delete(user)
        db.session.commit()
        return jsonify({'status': 'success'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

def login_required_pksr(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get('loggedin'):
            flash('Tim PKSR Anda harus login terlebih dahulu.', 'warning')
            return redirect(url_for('login_pksr'))
        return f(*args, **kwargs)
    return wrapper

# Config di sini
# app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://root:@127.0.0.1/telupsin_billing2'
# app.config['SQLALCHEMY_DATABASE_URI'] = (
#     f"mysql+pymysql://{db_config['user']}:{db_config['password']}@{db_config['host']}/{db_config['database']}"
# )
# app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# # 2. Konfigurasi database
# db_config = {
#     'host': 'localhost',
#     'user': 'root',
#     'password': '',
#     'database': 'billing_v2'
# }

# # 3. Set konfigurasi SQLALCHEMY
# app.config['SQLALCHEMY_DATABASE_URI'] = (
#     f"mysql+pymysql://{db_config['user']}:{db_config['password']}@{db_config['host']}/{db_config['database']}"
#     f"?charset=utf8mb4&connect_timeout=10&read_timeout=120&write_timeout=120"
# )
# app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
# app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
#     "pool_pre_ping": True,
#     "pool_recycle": 280,
#     "pool_use_lifo": True,
#     "pool_reset_on_return": "rollback",
#     "pool_size": 5,
#     "max_overflow": 10,
#     "connect_args": {
#         "read_timeout": 120,
#         "write_timeout": 120,
#     },
# }

# # 4. BUAT instance SQLAlchemy
# db = SQLAlchemy(app)  # ← INI YANG PENTING!

# # 5. Fungsi teardown (sudah ada di kode Anda)
# @app.teardown_appcontext
# def shutdown_session(exc=None):
#     try:
#         db.session.remove()
#     except Exception:
#         pass

# db_config = {
#     'host': 'localhost',
#     'user': 'root',
#     'password': '',
#     'database': 'billing_v2'
# }

# # Untuk pymysql langsung
# def get_db_connection2():
#     return pymysql.connect(
#         host=db_config['host'],
#         user=db_config['user'],
#         password=db_config['password'],
#         database=db_config['database'],
#         charset='utf8mb4',
#         cursorclass=pymysql.cursors.DictCursor
#     )

# # Untuk SQLAlchemy
# app.config['SQLALCHEMY_DATABASE_URI'] = (
#     f"mysql+pymysql://{db_config['user']}:{db_config['password']}@{db_config['host']}/{db_config['database']}?charset=utf8mb4"
# )
# app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
# db = SQLAlchemy(app)

# # Untuk Pandas read_sql / raw query
# engine = create_engine(app.config['SQLALCHEMY_DATABASE_URI'], pool_pre_ping=True)


# # Tambahkan charset + timeout driver langsung di URI
# app.config['SQLALCHEMY_DATABASE_URI'] = (
#     f"mysql+pymysql://{db_config['user']}:{db_config['password']}@{db_config['host']}/{db_config['database']}"
#     f"?charset=utf8mb4&connect_timeout=10&read_timeout=120&write_timeout=120"
# )
# app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# # Pool & koneksi: cegah stale connection, kurangi efek wait_timeout
# app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
#     "pool_pre_ping": True,         # auto ganti koneksi mati
#     "pool_recycle": 280,           # < wait_timeout (shared host sering 300s)
#     "pool_use_lifo": True,         # pakai koneksi paling “fresh”
#     "pool_reset_on_return": "rollback",
#     "pool_size": 5,
#     "max_overflow": 10,
#     "connect_args": {              # redundansi timeout (beberapa env butuh)
#         "read_timeout": 120,
#         "write_timeout": 120,
#     },
# }
# ==========================================
# DATABASE CONFIGURATION (FINAL CLEAN VERSION)
# ==========================================
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import create_engine
import pymysql

# Konfigurasi Database Utama
db_config = {
    'host': 'localhost',
    'user': 'root',
    'password': '',
    'database': 'billing_v2'
}

# Konfigurasi SQLAlchemy
app.config['SQLALCHEMY_DATABASE_URI'] = (
    f"mysql+pymysql://{db_config['user']}:{db_config['password']}@{db_config['host']}/{db_config['database']}?"
    "charset=utf8mb4&connect_timeout=10&read_timeout=120&write_timeout=120"
)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    "pool_pre_ping": True,
    "pool_recycle": 280,
    "pool_use_lifo": True,
    "pool_reset_on_return": "rollback",
    "pool_size": 5,
    "max_overflow": 10,
    "connect_args": {
        "read_timeout": 120,
        "write_timeout": 120,
    },
}

# Buat instance SQLAlchemy global
db = SQLAlchemy(app)

# Buat koneksi langsung (pymysql)
def get_db_connection2():
    return pymysql.connect(
        host=db_config['host'],
        user=db_config['user'],
        password=db_config['password'],
        database=db_config['database'],
        charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor
    )

# Buat engine global (untuk Pandas read_sql)
engine = create_engine(app.config['SQLALCHEMY_DATABASE_URI'], pool_pre_ping=True)

# Teardown untuk menutup koneksi aman
@app.teardown_appcontext
def shutdown_session(exc=None):
    try:
        db.session.remove()
    except Exception:
        pass



@app.teardown_appcontext
def shutdown_session(exc=None):
    try:
        db.session.remove()
    except Exception:
        pass




@app.template_filter('has_access')
def has_access(tables, target):
    return target in tables.split(',') if tables else False


@app.route('/user_manajemen')
def user_manajemen():
    # Ambil username login dari session
    username = session.get("username")

    # Query tb_user_pksr
    user = TbUserPksr.query.filter_by(username=username).first()
    
    if not user:
        return "User tidak ditemukan!", 403
    
    # Mulai query utama
    query = TbUserPksr.query

    # # Filter IDPEL kalau ada
    # idpel_filter = request.args.get('idpel', default="", type=str)
    # if idpel_filter:
    #     query = query.filter(KppPksr.IDPEL.contains(idpel_filter))
    
    # Filter UNITUP
    # if user.unitup != "ALL":
    #     query = query.filter(KppPksr.UNITUP == user.unitup)
    
    # Pagination
    page = request.args.get('page', 1, type=int)
    pagination = query.paginate(page=page, per_page=10)
    data = pagination.items

    return render_template(
        'pksr/user_manajemen.html',
        data=data,
        # idpel_filter=idpel_filter
        current_page=page,
        total_pages=pagination.pages
    )

# Define this at the top of your file, after imports
TABLE_MAPPING = {
    'tel_skw': 'billing_skw2',
    'tel_pmk': 'billing_pmk',
    'tel_sbs': 'billing_sbs',
    'tel_bkg': 'billing_bkg',
    'tel_skr': 'billing_skr',
    'tel_sdr': 'billing_sdr',
}

NOMET_MAPPING = {
    'tel_skw': 'nomet_skw',
    'tel_pmk': 'nomet_pmk',
    'tel_sbs': 'nomet_sbs',
    'tel_bkg': 'nomet_bkg',
    'tel_skr': 'nomet_skr',
    'tel_sdr': 'nomet_sdr',
}

DPM_MAPPING = {
    'tel_skw': 'dpm_skw',
    'tel_pmk': 'dpm_pmk',
    'tel_sbs': 'dpm_sbs',
    'tel_bkg': 'dpm_bkg',
    'tel_skr': 'dpm_skr',
    'tel_sdr': 'dpm_sdr',
}

app.secret_key = '040104'
# # Database config
# db_config = {
#     'host': '127.0.0.1',  # Instead of 'localhost'
#     # 'user': 'telupsin_telupsin',
#     'user': 'root',
#     # 'password': 'WjMD3#B88ce1',
#     'password': '',
#     'database': 'telupsin_billing2'
# }

# def get_db():
#     return pymysql.connect(**db_config)


# try:
#     conn = pymysql.connect(
#         host='127.0.0.1',
#         user='root',
#         # user='telupsin_telupsin',
#         # password='WjMD3#B88ce1',
#         password='',
#         database='telupsin_billing2',
#         port=3306
#     )
#     print("✅ Connection successful!")
#     conn.close()
# except Exception as e:
#     print(f"❌ Connection failed: {e}")

# engine = create_engine(
#     f"mysql+pymysql://{db_config['user']}:{db_config['password']}@{db_config['host']}/{db_config['database']}",
#     connect_args={"connect_timeout": 300},
#     pool_pre_ping=True
# )




def normalize_idpel(idpel):
    return str(idpel).replace('\xa0', '').replace(' ', '').replace(' ', '').strip().zfill(12)





user_access = {
    'Tel_Skw': {
        'password': 'telskw@123', 
        'tables': ['billing_skw2', 'nomet_skw', 'dpm_skw']
    },
    'Tel_Pmk': {
        'password': 'telpmk@123', 
        'tables': ['billing_pmk', 'nomet_pmk', 'dpm_pmk']
    },
    'Tel_Sbs': {
        'password': 'telsbs@123', 
        'tables': ['billing_sbs', 'nomet_sbs', 'dpm_sbs']
    },
    'Tel_Bkg': {
        'password': 'telbkg@123', 
        'tables': ['billing_bkg', 'nomet_bkg', 'dpm_bkg']
    },
    'Tel_Skr': {
        'password': 'telskr@123', 
        'tables': ['billing_skr', 'nomet_skr', 'dpm_skr']
    },
    'Tel_Sdr': {
        'password': 'telsdr@123', 
        'tables': ['billing_sdr', 'nomet_sdr', 'dpm_sdr']
    },
    'Tel_Up3': {
        'password': 'telup3@123', 
        'tables': [
            'billing_skw2', 'nomet_skw', 'dpm_skw',
            'billing_pmk', 'nomet_pmk', 'dpm_pmk',
            'billing_sbs', 'nomet_sbs', 'dpm_sbs',
            'billing_bkg', 'nomet_bkg', 'dpm_bkg',
            'billing_skr', 'nomet_skr', 'dpm_skr',
            'billing_sdr', 'nomet_sdr', 'dpm_sdr'
        ]
    }
}

@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        # Check if username exists and password matches
        if username in user_access and user_access[username]['password'] == password:
            session['loggedin'] = True
            session['username'] = username
            session['tables'] = user_access[username]['tables']  # Store the accessible tables for the user
            flash(f'Welcome, {username}!', 'success')
            return redirect(url_for('admin_v2'))
        else:
            flash('SnapBill Invalid username or password!', 'danger')
            return render_template('login.html')

    return render_template('login.html')

#dashboard monitoring
# Tables list
tables = [
    ("billing_skw2", "ULP Singkawang"),
    ("billing_pmk", "ULP Pemangkat"),
    ("billing_sbs", "ULP Sambas"),
    ("billing_bkg", "ULP Bengkayang"),
    ("billing_skr", "ULP Sekura"),
    ("billing_sdr", "ULP Sei Duri")
]



def fetch_status_recap(table, filter_marking=False):
    conn = get_db_connection2()
    cursor = conn.cursor()
    base_query = f"""
        SELECT `HASIL PEMERIKSAAN` AS status, COUNT(*) AS jumlah
        FROM {table}
        WHERE KET IN ('NAIK', 'TURUN', 'DIV/NA') 
          AND (KDKELOMPOK IN ('1','2','3','4','5','6','7','8','P'))
    """
    if filter_marking:
        base_query += " AND `DLPD_HITUNG` IS NOT NULL AND `DLPD_HITUNG` <> ''"

    base_query += " GROUP BY `HASIL PEMERIKSAAN`"

    cursor.execute(base_query)
    results = cursor.fetchall()
    cursor.close()
    conn.close()
    return results


def combine_recap_data(all_data):
    combined = {}
    total = 0
    for data in all_data:
        for row in data:
            status = row['status'] if row['status'] else 'BELUM DIISI'
            if status == 'AMAN':
                continue  # Abaikan status AMAN
            jumlah = row['jumlah']
            combined[status] = combined.get(status, 0) + jumlah
            total += jumlah

    recap_list = []
    for status, jumlah in combined.items():
        perc = round((jumlah / total) * 100, 1) if total > 0 else 0
        recap_list.append({
            'status': status,
            'jumlah': jumlah,
            'persentase': perc
        })

    # Urutkan dari jumlah terbesar
    recap_list = sorted(recap_list, key=lambda x: x['jumlah'], reverse=True)
    return recap_list, total

#UPLOAD ULANG KOREKSIAN
def generate_pivot_status_koreksi(table, filter_marking=False):
    conn = get_db_connection2()
    cursor = conn.cursor()

    base_query = f"""
        SELECT BLTH, KDKELOMPOK, `HASIL PEMERIKSAAN` AS status, COUNT(*) AS jumlah
        FROM `{table}`
        WHERE MARKING_KOREKSI > 0
          AND KDKELOMPOK IN ('1','2','3','4','5','6','7','8','P')
    """
    # if filter_marking:
    #     base_query += """
    #         AND `DLPD_HITUNG` IS NOT NULL
    #         AND LENGTH(TRIM(`DLPD_HITUNG`)) > 0
    #     """

    base_query += " GROUP BY BLTH, KDKELOMPOK, `HASIL PEMERIKSAAN`"

    cursor.execute(base_query)
    results = cursor.fetchall()
    cursor.close()
    conn.close()

    pivot = {}
    statuses = set()

    for row in results:
        status = row['status'] if row['status'] else 'BELUM DIISI'
        label = f"{row['BLTH']} - {row['KDKELOMPOK']}"
        jumlah = row['jumlah']
        statuses.add(status)

        if label not in pivot:
            pivot[label] = {}
        pivot[label][status] = pivot[label].get(status, 0) + jumlah

    return pivot, sorted(statuses)

#
def generate_pivot_status(table, filter_marking=False):
    conn = get_db_connection2()
    cursor = conn.cursor()
    base_query = f"""
        SELECT BLTH, KDKELOMPOK, `HASIL PEMERIKSAAN` AS status, COUNT(*) AS jumlah
        FROM {table}
        WHERE KET IN ('NAIK', 'TURUN', 'DIV/NA','AMAN')
          AND (KDKELOMPOK IN ('1','2','3','4','5','6','7','8','P'))
    """
    if filter_marking:
        base_query += " AND `DLPD_HITUNG` IS NOT NULL AND LENGTH(TRIM(`DLPD_HITUNG`)) > 0"

    base_query += " GROUP BY BLTH, KDKELOMPOK, `HASIL PEMERIKSAAN`"

    cursor.execute(base_query)
    results = cursor.fetchall()
    cursor.close()
    conn.close()

    pivot = {}
    statuses = set()

    for row in results:
        status = row['status'] if row['status'] else 'BELUM DIISI'
        # if status == 'AMAN':
        #     continue
        label = f"{row['BLTH']} - {row['KDKELOMPOK']}"
        jumlah = row['jumlah']
        statuses.add(status)
        if label not in pivot:
            pivot[label] = {}
        pivot[label][status] = pivot[label].get(status, 0) + jumlah

    return pivot, sorted(statuses)


#pivot dlpd hitung
def generate_pivot_status3(table, filter_marking=True):
    conn = get_db_connection2()
    cursor = conn.cursor()
    base_query = f"""
        SELECT `DLPD_HITUNG`, `HASIL PEMERIKSAAN` AS status, COUNT(*) AS jumlah
        FROM {table}
        WHERE KET IN ('NAIK', 'TURUN', 'DIV/NA','AMAN')
          AND KDKELOMPOK IN ('1','2','3','4','5','6','7','8','P')
    """
    if filter_marking:
        base_query += " AND `DLPD_HITUNG` IS NOT NULL AND LENGTH(TRIM(`DLPD_HITUNG`)) > 0"

    base_query += " GROUP BY `DLPD_HITUNG`, `HASIL PEMERIKSAAN`"

    cursor.execute(base_query)
    results = cursor.fetchall()
    cursor.close()
    conn.close()

    rows = [
        'JN_720up',
        'Cek_Pecahan',
        'Stan_Mundur',
        'Naik_50%Up',
        'Cek_DIV/NA',
        'Turun_50%Down',
        'kWh_Nol',
        'JN_40Down'
    ]

    pivot = {r: {} for r in rows}
    statuses = set()

    for row in results:
        dlpd = row['DLPD_HITUNG']
        status = row['status'] if row['status'] else 'BELUM DIISI'
        jumlah = row['jumlah']
        statuses.add(status)
        if dlpd in pivot:
            pivot[dlpd][status] = jumlah

    return pivot, sorted(statuses)

def generate_pivot_ganda(table):
    conn = get_db_connection2()
    cursor = conn.cursor()

    base_query = f"""
        SELECT 
            COALESCE(`DLPD_HITUNG`, '') AS DLPD_HITUNG, 
            COALESCE(`HASIL PEMERIKSAAN`, '') AS status, 
            COUNT(*) AS jumlah
        FROM {table}
        WHERE KET IN ('NAIK', 'TURUN', 'DIV/NA','AMAN')
          AND KDKELOMPOK IN ('I','A')
        GROUP BY DLPD_HITUNG, `HASIL PEMERIKSAAN`
    """

    cursor.execute(base_query)
    results = cursor.fetchall()
    cursor.close()
    conn.close()

    # Daftar nilai tetap yang diharapkan
    rows = [
        'JN_720up',
        'Cek_Pecahan',
        'Stan_Mundur',
        'Naik_50%Up',
        'Cek_DIV/NA',
        'Turun_50%Down',
        'kWh_Nol',
        'JN_40Down',
        ''  # Tambahkan entri kosong (tanpa DLPD_HITUNG)
    ]

    # Inisialisasi pivot
    pivot = {r: {} for r in rows}
    statuses = set()

    for row in results:
        dlpd = row['DLPD_HITUNG'] or ''  # Kosongkan NULL
        status = row['status'] or 'BELUM DIISI'
        jumlah = row['jumlah']
        statuses.add(status)

        # Tambahkan data ke pivot
        if dlpd not in pivot:
            pivot[dlpd] = {}
        pivot[dlpd][status] = jumlah

    return pivot, sorted(statuses)





def summarize_by_ket(data, total):
    summary = {}
    for key in ['NAIK', 'TURUN', 'DIV/NA']:
        jumlah = 0
        for row in data:
            if row['status'] == key:
                jumlah = row['jumlah']
                break
        persentase = round((jumlah / total) * 100, 1) if total > 0 else 0
        summary[key] = {'jumlah': jumlah, 'persentase': persentase}
    return summary

def fetch_ket_summary(table, filter_marking=False):
    conn = get_db_connection2()
    cursor = conn.cursor()
    base_query = f"""
        SELECT KET, COUNT(*) AS jumlah
        FROM {table}
        WHERE KET IN ('NAIK', 'TURUN', 'DIV/NA')
          AND KDKELOMPOK IN ('1','2','3','4','5','6','7','8','P')
    """
    if filter_marking:
        base_query += " AND DLPD_HITUNG IS NOT NULL AND LENGTH(TRIM(DLPD_HITUNG)) > 0"

    base_query += " GROUP BY KET"
    cursor.execute(base_query)
    results = cursor.fetchall()
    cursor.close()
    conn.close()
    return results
def summarize_ket_grouped(ket_data_list):
    total = 0
    summary = {'NAIK': 0, 'TURUN': 0, 'DIV/NA': 0}
    for data in ket_data_list:
        for row in data:
            ket = row['KET']
            jumlah = row['jumlah']
            if ket in summary:
                summary[ket] += jumlah
                total += jumlah

    result = {}
    for ket in ['NAIK', 'TURUN', 'DIV/NA']:
        jumlah = summary[ket]
        persen = round((jumlah / total) * 100, 1) if total > 0 else 0
        result[ket] = {'jumlah': jumlah, 'persentase': persen}
    return result, total

def fetch_ket_summary2(table, filter_marking=False):
    conn = get_db_connection2()
    cursor = conn.cursor()
    base_query = f"""
        SELECT DLPD_HITUNG, COUNT(*) AS jumlah
        FROM {table}
        WHERE KET IN ('NAIK', 'TURUN', 'DIV/NA')
        AND KDKELOMPOK IN ('1','2','3','4','5','6','7','8','P')
    """
    if filter_marking:
        base_query += " AND DLPD_HITUNG IS NOT NULL AND LENGTH(TRIM(DLPD_HITUNG)) > 0"

    base_query += " GROUP BY DLPD_HITUNG"
    cursor.execute(base_query)
    results = cursor.fetchall()
    cursor.close()
    conn.close()
    return results


def summarize_ket_grouped2(ket_data_list):
    total = 0
    summary = {

        'JN_720up':0,
        'Cek_Pecahan':0,
        'Stan_Mundur':0,
        'Naik_50%Up':0,
        'Cek_DIV/NA':0,
        'Turun_50%Down':0,
        'kWh_Nol':0,
        'JN_40Down':0
    }
    for data in ket_data_list:
        for row in data:
            ket = row['DLPD_HITUNG']
            jumlah = row['jumlah']
            if ket in summary:
                summary[ket] += jumlah
                total += jumlah

    result = {}
    for ket in summary.keys():
        jumlah = summary[ket]
        persen = round((jumlah / total) * 100, 1) if total > 0 else 0
        result[ket] = {'jumlah': jumlah, 'persentase': persen}
    return result, total
ulp_to_billing2 = {
    'ulp_skw': 'billing_skw2',
    'ulp_pmk': 'billing_pmk',
    'ulp_sbs': 'billing_sbs',
    'ulp_bkg': 'billing_bkg',
    'ulp_skr': 'billing_skr',
    'ulp_sdr': 'billing_sdr',
}

user_access2 = {
    'tel_skw': ['ulp_skw'],
    'tel_pmk': ['ulp_pmk'],
    'tel_sbs': ['ulp_sbs'],
    'tel_bkg': ['ulp_bkg'],
    'tel_skr': ['ulp_skr'],
    'tel_sdr': ['ulp_sdr'],
    'tel_up3': list(ulp_to_billing2.keys())
}

@app.route("/dashboard_monitoring")
@login_required
def index():
    user = session.get('username', '').lower()
    allowed_ulps = user_access2.get(user, [])

    tables = [(ulp_to_billing2[ulp], ulp.replace('ulp_', 'ULP ').upper()) for ulp in allowed_ulps]
    # Rekap Semua ULP tanpa filter
    all_tables_data = [fetch_status_recap(t[0], filter_marking=False) for t in tables]
    recap_all, total_all = combine_recap_data(all_tables_data)

    # Rekap Semua ULP dengan filter marking
    all_tables_marking_data = [fetch_status_recap(t[0], filter_marking=True) for t in tables]
    recap_marking, total_marking = combine_recap_data(all_tables_marking_data)

    # Rekap KET (NAIK/TURUN/DIV/NA) dari semua tabel
    ket_all_data = [fetch_ket_summary(t[0], filter_marking=False) for t in tables]
    summary_all, _ = summarize_ket_grouped(ket_all_data)

    #  # REKAP DLPD HITUNG (filter_marking=True)
    # ket_dlpd = [fetch_ket_summary2(t[0], filter_marking=True) for t in tables]
    # summary_dlpd, summary_dlpd_total = summarize_ket_grouped2(ket_dlpd)

    ket_marking_data = [fetch_ket_summary(t[0], filter_marking=True) for t in tables]
    summary_marking, _ = summarize_ket_grouped(ket_marking_data)


    # Per-table recap data
    per_table_recap = {}
    per_table_marking = {}
    for tbl, _ in tables:
        per_table_recap[tbl] = fetch_status_recap(tbl, filter_marking=False)
        per_table_marking[tbl] = fetch_status_recap(tbl, filter_marking=True)

    # Pivot per table NAIK TURUN DIV/NA
    pivot_tables_all = {}

    pivot_tables_marking = {}
    for tbl, _ in tables:
        pivot_tables_all[tbl] = generate_pivot_status(tbl, filter_marking=False)
        pivot_tables_marking[tbl] = generate_pivot_status(tbl, filter_marking=True)
    #PIVOT DLPD HITUNG SEMUA DATA
    pivot_tables_dlpd = {}
    for tbl, _ in tables:
        pivot_tables_dlpd[tbl] = generate_pivot_status3(tbl, filter_marking=True)
    

    #PIVOT DLPD HITUNG PLG GANDA
    pivot_tables_ganda = {}
    for tbl, _ in tables:
        pivot_tables_ganda[tbl] = generate_pivot_ganda(tbl)

     #PIVOT DLPD HITUNG SEMUA DATA 
    pivot_tables_koreksi = {}
    for tbl, _ in tables:
        pivot_tables_koreksi[tbl] = generate_pivot_status_koreksi(tbl, filter_marking=True)

    return render_template("index.html",
        tables=tables,
        recap_all=recap_all, total_all=total_all, summary_all=summary_all,
        recap_marking=recap_marking, total_marking=total_marking, summary_marking=summary_marking,
        per_table_recap=per_table_recap,
        per_table_marking=per_table_marking,
        # optional: bisa diaktifkan kalau ingin menampilkan
        # summary_dlpd=summary_dlpd, summary_dlpd_total=summary_dlpd_total,
        pivot_tables_all=pivot_tables_all,
        pivot_tables_marking=pivot_tables_marking,
        pivot_tables_dlpd=pivot_tables_dlpd,  # ⬅️ TAMBAHKAN INI
        pivot_tables_koreksi=pivot_tables_koreksi,
        pivot_tables_ganda=pivot_tables_ganda,  # ⬅️ TAMBAHKAN INI
        username=session.get('username', '').lower()
    )



#POP UP DI DASHBOARD MONITORING
@app.route('/get_detail_pelanggan')
@login_required
def get_detail_pelanggan():
    blth = request.args.get('blth')
    kdkelompok = request.args.get('kdkelompok')
    hasil_pemeriksaan = request.args.get('hasil_pemeriksaan')
    table = request.args.get('table')
    username = session.get('username', '').lower()

    # Validasi awal
    if not blth or not kdkelompok or not table:
        return jsonify({'error': 'Parameter wajib tidak lengkap'}), 400

    if kdkelompok not in [str(i) for i in range(1, 9)] + ['P']:
        return jsonify({'error': 'Invalid KDKELOMPOK value'}), 400

    ulp_to_billing = {
        'ulp_skw': 'billing_skw2',
        'ulp_pmk': 'billing_pmk',
        'ulp_sbs': 'billing_sbs',
        'ulp_bkg': 'billing_bkg',
        'ulp_skr': 'billing_skr',
        'ulp_sdr': 'billing_sdr',
    }

    user_access = {
        'tel_skw': ['ulp_skw'],
        'tel_pmk': ['ulp_pmk'],
        'tel_sbs': ['ulp_sbs'],
        'tel_bkg': ['ulp_bkg'],
        'tel_skr': ['ulp_skr'],
        'tel_sdr': ['ulp_sdr'],
        'tel_up3': list(ulp_to_billing.keys())
    }

    # Deteksi base table dan flag marking
    filter_marking = table.endswith('_marking')
    base_table = table[:-8] if filter_marking else table

    # Validasi hak akses user
    allowed_tables = [ulp_to_billing[ulp] for ulp in user_access.get(username, [])]
    if base_table not in allowed_tables:
        return jsonify({'error': 'Akses ditolak ke tabel ini'}), 403
    
    dlpd_hitung = request.args.get('dlpd_hitung')  # tambahkan ini

    # Query SQL
    conn = get_db_connection2()
    cursor = conn.cursor()

    query = f"""
        SELECT * FROM {base_table}
        WHERE BLTH = %s AND KDKELOMPOK = %s
          AND KET IN ('NAIK', 'TURUN', 'DIV/NA')
    """
    params = [blth, kdkelompok]

    if filter_marking:
        query += " AND DLPD_HITUNG IS NOT NULL AND LENGTH(TRIM(DLPD_HITUNG)) > 0"

    if hasil_pemeriksaan is not None:
        if hasil_pemeriksaan == 'BELUM DIISI':
            query += " AND (`HASIL PEMERIKSAAN` IS NULL OR `HASIL PEMERIKSAAN` = '')"
        else:
            query += " AND `HASIL PEMERIKSAAN` = %s"
            params.append(hasil_pemeriksaan)
    if dlpd_hitung:
        query += " AND DLPD_HITUNG = %s"
        params.append(dlpd_hitung)

    cursor.execute(query, params)
    data = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]

    cursor.close()
    conn.close()

    return jsonify({
        'columns': columns,
        'data': data
    })


# POP UP DI DASHBOARD MONITORING
@app.route('/get_detail_pelanggan_dlpd_hb')
@login_required
def get_detail_pelanggan_dlpd_hb():
    blth = request.args.get('blth')
    kdkelompok = request.args.get('kdkelompok')
    hasil_pemeriksaan = request.args.get('hasil_pemeriksaan')
    table = request.args.get('table')
    username = session.get('username', '').lower()
    dlpd_hitung = request.args.get('dlpd_hitung')

    # Validasi awal
    if not blth or not kdkelompok or not table:
        return jsonify({'error': 'Parameter wajib tidak lengkap'}), 400

    if kdkelompok not in [str(i) for i in range(1, 9)] + ['P']:
        return jsonify({'error': 'Invalid KDKELOMPOK value'}), 400

    ulp_to_billing = {
        'ulp_skw': 'billing_skw2',
        'ulp_pmk': 'billing_pmk',
        'ulp_sbs': 'billing_sbs',
        'ulp_bkg': 'billing_bkg',
        'ulp_skr': 'billing_skr',
        'ulp_sdr': 'billing_sdr',
    }

    user_access = {
        'tel_skw': ['ulp_skw'],
        'tel_pmk': ['ulp_pmk'],
        'tel_sbs': ['ulp_sbs'],
        'tel_bkg': ['ulp_bkg'],
        'tel_skr': ['ulp_skr'],
        'tel_sdr': ['ulp_sdr'],
        'tel_up3': list(ulp_to_billing.keys())
    }

    # Validasi hak akses user
    base_table = table.replace('_marking', '')
    allowed_tables = [ulp_to_billing[ulp] for ulp in user_access.get(username, [])]
    if base_table not in allowed_tables:
        return jsonify({'error': 'Akses ditolak ke tabel ini'}), 403

    # Query SQL
    conn = get_db_connection2()
    cursor = conn.cursor()

    query = f"""
        SELECT * FROM {base_table}
        WHERE BLTH = %s AND KDKELOMPOK = %s
          AND DLPD_HITUNG IS NOT NULL AND LENGTH(TRIM(DLPD_HITUNG)) > 0
          AND KET IN ('NAIK', 'TURUN', 'DIV/NA','AMAN')
    """
    params = [blth, kdkelompok]

    if hasil_pemeriksaan is not None:
        if hasil_pemeriksaan == 'BELUM DIISI':
            query += " AND (`HASIL PEMERIKSAAN` IS NULL OR `HASIL PEMERIKSAAN` = '')"
        else:
            query += " AND `HASIL PEMERIKSAAN` = %s"
            params.append(hasil_pemeriksaan)

    if dlpd_hitung:
        query += " AND DLPD_HITUNG = %s"
        params.append(dlpd_hitung)

    cursor.execute(query, params)
    data = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]

    cursor.close()
    conn.close()

    return jsonify({
        'columns': columns,
        'data': data
    })


#POP UP DI DASHBOARD MONITORING
@app.route('/get_detail_pelanggan_koreksi')
@login_required
def get_detail_pelanggan_koreksi():
    blth = request.args.get('blth')
    kdkelompok = request.args.get('kdkelompok')
    hasil_pemeriksaan = request.args.get('hasil_pemeriksaan')
    table = request.args.get('table')
    username = session.get('username', '').lower()

    # Validasi awal
    if not blth or not kdkelompok or not table:
        return jsonify({'error': 'Parameter wajib tidak lengkap'}), 400

    if kdkelompok not in [str(i) for i in range(1, 9)] + ['P']:
        return jsonify({'error': 'Invalid KDKELOMPOK value'}), 400

    ulp_to_billing = {
        'ulp_skw': 'billing_skw2',
        'ulp_pmk': 'billing_pmk',
        'ulp_sbs': 'billing_sbs',
        'ulp_bkg': 'billing_bkg',
        'ulp_skr': 'billing_skr',
        'ulp_sdr': 'billing_sdr',
    }

    user_access = {
        'tel_skw': ['ulp_skw'],
        'tel_pmk': ['ulp_pmk'],
        'tel_sbs': ['ulp_sbs'],
        'tel_bkg': ['ulp_bkg'],
        'tel_skr': ['ulp_skr'],
        'tel_sdr': ['ulp_sdr'],
        'tel_up3': list(ulp_to_billing.keys())
    }

    # Deteksi base table dan flag marking
    filter_marking = table.endswith('_marking')
    base_table = table[:-8] if filter_marking else table

    # Validasi hak akses user
    allowed_tables = [ulp_to_billing[ulp] for ulp in user_access.get(username, [])]
    if base_table not in allowed_tables:
        return jsonify({'error': 'Akses ditolak ke tabel ini'}), 403
    


    # Query SQL
    conn = get_db_connection2()
    cursor = conn.cursor()

    query = f"""
        SELECT * FROM {base_table}
        WHERE BLTH = %s AND KDKELOMPOK = %s AND MARKING_KOREKSI >0
    """
    params = [blth, kdkelompok]


    if hasil_pemeriksaan is not None:
        if hasil_pemeriksaan == 'BELUM DIISI':
            query += " AND (`HASIL PEMERIKSAAN` IS NULL OR `HASIL PEMERIKSAAN` = '')"
        else:
            query += " AND `HASIL PEMERIKSAAN` = %s"
            params.append(hasil_pemeriksaan)


    cursor.execute(query, params)
    data = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]

    cursor.close()
    conn.close()

    return jsonify({
        'columns': columns,
        'data': data
    })




@app.route('/search_idpel')
@login_required
def search_idpel():
    try:
       
        idpel = request.args.get('idpel')
        table = request.args.get('table')
        username = session.get('username', '').lower()
        ulp_to_billing = {
                'ulp_skw': 'billing_skw2',
                'ulp_pmk': 'billing_pmk',
                'ulp_sbs': 'billing_sbs',
                'ulp_bkg': 'billing_bkg',
                'ulp_skr': 'billing_skr',
                'ulp_sdr': 'billing_sdr',
            }

        user_access = {
                'tel_skw': ['ulp_skw'],
                'tel_pmk': ['ulp_pmk'],
                'tel_sbs': ['ulp_sbs'],
                'tel_bkg': ['ulp_bkg'],
                'tel_skr': ['ulp_skr'],
                'tel_sdr': ['ulp_sdr'],
                'tel_up3': list(ulp_to_billing.keys())
            }
        print(table)
        # Detect base table
        filter_marking = table.endswith('_marking')
        base_table = table[:-8] if filter_marking else table

        # Validate user access
        allowed_tables = [ulp_to_billing[ulp] for ulp in user_access.get(username, [])]
        if base_table not in allowed_tables:
            return jsonify({'error': 'Akses ditolak ke tabel ini'}), 403

        conn = get_db_connection2()
        cursor = conn.cursor()
        
        if not idpel:
            return jsonify({'error': 'IDPEL parameter is required'}), 400

        
        query = f"SELECT * FROM {table} WHERE IDPEL = %s"
        cursor.execute(query, (idpel,))

        data = cursor.fetchone()
        if not data:
            return jsonify({'error': 'Customer not found'}), 404

        return jsonify({'data': data})

    except Exception as e:
        app.logger.error(f"Error fetching customer detail: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()





@app.route('/get_detail_pelanggan_dlpd')
@login_required
def get_detail_pelanggan_dlpd():
    # Get parameters
    dlpd_hitung = request.args.get('dlpd_hitung')
    hasil_pemeriksaan = request.args.get('hasil_pemeriksaan', '')
    table = request.args.get('table')
    username = session.get('username', '').lower()

    # Validate required parameters
    if not table or not dlpd_hitung:
        return jsonify({'error': 'Parameter table dan dlpd_hitung wajib ada'}), 400

    # Validate table access
    ulp_to_billing = {
        'ulp_skw': 'billing_skw2',
        'ulp_pmk': 'billing_pmk',
        'ulp_sbs': 'billing_sbs',
        'ulp_bkg': 'billing_bkg',
        'ulp_skr': 'billing_skr',
        'ulp_sdr': 'billing_sdr',
    }

    user_access = {
        'tel_skw': ['ulp_skw'],
        'tel_pmk': ['ulp_pmk'],
        'tel_sbs': ['ulp_sbs'],
        'tel_bkg': ['ulp_bkg'],
        'tel_skr': ['ulp_skr'],
        'tel_sdr': ['ulp_sdr'],
        'tel_up3': list(ulp_to_billing.keys())
    }

    # Detect base table
    filter_marking = table.endswith('_marking')
    base_table = table[:-8] if filter_marking else table

    # Validate user access
    allowed_tables = [ulp_to_billing[ulp] for ulp in user_access.get(username, [])]
    if base_table not in allowed_tables:
        return jsonify({'error': 'Akses ditolak ke tabel ini'}), 403

    conn = get_db_connection2()
    cursor = conn.cursor()
    params = [dlpd_hitung]


    query = f"""
        SELECT * FROM {table}
        WHERE DLPD_HITUNG = %s
        AND KET IN ('NAIK', 'TURUN', 'DIV/NA', 'AMAN')
        AND KDKELOMPOK IN ('1','2','3','4','5','6','7','8','P')
    """

    # Filter hasil_pemeriksaan bila tersedia
    if hasil_pemeriksaan:
        query += " AND `HASIL PEMERIKSAAN` = %s"
        params.append(hasil_pemeriksaan)
    else:
        query += " AND (`HASIL PEMERIKSAAN` IS NULL OR `HASIL PEMERIKSAAN` = '')"

    if filter_marking:
        query += " AND DLPD_HITUNG IS NOT NULL AND LENGTH(TRIM(DLPD_HITUNG)) > 0"


    try:
        cursor.execute(query, params)
        data = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]

        return jsonify({
            'columns': columns,
            'data': data
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()


@app.route('/get_detail_pelanggan_ganda')
@login_required
def get_detail_pelanggan_ganda():
    # Get parameters
    dlpd_hitung = request.args.get('dlpd_hitung')
    hasil_pemeriksaan = request.args.get('hasil_pemeriksaan', '')
    table = request.args.get('table')
    username = session.get('username', '').lower()

    # Validate required parameters
    if not table or not dlpd_hitung:
        return jsonify({'error': 'Parameter table dan dlpd_hitung wajib ada'}), 400

    # Validate table access
    ulp_to_billing = {
        'ulp_skw': 'billing_skw2',
        'ulp_pmk': 'billing_pmk',
        'ulp_sbs': 'billing_sbs',
        'ulp_bkg': 'billing_bkg',
        'ulp_skr': 'billing_skr',
        'ulp_sdr': 'billing_sdr',
    }

    user_access = {
        'tel_skw': ['ulp_skw'],
        'tel_pmk': ['ulp_pmk'],
        'tel_sbs': ['ulp_sbs'],
        'tel_bkg': ['ulp_bkg'],
        'tel_skr': ['ulp_skr'],
        'tel_sdr': ['ulp_sdr'],
        'tel_up3': list(ulp_to_billing.keys())
    }

    # Detect base table
    filter_marking = table.endswith('_marking')
    base_table = table[:-8] if filter_marking else table

    # Validate user access
    allowed_tables = [ulp_to_billing[ulp] for ulp in user_access.get(username, [])]
    if base_table not in allowed_tables:
        return jsonify({'error': 'Akses ditolak ke tabel ini'}), 403

    conn = get_db_connection2()
    cursor = conn.cursor()
    params = [dlpd_hitung]


    query = f"""
        SELECT * FROM {table}
        WHERE DLPD_HITUNG = %s
        AND KET IN ('NAIK', 'TURUN', 'DIV/NA', 'AMAN')
        AND KDKELOMPOK IN ('A','I')
    """

    # Filter hasil_pemeriksaan bila tersedia
    if hasil_pemeriksaan:
        query += " AND `HASIL PEMERIKSAAN` = %s"
        params.append(hasil_pemeriksaan)
    else:
        query += " AND (`HASIL PEMERIKSAAN` IS NULL OR `HASIL PEMERIKSAAN` = '')"

    if filter_marking:
        query += " AND DLPD_HITUNG IS NOT NULL AND LENGTH(TRIM(DLPD_HITUNG)) > 0"


    try:
        cursor.execute(query, params)
        data = cursor.fetchall()
        columns = [desc[0] for desc in cursor.description]

        return jsonify({
            'columns': columns,
            'data': data
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        conn.close()



@app.route('/get_full_customer_detail')
@login_required
def get_full_customer_detail():
    try:
        idpel = request.args.get('idpel')
        table = request.args.get('table')
        username = session.get('username', '').lower()
        ulp_to_billing = {
                'ulp_skw': 'billing_skw2',
                'ulp_pmk': 'billing_pmk',
                'ulp_sbs': 'billing_sbs',
                'ulp_bkg': 'billing_bkg',
                'ulp_skr': 'billing_skr',
                'ulp_sdr': 'billing_sdr',
            }

        user_access = {
                'tel_skw': ['ulp_skw'],
                'tel_pmk': ['ulp_pmk'],
                'tel_sbs': ['ulp_sbs'],
                'tel_bkg': ['ulp_bkg'],
                'tel_skr': ['ulp_skr'],
                'tel_sdr': ['ulp_sdr'],
                'tel_up3': list(ulp_to_billing.keys())
            }
        print(table)
        # Detect base table
        filter_marking = table.endswith('_marking')
        base_table = table[:-8] if filter_marking else table

        # Validate user access
        allowed_tables = [ulp_to_billing[ulp] for ulp in user_access.get(username, [])]
        if base_table not in allowed_tables:
            return jsonify({'error': 'Akses ditolak ke tabel ini'}), 403

        conn = get_db_connection2()
        cursor = conn.cursor()
        
        if not idpel:
            return jsonify({'error': 'IDPEL parameter is required'}), 400

        
        query = f"SELECT * FROM {table} WHERE IDPEL = %s"
        cursor.execute(query, (idpel,))

        data = cursor.fetchone()
        if not data:
            return jsonify({'error': 'Customer not found'}), 404

        return jsonify({'data': data})

    except Exception as e:
        app.logger.error(f"Error fetching customer detail: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()



@app.route('/update_hasil_pemeriksaan', methods=['POST'])
def update_hasil_pemeriksaan():
    data = request.get_json()
    table = data.get('table')
    updates = data.get('updates', [])

    allowed_tables = [t[0] for t in tables]

    # Tangani suffix _marking jika ada
    if table.endswith('_marking'):
        base_table = table[:-8]
        if base_table not in allowed_tables:
            return jsonify({"status": "error", "message": "Tabel tidak dikenali"}), 400
        table_to_update = base_table
    else:
        if table not in allowed_tables:
            return jsonify({"status": "error", "message": "Tabel tidak dikenali"}), 400
        table_to_update = table

    conn = get_db_connection2()
    cursor = conn.cursor()

    try:
        for row in updates:
            cursor.execute(f"""
                UPDATE {table_to_update}
                SET `HASIL PEMERIKSAAN` = %s,
                    `TINDAK LANJUT` = %s,
                    `STAN_VERIFIKASI` = %s
                WHERE IDPEL = %s
            """, (row['HASIL'], row['TINDAK'], row.get('STAN', ''), row['IDPEL']))
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

    return jsonify({"status": "success"})



def update_dpm_table(df, table_name):
    """
    Inserts or replaces DPM data (BLTH, IDPEL, LWBPPAKAI) into the specified table.
    Automatically replaces existing rows with the same BLTH and IDPEL.
    Also deletes rows from exactly 6 months before the latest BLTH.
    """
    try:
        print("Connecting to database...")
        connection = pymysql.connect(
            host=db_config['host'],
            user=db_config['user'],
            password=db_config['password'],
            database='billing_v2',
            charset='utf8mb4',
            cursorclass=pymysql.cursors.DictCursor
        )
        print("Connection established.")

        with connection.cursor() as cursor:
            print(f"Checking if table '{table_name}' exists...")
            cursor.execute(f"SHOW TABLES LIKE %s", (table_name,))
            if not cursor.fetchone():
                raise ValueError(f"Table '{table_name}' does not exist")
            print("Table found.")

            print("Determining latest BLTH in uploaded data...")
            df['BLTH'] = df['BLTH'].astype(str).str.strip()
            latest_blth_str = df['BLTH'].dropna().max()
            print(f"Latest BLTH found: {latest_blth_str}")

            latest_blth_dt = datetime.strptime(latest_blth_str, "%Y%m")
            cutoff_blth_dt = latest_blth_dt - relativedelta(months=6)
            cutoff_blth = cutoff_blth_dt.strftime("%Y%m")
            print(f"Calculated cutoff BLTH to delete: {cutoff_blth}")

            print(f"Deleting rows from {table_name} where BLTH = {cutoff_blth}...")
            delete_sql = f"DELETE FROM {table_name} WHERE BLTH = %s"
            cursor.execute(delete_sql, (cutoff_blth,))
            deleted_rows = cursor.rowcount
            print(f"Deleted {deleted_rows} rows.")

            print("Preparing batch REPLACE of records...")
            batch_data = []

            for index, row in df.iterrows():
                try:
                    blth = str(row['BLTH']).strip() if pd.notna(row['BLTH']) else None
                    idpel = str(row['IDPEL']).strip() if pd.notna(row['IDPEL']) else None
                    lwbp = float(row['LWBPPAKAI']) if pd.notna(row['LWBPPAKAI']) else 0.0

                    if not blth or not idpel:
                        print(f"Skipping row {index}: Missing BLTH or IDPEL.")
                        continue

                    batch_data.append((blth, idpel, lwbp))

                except Exception as e:
                    print(f"Error processing row {index}: {str(e)}")
                    continue

            print(f"Prepared {len(batch_data)} records for REPLACE.")

            if batch_data:
                print("Replacing records (insert or update duplicates)...")
                replace_sql = f"""
                REPLACE INTO {table_name} (BLTH, idpel, LWBPPAKAI) 
                VALUES (%s, %s, %s)
                """
                cursor.executemany(replace_sql, batch_data)
                inserted = len(batch_data)
                connection.commit()
                print(f"Successfully REPLACED {inserted} records.")
            else:
                inserted = 0
                print("No valid records to insert.")

        return {
                'inserted_or_updated': inserted,
                'deleted_old_blth': cutoff_blth,
                'deleted_rows': deleted_rows,
                'total_processed': len(df)
            }

    except Exception as e:
        print(f"Database error: {str(e)}")
        return None


def process_dpm(df):
    """Process DPM dataframe to ensure required columns exist and are properly formatted"""
    required_columns = ['BLTH', 'IDPEL', 'LWBPPAKAI']
    
    # Check if all required columns exist
    missing_cols = [col for col in required_columns if col not in df.columns]
    if missing_cols:
        raise ValueError(f"Kolom yang diperlukan tidak ditemukan: {', '.join(missing_cols)}")
    
    # Convert BLTH to string (format: YYYYMM)
    df['BLTH'] = df['BLTH'].astype(str).str.strip()
    
    # Ensure IDPEL is string
    df['IDPEL'] = df['IDPEL'].astype(str).str.strip()
    
    # Convert LWBPPAKAI to numeric
    df['LWBPPAKAI'] = pd.to_numeric(df['LWBPPAKAI'], errors='coerce')
    
    return df[required_columns]

#UPDATE NO METER
def update_mysql_table(df, table_name):
    print(f"Attempting to update {table_name}")
    print(f"Data to insert:\n{df.head()}")

    try:
        connection = pymysql.connect(
            host=db_config['host'],
            user=db_config['user'],
            password=db_config['password'],
            database='billing_v2',
            charset='utf8mb4',
            cursorclass=pymysql.cursors.DictCursor
        )

        with connection.cursor() as cursor:
            # Check if table exists
            cursor.execute(f"SHOW TABLES LIKE %s", (table_name,))
            if not cursor.fetchone():
                raise Exception(f"Table '{table_name}' doesn't exist in database 'billing_v2'")

            updated_count = 0
            for index, row in df.iterrows():
                idpel = str(row['idpel']) if pd.notna(row['idpel']) else None
                no_meter = str(row['no_meter']) if pd.notna(row['no_meter']) else None

                if idpel and no_meter:
                    # First try to update existing record
                    sql_update = f"""
                    UPDATE {table_name} 
                    SET no_meter = %s, created_at = CURRENT_TIMESTAMP 
                    WHERE idpel = %s
                    """
                    cursor.execute(sql_update, (no_meter, idpel))
                    
                    # If no rows were updated, insert new record
                    if cursor.rowcount == 0:
                        sql_insert = f"""
                        INSERT INTO {table_name} (idpel, no_meter) 
                        VALUES (%s, %s)
                        """
                        cursor.execute(sql_insert, (idpel, no_meter))
                    
                    updated_count += 1

            connection.commit()

            # if updated_count > 0:
            #     flash(f"Sukses menyimpan {updated_count} data ke tabel {table_name}.", 'success')
            # else:
            #     flash("Tidak ada data yang disimpan.", 'warning')

            return {
                'success_count': updated_count,
                'duplicate_count': 0,
                'total_saved': updated_count
            }

    except Exception as e:
        print(f"Database error: {str(e)}")
        flash(f"Database error: {str(e)}", 'danger')
        if 'connection' in locals() and connection:
            connection.rollback()
        raise

    finally:
        if 'connection' in locals() and connection:
            connection.close()



# Define mappings at module level (outside the route function)
ULP_MAPPING = {
    'file_nomet_singkawang': 'nomet_skw',
    'file_nomet_pemangkat': 'nomet_pmk',
    'file_nomet_sambas': 'nomet_sbs',
    'file_nomet_bengkayang': 'nomet_bkg',
    'file_nomet_sekura': 'nomet_skr',
    'file_nomet_seiduri': 'nomet_sdr'
}

DPM_MAPPING = {
    'file_dpm_skw': 'dpm_skw',
    'file_dpm_pmk': 'dpm_pmk',
    'file_dpm_sbs': 'dpm_sbs',
    'file_dpm_bkg': 'dpm_bkg',
    'file_dpm_skr': 'dpm_skr',
    'file_dpm_sdr': 'dpm_sdr'
}

@app.route("/admin_v2", methods=["GET", "POST"])
@login_required
def admin_v2():
    if request.method == "POST":
        # Get tables user can access
        accessible_tables = session.get('tables', [])
        processed_files = False

        # Process No Meter uploads
        for file_field, table_name in ULP_MAPPING.items():
            if file_field in request.files:
                file = request.files[file_field]
                
                if file.filename == '':
                    continue
                    
                if table_name not in accessible_tables:
                    flash(f"Anda tidak memiliki akses ke tabel {table_name}", 'warning')
                    continue

                if not allowed_file(file.filename):
                    flash("Format file tidak valid. Hanya file Excel (.xlsx, .xls) yang diperbolehkan", 'error')
                    continue

                try:
                    filename = secure_filename(file.filename)
                    temp_path = os.path.join('temp', filename)
                    os.makedirs('temp', exist_ok=True)
                    file.save(temp_path)
                    
                    # Read Excel with engine specification
                    try:
                        df = pd.read_excel(temp_path, engine='openpyxl')
                    except:
                        df = pd.read_excel(temp_path, engine='xlrd')
                    #UPLOAD NO METER
                    processed_df = process_nomet(df)
                    result = update_mysql_table(processed_df, table_name)
                    
                    if result['success_count'] > 0:
                        flash(f"Sukses update {result['success_count']} data No Meter ke {table_name}", 'success')
                    if result['duplicate_count'] > 0:
                        flash(f"Ditemukan {result['duplicate_count']} data duplikat di {table_name}", 'warning')
                        
                    processed_files = True
                except Exception as e:
                    flash(f"Gagal memproses file {file_field}: {str(e)}", 'error')
                finally:
                    if os.path.exists(temp_path):
                        os.remove(temp_path)

        # Process DPM uploads
        for file_field, table_name in DPM_MAPPING.items():
            if file_field in request.files:
                file = request.files[file_field]
                
                if file.filename == '':
                    continue
                    
                if table_name not in accessible_tables:
                    flash(f"DPM: Anda tidak memiliki akses ke tabel {table_name}", 'warning')
                    continue

                if not allowed_file(file.filename):
                    flash("DPM: Format file tidak valid. Hanya file Excel (.xlsx, .xls) yang diperbolehkan", 'error')
                    continue

                try:
                    filename = secure_filename(file.filename)
                    temp_path = os.path.join('temp', filename)
                    os.makedirs('temp', exist_ok=True)
                    file.save(temp_path)
                    
                    # Read Excel with engine specification
                    try:
                        df = pd.read_excel(temp_path, engine='openpyxl')
                    except:
                        df = pd.read_excel(temp_path, engine='xlrd')
                    #UPDATE DPM
                    processed_df = process_dpm(df)
                    result = update_dpm_table(processed_df, table_name)

                    if result['inserted_or_updated'] > 0:
                        flash(f"DPM: Sukses upload {result['inserted_or_updated']} data ke {table_name}", 'success')

                        
                    processed_files = True
                except Exception as e:
                    flash(f"DPM: Gagal memproses file {file_field}: {str(e)}", 'error')
                finally:
                    if os.path.exists(temp_path):
                        os.remove(temp_path)

        if not processed_files:
            flash("Tidak ada file yang diproses", 'info')

        return redirect(url_for('admin_v2'))

        # GET method - Fetch last upload time and BLTH information for each table
    else:
        last_uploads = {}
        dpm_blth_info = {}  # To store BLTH information from DPM tables
        
        try:
            conn = pymysql.connect(
                host=db_config['host'],
                user=db_config['user'],
                password=db_config['password'],
                db='billing_v2'
            )
            cursor = conn.cursor()

            # Check both No Meter and DPM tables
            tables_to_check = list(ULP_MAPPING.values()) + list(DPM_MAPPING.values())
            
            for table in tables_to_check:
                try:
                    # Get last upload time
                    cursor.execute(f"SELECT created_at FROM {table} ORDER BY created_at DESC LIMIT 1")
                    result = cursor.fetchone()
                    last_uploads[table] = result[0].strftime('%d-%m-%Y pukul %H:%M') if result else "Belum ada data"

                    # For DPM tables, also get unique BLTH values
                    if table.startswith('dpm_'):
                        cursor.execute(f"SELECT DISTINCT BLTH FROM {table} ORDER BY BLTH DESC LIMIT 12")
                        blth_results = cursor.fetchall()
                        dpm_blth_info[table] = [row[0] for row in blth_results]

                except Exception as e:
                    print(f"Error checking table {table}: {e}")
                    last_uploads[table] = "Tidak dapat diakses"
                    if table.startswith('dpm_'):
                        dpm_blth_info[table] = []

            cursor.close()
            conn.close()
        except Exception as e:
            print(f"Database error: {e}")
            flash(f"Error koneksi database: {str(e)}", 'danger')
            # Initialize empty BLTH info if connection fails
            for table in DPM_MAPPING.values():
                dpm_blth_info[table] = []

    return render_template(
        'admin_v2.html',
        name=session['username'],
        last_uploads=last_uploads,
        dpm_blth_info=dpm_blth_info,
        accessible_tables=session.get('tables', [])
    )



@app.route('/admin')

@login_required

def admin():

    return render_template('admin.html', name=session['username'])



# Halaman Logout

@app.route('/logout')

def logout():

    session.pop('loggedin', None)

    session.pop('username', None)

    flash('Anda telah logout.', 'success')

    return redirect(url_for('login'))

@app.route('/logout_pksr')

def logout_pksr():

    session.pop('loggedin', None)

    session.pop('username', None)

    flash('Anda telah logout.', 'success')

    return redirect(url_for('login_pksr'))


def allowed_file(filename):
    """Check if the file has an allowed extension"""
    allowed_extensions = {'xls', 'xlsx'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

def process_nomet(df):
    # Standardize column names and clean data
    df = df.rename(columns={
        'IDPEL': 'idpel',
        'NOMORKWH': 'no_meter',
        'NOMOR KWH': 'no_meter',
        'NO METER': 'no_meter'
    })
    
    # Select only the columns we need
    return df[['idpel', 'no_meter']]

@app.route('/grafik/<idpel>')
def show_graph(idpel):
    blth = request.args.get('blth')
    username = session.get('username', '').lower()
    ulp_param = request.args.get("ulp", "").lower()

    # Mapping ULP ke Tabel
    dpm_to_billing = {
        'ulp_skw': 'dpm_skw',
        'ulp_pmk': 'dpm_pmk',
        'ulp_sbs': 'dpm_sbs',
        'ulp_bkg': 'dpm_bkg',
        'ulp_skr': 'dpm_skr',
        'ulp_sdr': 'dpm_sdr',
    }

    user_access = {
        'tel_skw': ['ulp_skw'],
        'tel_pmk': ['ulp_pmk'],
        'tel_sbs': ['ulp_sbs'],
        'tel_bkg': ['ulp_bkg'],
        'tel_skr': ['ulp_skr'],
        'tel_sdr': ['ulp_sdr'],
        'tel_up3': list(dpm_to_billing.keys())
    }

    allowed_ulp = user_access.get(username)
    if not allowed_ulp:
        return jsonify({"error": "Akun tidak memiliki akses ULP"}), 403

   
    ulp_key = ulp_param if ulp_param.startswith("ulp_") else f"ulp_{ulp_param}"
  
    dpm_table = dpm_to_billing.get(ulp_key)
    if not dpm_table:
        return jsonify({"error": "Tabel tidak ditemukan untuk ULP ini"}), 403


    query = f"""
    SELECT 
        BLTH, + 
        LWBPPAKAI,
        DATE_FORMAT(STR_TO_DATE(CONCAT(BLTH, '01'), '%Y%m%d'), '%b-%Y') AS BLTH_FORMATTED
    FROM {dpm_table}
    WHERE 
        IDPEL = :idpel AND
        BLTH <= :blth
    ORDER BY BLTH DESC
    LIMIT 7
    """

    try:
        usage_data = pd.read_sql(text(query), engine, params={'idpel': idpel, 'blth': blth})
        
        if usage_data.empty:
            return "<h3>Tidak ada data grafik untuk IDPEL ini.</h3>"

        plt.figure(figsize=(7, 4))
        plt.plot(usage_data['BLTH_FORMATTED'], usage_data['LWBPPAKAI'], 'bo-')
        plt.title(f"Grafik Pemakaian IDPEL: {idpel}")
        plt.xlabel("BLTH")
        plt.ylabel("kWh")
        plt.grid(True)

        # Tambahkan label angka di setiap titik
        for x, y in zip(usage_data['BLTH_FORMATTED'], usage_data['LWBPPAKAI']):
            plt.text(x, y + (max(usage_data['LWBPPAKAI'])*0.02), f"{y:.0f}", ha='center', fontsize=12)

        buf = io.BytesIO()
        plt.savefig(buf, format='png', bbox_inches='tight')
        buf.seek(0)
        plt.close()


        return send_file(buf, mimetype='image/png')

    except Exception as e:
        print("[ERROR] Grafik gagal dibuat:", e)
        return f"<h3>❌ Error menampilkan grafik: {e}</h3>", 500

# ALTER TABLE billing_skw2 
# ADD COLUMN KDKELOMPOK VARCHAR(20) AFTER JAMNYALA600;
def copy_dataframe2(
    lalu3, lalulalu, lalu, akhir,
    blth_lalulalu, blth_lalu, blth_kini, blth_lalu3,
    sortir_naik, sortir_turun, ulp_param, dpm_table):
    
    def safe_float_column(df, col):
        return df.get(col, pd.Series(dtype='float64'))

    def prepare_dataframe(raw, keys):
        return pd.DataFrame({
            k: raw.get(k, pd.Series(dtype='object' if k in ['BLTH', 'IDPEL'] else 'float64'))
            for k in keys
        })

    try:
        juruslalu3 = prepare_dataframe(lalu3, ['BLTH', 'IDPEL', 'LWBPPAKAI'])
        juruslalulalu = prepare_dataframe(lalulalu, ['BLTH', 'IDPEL', 'LWBPPAKAI'])
        juruslalu = prepare_dataframe(lalu, ['BLTH', 'IDPEL', 'LWBPPAKAI'])
        jurusakhir = prepare_dataframe(akhir, [
            'BLTH', 'IDPEL', 'NAMA', 'TARIF', 'DAYA', 'KDKELOMPOK', 'SLALWBP', 'LWBPCABUT',
            'LWBPPASANG', 'SAHLWBP', 'DLPD','LWBPPAKAI'
        ])
          
        print("---- SEBELUM MERGE (copy_dataframe_v2) ----")
        print("jurus lalu3 :", juruslalu3.shape)
        print("Juruslalulalu:", juruslalulalu.shape)
        print("Juruslalu:", juruslalu.shape)
        print("Jurusakhir:", jurusakhir.shape)

        duplikat_lalu3 = juruslalu3['IDPEL'].duplicated().sum()
        duplikat_lalulalu = juruslalulalu['IDPEL'].duplicated().sum()
        duplikat_lalu = juruslalu['IDPEL'].duplicated().sum()
        duplikat_akhir = jurusakhir['IDPEL'].duplicated().sum()

        print("Duplikat di juruslalu3:", duplikat_lalu3)
        print("Duplikat di juruslalulalu:", duplikat_lalulalu)
        print("Duplikat di juruslalu:", duplikat_lalu)
        print("Duplikat di jurusakhir:", duplikat_akhir)

        if duplikat_lalu3 > 0:
            flash(f"⚠️ Ditemukan {duplikat_lalu3} IDPEL duplikat di DPM N - 3.", "warning")
        if duplikat_lalulalu > 0:
            flash(f"⚠️ Ditemukan {duplikat_lalulalu} IDPEL duplikat di DPM N - 2.", "warning")
        if duplikat_lalu > 0:
            flash(f"⚠️ Ditemukan {duplikat_lalu} IDPEL duplikat di DPM N - 1.", "warning")
        if duplikat_akhir > 0:
            flash(f"⚠️ Ditemukan {duplikat_akhir} IDPEL duplikat di DPM bulan N.", "warning")


        # Pastikan semua IDPEL sudah distandarisasi
        for df in [juruslalu3, juruslalulalu,juruslalu, jurusakhir]:
            df['IDPEL'] = df['IDPEL'].astype(str).str.strip().str.lower()

        # Rename dulu kolom LWBPPAKAI di juruslalulalu dan juruslalu agar tidak bentrok
        lalu3_renamed = juruslalu3.rename(columns={'LWBPPAKAI': 'LWBPPAKAI_z'})
        lalulalu_renamed = juruslalulalu.rename(columns={'LWBPPAKAI': 'LWBPPAKAI_x'})
        lalu_renamed     = juruslalu.rename(columns={'LWBPPAKAI': 'LWBPPAKAI_y'})

         # Merge perbaikan: pastikan nama kolom tidak bentrok
        kroscek_temp = (
            jurusakhir
            .merge(lalu_renamed[['IDPEL', 'LWBPPAKAI_y']], on='IDPEL', how='left')
            .merge(lalulalu_renamed[['IDPEL', 'LWBPPAKAI_x']], on='IDPEL', how='left')
            .merge(lalu3_renamed[['IDPEL', 'LWBPPAKAI_z']], on='IDPEL', how='left')
        )
        print("Duplikat setelah merge :", kroscek_temp['IDPEL'].duplicated().sum())

        # Simpan kondisi awal
        lwbp_kosong = kroscek_temp['LWBPPAKAI'].isna()

        # Hitung berapa yang kosong sebelum dihitung ulang
        count_replaced = lwbp_kosong.sum()
        print(f"Jumlah data LWBPPAKAI yang dihitung ulang: {count_replaced}")

        # Isi nilai hanya kalau kosong
        kroscek_temp['LWBPPAKAI'] = np.where(
            lwbp_kosong,
            (kroscek_temp['LWBPCABUT'].fillna(0)
            - kroscek_temp['SLALWBP'].fillna(0)
            + kroscek_temp['SAHLWBP'].fillna(0)
            - kroscek_temp['LWBPPASANG'].fillna(0)),
            kroscek_temp['LWBPPAKAI']
        )

        print(f"[DEBUG] Step 1 - Original rows in kroscek_temp: {len(kroscek_temp)}")

        delta = kroscek_temp['LWBPPAKAI'] - kroscek_temp['LWBPPAKAI_y'].fillna(0)
        rerata = (
            kroscek_temp[['LWBPPAKAI_y', 'LWBPPAKAI_x', 'LWBPPAKAI_z']]
            .fillna(0)
            .mean(axis=1)
            )
        print("[DEBUG] Cek NaN di LWBPPAKAI:", kroscek_temp['LWBPPAKAI'].isna().sum())
        print("[DEBUG] Cek NaN di rerata:", rerata.isna().sum())

        
        kroscek_temp['DLPD_3BLN'] = np.where(
            (kroscek_temp['LWBPPAKAI'].fillna(0) > 1.5 * rerata.fillna(0)),
            'Naik50% R3BLN',
            'Turun=50% R3BLN'
            )
        with np.errstate(divide='ignore', invalid='ignore'):
            percentage = (delta / kroscek_temp['LWBPPAKAI_y'].replace(0, np.nan)) * 100
            percentage = np.nan_to_num(percentage, nan=0)
        
        daya_kw = kroscek_temp['DAYA'] / 1000
        jam_nyala = (kroscek_temp['LWBPPAKAI'] / daya_kw).replace([np.inf, -np.inf], 0).fillna(0)
        
        # Kondisi stan mundur
        stan_mundur_condition = (
            (kroscek_temp['SAHLWBP'] < kroscek_temp['SLALWBP']) &
            (kroscek_temp['LWBPCABUT'].fillna(0) == 0) &
            (kroscek_temp['LWBPPASANG'].fillna(0) == 0)
        )

        # Kondisi cek pecahan (ada nilai cabut atau pasang)
        cek_pecahan_condition = (
            (kroscek_temp['LWBPCABUT'].fillna(0) != 0) |
            (kroscek_temp['LWBPPASANG'].fillna(0) != 0)
        )

        sortir_naik = 40
        sortir_turun = 40
        is_na = kroscek_temp['LWBPPAKAI_y'].isna() | (kroscek_temp['LWBPPAKAI_y'] == 0)
        is_naik = (~is_na) & (percentage >= sortir_naik)
        is_turun = (~is_na) & (percentage <= -sortir_turun)
        
        # Daftar kondisi dan klasifikasi
        conditions = [
            (jam_nyala >= 720),
            cek_pecahan_condition,
            stan_mundur_condition,
            (percentage > 50),
            (is_na & (jam_nyala > 40)), 
            (percentage < -50),
            # (delta == 0),
            (kroscek_temp['LWBPPAKAI'] == 0),
            (jam_nyala > 0) & (jam_nyala < 40)
            
        ]

        choices = [
            'JN_720up',
            'Cek_Pecahan',
            'Stan_Mundur',
            'Naik_50%Up',
            'Cek_DIV/NA',
            'Turun_50%Down',
            'kWh_Nol',
            'JN_40Down'
        ]
        # Terapkan ke kolom baru
        kroscek_temp['DLPD_HITUNG'] = np.select(conditions, choices, default='')
 

        # # processed_df = process_dpm(kroscek_temp)
        # processed_df = process_dpm(kroscek_temp[['BLTH', 'IDPEL', 'LWBPPAKAI']])
        # #UPDATE DPM
        # processed_df = process_dpm(df)
        # result = update_dpm_table(processed_df, dpm_table)

        # if result['inserted_or_updated'] > 0:
        #     flash(f"DPM: Sukses upload {result['inserted_or_updated']} data ke {dpm_table}", 'success')

        processed_df = process_dpm(kroscek_temp.rename(columns={'KWH SEKARANG': 'LWBPPAKAI'})[['BLTH', 'IDPEL', 'LWBPPAKAI']])
        result = update_dpm_table(processed_df, dpm_table)


        if result['inserted_or_updated'] > 0:
            flash(f"DPM: Sukses upload {result['inserted_or_updated']} data ke {dpm_table}", 'success')
        
        
        kroscek = pd.DataFrame({
            'BLTH': blth_kini,
            'IDPEL': kroscek_temp['IDPEL'].astype(str).str.strip().str.lower(),
            'NAMA': kroscek_temp['NAMA'],
            'TARIF': kroscek_temp['TARIF'],
            'DAYA': kroscek_temp['DAYA'].fillna(0).astype(int),
            'SLALWBP': kroscek_temp['SLALWBP'].fillna(0).astype(int),
            'LWBPCABUT': kroscek_temp['LWBPCABUT'].fillna(0).astype(int),
            'SELISIH STAN BONGKAR': (kroscek_temp['SLALWBP'].fillna(0) - kroscek_temp['LWBPCABUT'].fillna(0)).astype(int),
            'LWBP PASANG': kroscek_temp['LWBPPASANG'].fillna(0).astype(int),
            'KWH SEKARANG': kroscek_temp['LWBPPAKAI'].fillna(0).astype(int),
            'KWH 1 BULAN LALU': kroscek_temp['LWBPPAKAI_y'].fillna(0).astype(int),
            'KWH 2 BULAN LALU': kroscek_temp['LWBPPAKAI_x'].fillna(0).astype(int),
            'DELTA PEMKWH': delta.fillna(0).astype(int),
            'DLPD_HITUNG':kroscek_temp['DLPD_HITUNG'],
            'DLPD_3BLN': kroscek_temp['DLPD_3BLN']  # Add this line
        })  
        print(f"[DEBUG] Step 2 - After DataFrame creation: {len(kroscek)} rows")

       

        kroscek['%'] = pd.Series(percentage).round(1).astype(str) + '%'

        
        kroscek['KET'] = np.select([is_na, is_naik, is_turun], ['DIV/NA', 'NAIK', 'TURUN'], default='AMAN')
        print(f"[DEBUG] Step 3 - After classification: {len(kroscek)} rows")

        kroscek['DLPD'] = kroscek_temp['DLPD'].fillna('')
        print("[DEBUG] Head setelah klasifikasi:")
        print(
            kroscek_temp[
                ['IDPEL', 'LWBPPAKAI', 'LWBPPAKAI_y', 'LWBPPAKAI_x', 'LWBPPAKAI_z','DLPD_3BLN']
            ].head(10)
        )

        kroscek['SAHLWBP'] = kroscek_temp['SAHLWBP'].fillna(0).astype(int)
        
        kroscek['JAM NYALA'] = jam_nyala.round(1)
        kroscek['JAMNYALA600'] = np.select([jam_nyala > 600, jam_nyala <= 600], ['600Up', '600Down'], default='UNKNOWN')
        print(f"[DEBUG] Step 4 - After JAM NYALA calculation: {len(kroscek)} rows")
        
        # kroscek['GRAFIK'] = kroscek['IDPEL'].apply(lambda x: 
        #     f'<a href="/grafik/{x}?blth={blth_kini}" target="popup" onclick="window.open(\'/grafik/{x}?blth={blth_kini}\', \'popup\', \'width=800,height=600\'); return false;">LIHAT GRAFIK</a>'
        # )
        kroscek['GRAFIK'] = kroscek['IDPEL'].apply(
        lambda idpel: f'<a href="https://telup3singkawang.com/grafik/{idpel}?blth={blth_kini}&ulp={ulp_param.lower()}" '
                  f'target="popup" onclick="window.open(\'/grafik/{idpel}?blth={blth_kini}&ulp={ulp_param.lower()}\', \'popup\', \'width=800,height=600\'); return false;">LIHAT GRAFIK</a>'
        )

      
        df['IDPEL'] = df['IDPEL'].apply(normalize_idpel)
        #jaga2 dupliaksi
        kroscek = kroscek.drop_duplicates(subset='IDPEL', keep='first')

        kroscek['NOMET'] = ""
        
      # URL dasar untuk foto
        path_foto1 = 'https://portalapp.iconpln.co.id/acmt/DisplayBlobServlet1?idpel='
        path_foto2 = '&blth='

        # Link per bulan
        kroscek['FOTO AKHIR'] = kroscek['IDPEL'].apply(lambda x: f'<a href="{path_foto1}{x}{path_foto2}{blth_kini}" target="popup" onclick="window.open(\'{path_foto1}{x}{path_foto2}{blth_kini}\', \'popup\', \'width=700,height=700,scrollbars=no,toolbar=no\'); return false;">LINK FOTO</a>')
        kroscek['FOTO LALU'] = kroscek['IDPEL'].apply(lambda x: f'<a href="{path_foto1}{x}{path_foto2}{blth_lalu}" target="popup" onclick="window.open(\'{path_foto1}{x}{path_foto2}{blth_lalu}\', \'popup\', \'width=700,height=700,scrollbars=no,toolbar=no\'); return false;">LINK FOTO</a>')
        kroscek['FOTO LALU2'] = kroscek['IDPEL'].apply(lambda x: f'<a href="{path_foto1}{x}{path_foto2}{blth_lalulalu}" target="popup" onclick="window.open(\'{path_foto1}{x}{path_foto2}{blth_lalulalu}\', \'popup\', \'width=700,height=700,scrollbars=no,toolbar=no\'); return false;">LINK FOTO</a>')

        # Link 3 foto sekaligus, pakai 5 digit terakhir IDPEL sebagai label link
        kroscek['FOTO 3BLN'] = kroscek['IDPEL'].apply(lambda x: f'<a href="#" onclick="open3Foto(\'{x}\', \'{blth_kini}\'); return false;">{str(x)[-5:]}</a>')



        kroscek['HASIL PEMERIKSAAN'] = ""
        kroscek['TINDAK LANJUT'] = ""
        kroscek['KDKELOMPOK'] = kroscek_temp['KDKELOMPOK'].fillna('')
        print(f"[DEBUG] Step 7 - Final row count in kroscek: {len(kroscek)} rows")

        return kroscek
    except Exception as e:
        print(f"Error in copy_dataframe2: {str(e)}")
        return pd.DataFrame(columns=[
            'BLTH', 'IDPEL', 'NAMA', 'TARIF', 'DAYA', 'KDKELOMPOK', 'SLALWBP', 'LWBPCABUT',
            'SELISIH STAN BONGKAR', 'LWBP PASANG', 'KWH SEKARANG', 'KWH 1 BULAN LALU',
            'KWH 2 BULAN LALU', 'DELTA PEMKWH', '%', 'KET', 'JAM NYALA', 'DLPD', 'DLPD_3BLN', 'DLPD_HITUNG',
            'SAHLWBP', 'FOTO AKHIR', 'FOTO LALU', 'FOTO LALU2','FOTO 3BLN', 'HASIL PEMERIKSAAN',
            'TINDAK LANJUT', 'JAMNYALA600', 'NOMET', 'GRAFIK'
        ])


def copy_dataframe_update(
    data_lalu, data_akhir, blth_kini,
    sortir_naik, sortir_turun, ulp_param, dpm_table, engine,billing_table):
    
    def prepare_dataframe(data, keys):
        """Handle both DataFrame and file inputs"""
        if isinstance(data, pd.DataFrame):
            df = data.copy()
        else:  # Assume it's a file object
            try:
                data.seek(0)  # Rewind file pointer if needed
                df = pd.read_excel(data)
            except Exception as e:
                print(f"Error reading data: {str(e)}")
                return pd.DataFrame(columns=keys)
        
        # Ensure all requested columns exist
        for col in keys:
            if col not in df.columns:
                if col in ['BLTH', 'IDPEL']:
                    df[col] = pd.Series(dtype='object')
                else:
                    df[col] = pd.Series(dtype='float64')
        return df[keys]

    try:
         # Validate ulp_param
        if not ulp_param or not isinstance(ulp_param, str):
            raise ValueError("ULP parameter is missing or invalid")
        
        # billing_table = f"billing_{ulp_param.strip().lower()}"
        print(f"[DEBUG] Using billing table: {billing_table}")

        # Read data from either DataFrames or files
        juruslalu = prepare_dataframe(data_lalu, ['BLTH', 'IDPEL', 'LWBPPAKAI'])
        jurusakhir = prepare_dataframe(data_akhir, [
            'BLTH', 'IDPEL', 'LWBPCABUT',
            'LWBPPASANG', 'SAHLWBP'
        ])
        
        # Normalize IDPEL
        for df in [juruslalu, jurusakhir]:
            df['IDPEL'] = df['IDPEL'].astype(str).str.strip().str.zfill(12)
        

        query = f"SELECT IDPEL, SLALWBP, DAYA FROM {billing_table} WHERE BLTH = '{blth_kini}'"

        update_jurusakhir = pd.read_sql(query, engine)

        update_jurusakhir['IDPEL'] = update_jurusakhir['IDPEL'].astype(str).str.strip().str.zfill(12)
        
        # Merge with current data
        jurusakhir = jurusakhir.merge(
            update_jurusakhir[['IDPEL', 'SLALWBP','DAYA']], 
            on='IDPEL', 
            how='left'
        )
        
        # Recalculate LWBPPAKAI
        jurusakhir['LWBPPAKAI'] = (
            jurusakhir['LWBPCABUT'].fillna(0) 
            - jurusakhir['SLALWBP'].fillna(0) 
            + jurusakhir['SAHLWBP'].fillna(0) 
            - jurusakhir['LWBPPASANG'].fillna(0)
        )
        
        # Merge with previous data
        lalu_renamed = juruslalu.rename(columns={'LWBPPAKAI': 'LWBPPAKAI_y'})
        kroscek_temp = jurusakhir.merge(
            lalu_renamed[['IDPEL', 'LWBPPAKAI_y']], 
            on='IDPEL', 
            how='left'
        )
        
        # Hitung delta dan persentase
        delta = kroscek_temp['LWBPPAKAI'] - kroscek_temp['LWBPPAKAI_y'].fillna(0)
        with np.errstate(divide='ignore', invalid='ignore'):
            percentage = (delta / kroscek_temp['LWBPPAKAI_y'].replace(0, np.nan)) * 100
            percentage = np.nan_to_num(percentage, nan=0)
        
        daya_kw = kroscek_temp['DAYA'] / 1000
        jam_nyala = (kroscek_temp['LWBPPAKAI'] / daya_kw).replace([np.inf, -np.inf], 0).fillna(0)
        
        # Kondisi stan mundur
        stan_mundur_condition = (
            (kroscek_temp['SAHLWBP'] < kroscek_temp['SLALWBP']) &
            (kroscek_temp['LWBPCABUT'].fillna(0) == 0) &
            (kroscek_temp['LWBPPASANG'].fillna(0) == 0)
        )

        # Kondisi cek pecahan (ada nilai cabut atau pasang)
        cek_pecahan_condition = (
            (kroscek_temp['LWBPCABUT'].fillna(0) != 0) |
            (kroscek_temp['LWBPPASANG'].fillna(0) != 0)
        )

        sortir_naik = 40
        sortir_turun = 40
        is_na = kroscek_temp['LWBPPAKAI_y'].isna() | (kroscek_temp['LWBPPAKAI_y'] == 0)
        is_naik = (~is_na) & (percentage >= sortir_naik)
        is_turun = (~is_na) & (percentage <= -sortir_turun)
        
        # Daftar kondisi dan klasifikasi
        conditions = [
            (jam_nyala >= 720),
            cek_pecahan_condition,
            stan_mundur_condition,
            (percentage > 50),
            (is_na & (jam_nyala > 40)), 
            (percentage < -50),
            # (delta == 0),
            (kroscek_temp['LWBPPAKAI'] == 0),
            (jam_nyala > 0) & (jam_nyala < 40)
            
        ]

        choices = [
            'JN_720up',
            'Cek_Pecahan',
            'Stan_Mundur',
            'Naik_50%Up',
            'Cek_DIV/NA',
            'Turun_50%Down',
            'kWh_Nol',
            'JN_40Down'
        ]
        # Terapkan ke kolom baru
        kroscek_temp['DLPD_HITUNG'] = np.select(conditions, choices, default='')
        
       
        # Buat DataFrame hasil
        kroscek = pd.DataFrame({
            'BLTH': blth_kini,
            'IDPEL': kroscek_temp['IDPEL'].astype(str).str.strip().str.lower(),
            'LWBPCABUT': kroscek_temp['LWBPCABUT'].fillna(0).astype(int),
            'SELISIH STAN BONGKAR': (kroscek_temp['SLALWBP'].fillna(0) - kroscek_temp['LWBPCABUT'].fillna(0)).astype(int),
            'LWBP PASANG': kroscek_temp['LWBPPASANG'].fillna(0).astype(int),
            'KWH SEKARANG': kroscek_temp['LWBPPAKAI'].fillna(0).astype(int),
            'DELTA PEMKWH': delta.fillna(0).astype(int),
            '%': pd.Series(percentage).round(1).astype(str) + '%',
            'KET': np.select([is_na, is_naik, is_turun], ['DIV/NA', 'NAIK', 'TURUN'], default='AMAN'),
            'DLPD_HITUNG': kroscek_temp['DLPD_HITUNG'],
            'SAHLWBP': kroscek_temp['SAHLWBP'].fillna(0).astype(int),
            'JAM NYALA': jam_nyala.round(1),
            'JAMNYALA600': np.where(jam_nyala > 600, '600Up', '600Down')
        })
        
        # Ambil MARKING_KOREKSI lama dari DB
        idpel_list = kroscek['IDPEL'].astype(str).tolist()
        format_ids = ','.join([f"'{idpel}'" for idpel in idpel_list])
        query = f"SELECT IDPEL, MARKING_KOREKSI FROM {billing_table} WHERE IDPEL IN ({format_ids})"
        df_marking = pd.read_sql(query, engine)

        # Buat dict mapping IDPEL ke nilai lama
        marking_dict = dict(zip(df_marking['IDPEL'], df_marking['MARKING_KOREKSI']))

        # Tambahkan kolom MARKING_KOREKSI ke hasil kroscek
        def add_marking(idpel):
            return marking_dict.get(idpel, 0) + 1

        kroscek['MARKING_KOREKSI'] = kroscek['IDPEL'].apply(add_marking)
         # processed_df = process_dpm(kroscek)
        # Proses DPM
        # Proses DPM
        processed_df = process_dpm(kroscek.rename(columns={'KWH SEKARANG': 'LWBPPAKAI'})[['BLTH', 'IDPEL', 'LWBPPAKAI']])
        result = update_dpm_table(processed_df, dpm_table)


        if result['inserted_or_updated'] > 0:
            flash(f"DPM: Sukses upload {result['inserted_or_updated']} data ke {dpm_table}", 'success')


        return kroscek

    except Exception as e:
        print(f"Error in copy_dataframe_update: {str(e)}")
        return pd.DataFrame(columns=[
            'IDPEL','LWBPCABUT', 'SELISIH STAN BONGKAR', 'LWBP PASANG', 'KWH SEKARANG', 
            'DELTA PEMKWH', '%', 'KET', 'JAM NYALA', 'DLPD_HITUNG',
            'SAHLWBP', 'JAMNYALA600'
        ])

############################################################





def save_to_database_v2(df, table_name, engine,
                       batch_size=100,
                       max_retries=3,
                       retry_delay=2):
    try:
        print(f"📥 Jumlah baris awal input: {len(df)}")

        # 1. Normalisasi IDPEL dan BLTH, hapus duplikat berdasarkan keduanya
        df['IDPEL'] = df['IDPEL'].astype(str).str.strip().str.zfill(12)
        df['BLTH'] = df['BLTH'].astype(str).str.strip()
        
        # Drop duplicates berdasarkan kombinasi IDPEL + BLTH
        df = df.drop_duplicates(subset=['IDPEL', 'BLTH'], keep='last')
        print(f"🧹 Setelah normalisasi & drop duplikat berdasarkan IDPEL+BLTH: {len(df)} baris")

        # 2. Ambil kombinasi IDPEL+BLTH yang sudah ada di DB
        existing_keys = set()
        chunk_size = 10000
        with engine.connect() as conn:
            result = conn.execution_options(stream_results=True).execute(
                text(f"SELECT DISTINCT idpel, blth FROM {table_name}")
            )
            
            while True:
                chunk = result.fetchmany(chunk_size)
                if not chunk:
                    break
                existing_keys.update({
                    (str(row[0]).strip().zfill(12), str(row[1]).strip()) for row in chunk
                })

        print(f"📊 Kombinasi IDPEL+BLTH sudah ada di DB: {len(existing_keys)}")
        if existing_keys:
            print(f"Contoh: {list(existing_keys)[:5]}")

        # 3. Pisahkan data baru vs duplikat berdasarkan tuple (IDPEL, BLTH)
        clean_rows = []
        duplicates = []
        
        for _, row in df.iterrows():
            key = (row['IDPEL'], row['BLTH'])
            
            if key in existing_keys:
                duplicates.append(f"{key}")
            else:
                clean_rows.append(row)
                existing_keys.add(key)  # Prevent duplicates in same batch

        success_count = len(clean_rows)
        duplicate_count = len(duplicates)

        print(f"✅ Data baru siap disimpan: {success_count}")
        print(f"⚠️ Duplikat (kombinasi IDPEL+BLTH sudah ada di DB): {duplicate_count}")

        # 4. Simpan ke database jika ada data baru
        if clean_rows:
            clean_df = pd.DataFrame(clean_rows)
            
            # # Clean HTML content if needed
            # for col in clean_df.columns:
            #     if clean_df[col].dtype == object:
            #         clean_df[col] = clean_df[col].str.replace(r'<[^>]*>', '', regex=True)
            
            # Process in smaller batches
            for start in range(0, success_count, batch_size):
                batch = clean_df.iloc[start:start + batch_size]
                
                for attempt in range(1, max_retries + 1):
                    try:
                        with engine.begin() as conn:
                            batch.to_sql(
                                name=table_name,
                                con=conn,
                                if_exists='append',
                                index=False,
                                method='multi',
                                chunksize=min(50, batch_size)
                            )
                        print(f"✅ Berhasil insert batch {start//batch_size + 1}")
                        break
                    except Exception as e:
                        if attempt < max_retries:
                            print(f"🔁 Gagal insert attempt {attempt}, retrying in {retry_delay}s... Error: {str(e)}")
                            time.sleep(retry_delay)
                        else:
                            # Identify which records caused the failure
                            problem_ids = []
                            for _, row in batch.iterrows():
                                try:
                                    with engine.begin() as conn:
                                        row.to_frame().T.to_sql(
                                            name=table_name,
                                            con=conn,
                                            if_exists='append',
                                            index=False
                                        )
                                except Exception as single_error:
                                    problem_ids.append((row['IDPEL'], row['BLTH']))
                                    print(f"⚠️ Gagal insert IDPEL+BLTH {row['IDPEL']}, {row['BLTH']}: {str(single_error)}")
                            
                            raise Exception(f"Gagal insert batch setelah {max_retries} percobaan. Masalah dengan IDPEL+BLTH: {problem_ids}")

            print(f"✅ Sukses insert {success_count} baris ke {table_name}")
        else:
            print("Tidak ada data baru (semua kombinasi IDPEL+BLTH sudah ada).")

        # 5. Return summary tanpa data besar
        return {
            'success_count': success_count,
            'duplicate_count': duplicate_count,
            'duplicates_sample': duplicates[:10],
            'table_name': table_name
        }

    except Exception as e:
        error_msg = f"Error saving to database: {str(e)}"
        print(error_msg)
        flash(error_msg, 'danger')
        raise





def save_to_database_update(df, table_name, engine, batch_size=500, max_retries=3, retry_delay=2):
    try:
        print(f"📥 Jumlah baris awal input: {len(df)}")

        # Normalisasi IDPEL dan BLTH jadi string, hapus duplikat kombinasi keduanya
        df['IDPEL'] = df['IDPEL'].astype(str).str.strip().str.zfill(12)
        df['BLTH'] = df['BLTH'].astype(str).str.strip()
        df = df.drop_duplicates(subset=['IDPEL', 'BLTH'])
        print(f"🧹 Setelah normalisasi & drop duplikat IDPEL+BLTH: {len(df)} baris")

        # Mapping kolom asli ke placeholder SQL (underscore)
        mapping = {
            'BLTH': 'BLTH',
            'LWBPCABUT': 'LWBPCABUT',
            'SELISIH STAN BONGKAR': 'SELISIH_STAN_BONGKAR',
            'LWBP PASANG': 'LWBP_PASANG',
            'KWH SEKARANG': 'KWH_SEKARANG',
            'DELTA PEMKWH': 'DELTA_PEMKWH',
            'DLPD_HITUNG': 'DLPD_HITUNG',
            '%': 'PERSEN',
            'KET': 'KET',
            'SAHLWBP': 'SAHLWBP',
            'JAM NYALA': 'JAM_NYALA',
            'JAMNYALA600': 'JAMNYALA600',
            'MARKING_KOREKSI': 'MARKING_KOREKSI'
        }

        # Kolom yang tidak boleh dioverwrite
        immutable_cols = ['SLALWBP', 'DAYA', 'HASIL PEMERIKSAAN', 'TINDAK LANJUT']

        # Kolom yang ada di df dan di mapping, kecuali immutable
        update_columns = [col for col in mapping if col in df.columns and col not in immutable_cols]
        print(f"📝 Kolom yang akan di-update: {update_columns}")

        updated_rows = 0
        failed_rows = []
        updated_detail = []

        for idx, row in df.iterrows():
            idpel = row['IDPEL']
            blth = row['BLTH']
            update_values = {}
            updated_cols = []

            for col in update_columns:
                param_name = mapping[col]
                value = row[col]

                # Tambah ke update hanya jika tidak NaN
                if pd.notna(value):
                    update_values[param_name] = value
                    updated_cols.append(col)

            if not updated_cols:
                print(f"⚠️ Lewati {idpel}-{blth} karena tidak ada kolom valid untuk update.")
                continue

            # Tambah kunci where
            update_values['idpel'] = idpel
            update_values['blth'] = blth

            # Buat set clause pakai nama kolom asli (pakai spasi)
            set_clause = ', '.join([f"`{col}` = :{mapping[col]}" for col in updated_cols])

            sql = text(f"""
                UPDATE `{table_name}`
                SET {set_clause}
                WHERE `IDPEL` = :idpel AND `BLTH` = :blth
            """)

            for attempt in range(1, max_retries + 1):
                try:
                    with engine.begin() as conn:
                        conn.execute(sql, update_values)
                    updated_rows += 1
                    updated_detail.append((f"{idpel}-{blth}", updated_cols))
                    print(f"✅ Update {idpel}-{blth}: kolom {updated_cols}")
                    break
                except OperationalError as oe:
                    print(f"🔁 Gagal update {idpel}-{blth} attempt {attempt}, retrying in {retry_delay}s...")
                    time.sleep(retry_delay)
                    if attempt == max_retries:
                        failed_rows.append(f"{idpel}-{blth}")
                        print(f"[❌ ERROR] Gagal update {idpel}-{blth}: {oe}")
                except Exception as e:
                    failed_rows.append(f"{idpel}-{blth}")
                    print(f"[❌ ERROR] Gagal update {idpel}-{blth}: {e}")
                    break

        print(f"✅ Sukses update: {updated_rows}")
        print(f"⚠️ Gagal update: {len(failed_rows)}")
        if failed_rows:
            preview = ', '.join(failed_rows[:10]) + ('...' if len(failed_rows) > 10 else '')
            print(f"🧨 IDPEL-BLTH gagal update: {preview}")

        return {
            'updated': updated_rows,
            'failed': failed_rows,
            'detail': updated_detail,
            'table': table_name
        }

    except Exception as e:
        print(f"❌ Fatal error in save_to_database_update: {str(e)}")
        raise




# Global DataFrames to store results

# result_df = None

naik_df = None
turun_df = None
# aman_df = None
div_df = None
kroscek_df = None


@app.route("/1_v2", methods=["GET", "POST"])
@login_required
def index1_v2():
    try:
        # Debug: Start request logging
        print(f"[DEBUG] Starting request - Method: {request.method}, User: {session.get('username')}")
        
        # Authentication and access control
        username = session.get('username', '').lower()
        ulp_param = request.args.get("ulp", "").lower()
        
        # Debug: Log initial parameters
        print(f"[DEBUG] User: {username}, ULP param: {ulp_param}")

        # Configuration mappings
        ulp_to_billing = {
            'ulp_skw': 'billing_skw2',
            'ulp_pmk': 'billing_pmk',
            'ulp_sbs': 'billing_sbs',
            'ulp_bkg': 'billing_bkg',
            'ulp_skr': 'billing_skr',
            'ulp_sdr': 'billing_sdr',
        }
        
        dpm_to_billing = {
            'ulp_skw': 'dpm_skw',
            'ulp_pmk': 'dpm_pmk',
            'ulp_sbs': 'dpm_sbs',
            'ulp_bkg': 'dpm_bkg',
            'ulp_skr': 'dpm_skr',
            'ulp_sdr': 'dpm_sdr',
        }

        user_access = {
            'tel_skw': ['ulp_skw'],
            'tel_pmk': ['ulp_pmk'],
            'tel_sbs': ['ulp_sbs'],
            'tel_bkg': ['ulp_bkg'],
            'tel_skr': ['ulp_skr'],
            'tel_sdr': ['ulp_sdr'],
            'tel_up3': list(ulp_to_billing.keys())
        }

        # Check user access
        allowed_ulp_keys = user_access.get(username, [])
        if not allowed_ulp_keys:
            print(f"[WARNING] Access denied for user: {username}")
            flash("Anda tidak memiliki akses ke data ini.", "danger")
            return redirect(url_for('admin_v2'))

        # Determine tables
        if username == 'tel_up3':
            if not ulp_param:
                print("[WARNING] UP3 user didn't specify ULP")
                flash("User UP3 harus memilih ULP yang valid melalui tombol navigasi (?ulp=...).", "danger")
                return redirect(url_for('admin_v2'))
            
            ulp_key = f"ulp_{ulp_param}"
            if ulp_key not in allowed_ulp_keys:
                print(f"[WARNING] Invalid ULP selected: {ulp_param}")
                flash("ULP yang dipilih tidak valid.", "danger")
                return redirect(url_for('admin_v2'))
        else:
            ulp_key = allowed_ulp_keys[0] if not ulp_param else f"ulp_{ulp_param}"
            if ulp_param and ulp_key not in allowed_ulp_keys:
                print(f"[WARNING] User {username} tried accessing unauthorized ULP: {ulp_param}")
                flash("Akses tidak diizinkan untuk ULP ini.", "danger")
                return redirect(url_for('admin_v2'))

        billing_table = ulp_to_billing.get(ulp_key)
        dpm_table = dpm_to_billing.get(ulp_key)
        # nomet_table = nomet_to_billing[ulp_key]

        if not billing_table or not dpm_table:
            print(f"[ERROR] Table mapping not found for ULP: {ulp_key}")
            flash("Tabel tujuan tidak ditemukan.", "danger")
            return redirect(url_for('admin_v2'))

        # Handle GET request
        if request.method == "GET":
            print("[DEBUG] Handling GET request")
            flash(f"Sugeng Rawuh, {username}!", "info")
            return render_template('index1_v2.html')

        # Handle POST request
        print("[DEBUG] Handling POST request")
        
        # Validate form data
        required_fields = ['blth_kini']
        for field in required_fields:
            if field not in request.form:
                print(f"[ERROR] Missing required field: {field}")
                flash(f"Data {field} harus diisi.", "danger")
                return redirect(request.url)
        try:
            blth_kini = request.form['blth_kini']
            date_blth_kini = datetime.strptime(blth_kini, '%Y%m')
            
            blth_lalu = (date_blth_kini - relativedelta(months=1)).strftime('%Y%m')
            blth_lalulalu = (date_blth_kini - relativedelta(months=2)).strftime('%Y%m')
            blth_lalu3 = (date_blth_kini - relativedelta(months=3)).strftime('%Y%m')
                    
           
            sortir_naik = float(request.form.get('sortir_naik', 0))
            sortir_turun = float(request.form.get('sortir_turun', 0))
            # sortir_naik = 40
            # sortir_turun = 40
        except ValueError as e:
            print(f"[ERROR] Invalid form data: {str(e)}")
            flash("Format data tidak valid.", "danger")
            return redirect(request.url)

        # Debug log
        print(f"[DEBUG] Form params - BLTH: {blth_lalu3}, {blth_lalulalu}, {blth_lalu}, {blth_kini}")
        print(f"[DEBUG] Thresholds - Naik: {sortir_naik}, Turun: {sortir_turun}")

        file_akhir = request.files['file_akhir']
        if file_akhir.filename == '':
            print("[WARNING] No file uploaded")
            flash("Tidak ada file yang dipilih.", "danger")
            return redirect(request.url)

        try:
            print("[DEBUG] Reading uploaded Excel file")
            akhir = pd.read_excel(file_akhir)
            print(f"[DEBUG] Uploaded file contains {len(akhir)} rows")
            

        except Exception as file_error:
            print(f"[ERROR] File processing error: {str(file_error)}")
            flash("Gagal memproses file yang diunggah.", "danger")
            return redirect(request.url)


        # Get data from database with error handling
        try:
            base_query = text(f"SELECT * FROM {dpm_table} WHERE BLTH = :blth")

            print(f"[DEBUG] Fetching lalulalu-1 data for BLTH: {blth_lalu3}")
            file_lalu3 = pd.read_sql(base_query, engine, params={'blth': blth_lalu3})
            print(f"[DEBUG] Retrieved {len(file_lalu3)} rows from lalu3")
            
            print(f"[DEBUG] Fetching lalu3 data for BLTH: {blth_lalulalu}")
            file_lalulalu = pd.read_sql(base_query, engine, params={'blth': blth_lalulalu})
            print(f"[DEBUG] Retrieved {len(file_lalulalu)} rows from lalulalu")
            
            print(f"[DEBUG] Fetching lalu data for BLTH: {blth_lalu}")
            file_lalu = pd.read_sql(base_query, engine, params={'blth': blth_lalu})
            print(f"[DEBUG] Retrieved {len(file_lalu)} rows from lalu")

        except Exception as db_error:
            print(f"[ERROR] Database error: {str(db_error)}")
            flash("Gagal mengambil data dari database.", "danger")
            return redirect(request.url)

        # Process data
        try:
            print("[DEBUG] Starting data processing with copy_dataframe2")
            kroscek_df = copy_dataframe2(
                file_lalu3, file_lalulalu, file_lalu, akhir,
                blth_lalulalu, blth_lalu, blth_kini,blth_lalu3,
                sortir_naik, sortir_turun,
                ulp_param, dpm_table
            )

            
            print(f"[DEBUG] Processed dataframe contains {len(kroscek_df)} rows")
            
            if kroscek_df.empty:
                print("[ERROR] Resulting dataframe is empty")
                flash("Tidak ada data yang diproses.", "warning")
                return redirect(request.url)
                
        except Exception as process_error:
            print(f"[ERROR] Data processing error: {str(process_error)}")
            flash("Terjadi kesalahan saat memproses data.", "danger")
            return redirect(request.url)

        # Save to database
        try:
            print(f"[DEBUG] Saving data to {billing_table}")     
            save_to_database_v2(kroscek_df, billing_table, engine)
            print(f"[INFO] Successfully saved {len(kroscek_df)} rows to {billing_table}")
            flash(f"Data berhasil disimpan ke tabel {billing_table}", "success")
        except Exception as save_error:
            print(f"[ERROR] Database save error: {str(save_error)}")
            flash("Gagal menyimpan data ke database.", "danger")
            return redirect(request.url)
        print("=== DEBUG: data akhir billing:", len(kroscek_df))
        print(kroscek_df['IDPEL'].duplicated().sum(), "Data billing uplikat ditemukan")
        return redirect(url_for('index1_v2', ulp=ulp_param))
    
    except Exception as e:
        print(f"[CRITICAL] Unexpected error in index1_v2: {str(e)}")
        import traceback
        traceback.print_exc()
        flash("Terjadi kesalahan sistem.", "danger")
        return redirect(url_for('index1_v2', ulp=ulp_param))
    

@app.route("/index_update", methods=["GET", "POST"])
@login_required
def index_update():
    try:
        print(f"[DEBUG] Starting request Index_update - Method: {request.method}, User: {session.get('username')}")
        username = session.get('username', '').lower()
        ulp_param = request.args.get("ulp", "").lower()

        ulp_to_billing = {
            'ulp_skw': 'billing_skw2',
            'ulp_pmk': 'billing_pmk',
            'ulp_sbs': 'billing_sbs',
            'ulp_bkg': 'billing_bkg',
            'ulp_skr': 'billing_skr',
            'ulp_sdr': 'billing_sdr',
        }
        dpm_to_billing = {
            'ulp_skw': 'dpm_skw',
            'ulp_pmk': 'dpm_pmk',
            'ulp_sbs': 'dpm_sbs',
            'ulp_bkg': 'dpm_bkg',
            'ulp_skr': 'dpm_skr',
            'ulp_sdr': 'dpm_sdr',
        }
        user_access = {
            'tel_skw': ['ulp_skw'],
            'tel_pmk': ['ulp_pmk'],
            'tel_sbs': ['ulp_sbs'],
            'tel_bkg': ['ulp_bkg'],
            'tel_skr': ['ulp_skr'],
            'tel_sdr': ['ulp_sdr'],
            'tel_up3': list(ulp_to_billing.keys())
        }

        allowed_ulp_keys = user_access.get(username, [])
        if not allowed_ulp_keys:
            flash("Anda tidak memiliki akses ke data ini.", "danger")
            return redirect(url_for('admin_v2'))

        if username == 'tel_up3':
            if not ulp_param:
                flash("User UP3 harus memilih ULP yang valid melalui tombol navigasi (?ulp=...).", "danger")
                return redirect(url_for('admin_v2'))
            ulp_key = f"ulp_{ulp_param}"
            if ulp_key not in allowed_ulp_keys:
                flash("ULP yang dipilih tidak valid.", "danger")
                return redirect(url_for('admin_v2'))
        else:
            ulp_key = allowed_ulp_keys[0] if not ulp_param else f"ulp_{ulp_param}"
            if ulp_param and ulp_key not in allowed_ulp_keys:
                flash("Akses tidak diizinkan untuk ULP ini.", "danger")
                return redirect(url_for('admin_v2'))

        billing_table = ulp_to_billing.get(ulp_key)
        dpm_table = dpm_to_billing.get(ulp_key)

        if not billing_table or not dpm_table:
            flash("Tabel tujuan tidak ditemukan.", "danger")
            return redirect(url_for('admin_v2'))

        if request.method == "GET":
            flash(f"Sugeng Rawuh, {username}!", "info")
            return render_template('index_update.html')

        # POST method
        required_fields = ['blth_kini']
        for field in required_fields:
            if field not in request.form:
                flash(f"Data {field} harus diisi.", "danger")
                return redirect(request.url)

        try:
            blth_kini = request.form['blth_kini']
            date_blth_kini = datetime.strptime(blth_kini, '%Y%m')
            blth_lalu = (date_blth_kini - relativedelta(months=1)).strftime('%Y%m')
            blth_lalulalu = (date_blth_kini - relativedelta(months=2)).strftime('%Y%m')
            blth_lalu3 = (date_blth_kini - relativedelta(months=3)).strftime('%Y%m')

            sortir_naik = float(request.form.get('sortir_naik', 0))
            sortir_turun = float(request.form.get('sortir_turun', 0))
        except ValueError:
            flash("Format data tidak valid.", "danger")
            return redirect(request.url)

        file_akhir = request.files['file_akhir']
        if file_akhir.filename == '':
            flash("Tidak ada file yang dipilih.", "danger")
            return redirect(request.url)

        try:
            akhir = pd.read_excel(file_akhir)
        except Exception:
            flash("Gagal memproses file yang diunggah.", "danger")
            return redirect(request.url)

        try:
            base_query = text(f"SELECT * FROM {dpm_table} WHERE BLTH = :blth")
            file_lalu = pd.read_sql(base_query, engine, params={'blth': blth_lalu})
        except Exception:
            flash("Gagal mengambil data dari database.", "danger")
            return redirect(request.url)

        try:
            kroscek_df = copy_dataframe_update(
                data_lalu=file_lalu,
                data_akhir=akhir,
                blth_kini=blth_kini,
                sortir_naik=sortir_naik,
                sortir_turun=sortir_turun,
                ulp_param=ulp_param,
                dpm_table=dpm_table,
                engine=engine,
                billing_table=billing_table
            )
        except Exception:
            flash("Terjadi kesalahan saat memproses data.", "danger")
            return redirect(request.url)

        if kroscek_df.empty:
            flash("Tidak ada data yang diproses.", "warning")
            return redirect(request.url)

        try:
            result = save_to_database_update(kroscek_df, billing_table, engine)
            updated_ids = kroscek_df['IDPEL'].astype(str).unique().tolist()
            flash(f"Data berhasil disimpan ke tabel {billing_table}", "success")

            # Redirect ke view_update dengan updated_ids di URL param
            return redirect(url_for('view_update', ulp=ulp_param, updated_ids=','.join(updated_ids)))

        except Exception as save_error:
            print(f"[ERROR] Database save error: {str(save_error)}")
            flash("Gagal menyimpan data ke database.", "danger")
            return redirect(request.url)

    except Exception as e:
        print(f"[CRITICAL] Unexpected error in index_update: {str(e)}")
        import traceback
        traceback.print_exc()
        flash("Terjadi kesalahan sistem.", "danger")
        return redirect(url_for('index_update', ulp=ulp_param))
    
    
#DATA AMAN JADI TAMPILKAN SELURUH DATA
@app.route("/view_update", methods=["GET", "POST"])
@login_required
def view_update():
    try:
        active_tab = request.args.get('tab', 'dlpd_3bln')
        username = session.get('username', '').lower()
        ulp_param = request.args.get('ulp', '').lower()


        ulp_to_billing = {
            'ulp_skw': 'billing_skw2',
            'ulp_pmk': 'billing_pmk',
            'ulp_sbs': 'billing_sbs',
            'ulp_bkg': 'billing_bkg',
            'ulp_skr': 'billing_skr',
            'ulp_sdr': 'billing_sdr',
        }

        user_access = {
            'tel_skw': ['ulp_skw'],
            'tel_pmk': ['ulp_pmk'],
            'tel_sbs': ['ulp_sbs'],
            'tel_bkg': ['ulp_bkg'],
            'tel_skr': ['ulp_skr'],
            'tel_sdr': ['ulp_sdr'],
            'tel_up3': list(ulp_to_billing.keys())
        }

        # Validasi akses
        if username == 'tel_up3':
            ulp_key = f"ulp_{ulp_param}"
            if ulp_key not in user_access['tel_up3']:
                flash("ULP tidak valid untuk user UP3.", "danger")
                return redirect(url_for('admin_v2'))
            billing_table = ulp_to_billing[ulp_key]
        else:
            allowed_ulp = user_access.get(username)
            if not allowed_ulp or (ulp_param and f"ulp_{ulp_param}" not in allowed_ulp):
                flash("Anda tidak diperbolehkan melihat ULP lain.", "danger")
                return redirect(url_for('admin_v2'))
            billing_table = ulp_to_billing[allowed_ulp[0]]

        # Jika tidak ada updated_ids, tampilkan semua data normal
        selected_kelompok = request.args.get("kdkelompok")
        # base_query = f"SELECT * FROM {billing_table} WHERE 1=1"
        base_query = f"SELECT * FROM {billing_table} WHERE MARKING_KOREKSI > 0"

        if selected_kelompok:
            base_query += f" AND MARKING_KOREKSI = '{selected_kelompok}'"

        data_naik = pd.read_sql(text(base_query + " AND KET = 'NAIK'"), engine)
        data_turun = pd.read_sql(text(base_query + " AND KET = 'TURUN'"), engine)
        data_div = pd.read_sql(text(base_query + " AND KET = 'DIV/NA'"), engine)
        data_aman = pd.read_sql(text(base_query + " AND KET IN ('AMAN','NAIK','TURUN','DIV/NA')"), engine)


        jam_nyala_min = request.args.get("jam_nyala_min", default=599, type=float)
        jam_nyala_max = request.args.get("jam_nyala_max", default=9999, type=float)
        jam_nyala_query = base_query + " AND `JAM NYALA` IS NOT NULL"
        if jam_nyala_min and jam_nyala_max:
            jam_nyala_query += f" AND `JAM NYALA` BETWEEN {jam_nyala_min} AND {jam_nyala_max}"
        data_jam_nyala = pd.read_sql(text(jam_nyala_query), engine)

        data_dlpd_3bln = pd.read_sql(
            text(base_query + " AND DLPD_3BLN = :condition"),
            engine, params={'condition': 'Naik50% R3BLN'}
        )
        hasil_options = [
            "SESUAI", "TEMPER NYALA", "SALAH STAN", "SALAH FOTO", "FOTO BURAM",
            "ANOMALI PDL", "LEBIH TAGIH", "KURANG TAGIH", "BKN FOTO KWH",
            "BENCANA", "3BLN TANPA STAN", "BACA ULANG", "MASUK 720JN"
        ]

        def create_editable_df(df):
            if df.empty:
                return df
            df_display = df.copy()
            df_display['HASIL PEMERIKSAAN'] = [
                f'<select name="hasil_pemeriksaan_{row["IDPEL"]}" class="form-select">'
                '<option value="" selected hidden></option>' +
                ''.join([
                    f'<option value="{opt}" {"selected" if str(row.get("HASIL PEMERIKSAAN", "")) == opt else ""}>{opt}</option>'
                    for opt in hasil_options
                ]) +
                '</select>'
                for _, row in df.iterrows()
            ]
            df_display['TINDAK LANJUT'] = [
                f'<textarea name="tindak_lanjut_{row["IDPEL"]}" class="tindak-lanjut" rows="2" cols="15">'
                f'{"" if pd.isna(row.get("TINDAK LANJUT")) else escape(str(row.get("TINDAK LANJUT", "")))}</textarea>'
                for _, row in df.iterrows()
            ]
            return df_display

        flash("PILIH KOREKSI KE?", "danger")

        dlpd_3bln_html = create_editable_df(data_dlpd_3bln).to_html(
            classes="table table-striped", index=False, escape=False, na_rep=''
        ) if not data_dlpd_3bln.empty else "<p>Tidak ada data DLPD_3BLN</p>"

        naik_html_v2 = create_editable_df(data_naik).to_html(
            classes="table table-striped", index=False, escape=False, na_rep=''
        ) if not data_naik.empty else "<p>Tidak ada data sortir naik</p>"

        turun_html_v2 = create_editable_df(data_turun).to_html(
            classes="table table-striped", index=False, escape=False, na_rep=''
        ) if not data_turun.empty else "<p>Tidak ada data sortir turun</p>"

        div_html_v2 = create_editable_df(data_div).to_html(
            classes="table table-striped", index=False, escape=False, na_rep=''
        ) if not data_div.empty else "<p>Tidak ada data sortir DIV/NA</p>"

        aman_html_v2 = create_editable_df(data_aman).to_html(
            classes="table table-striped", index=False, escape=False, na_rep=''
        ) if not data_aman.empty else "<p>Tidak ada data sortir AMAN</p>"

        jam_nyala_html = create_editable_df(data_jam_nyala).to_html(
            classes="table table-striped", index=False, escape=False, na_rep=''
        ) if not data_jam_nyala.empty else "<p>Tidak ada data sortir JN 600 up</p>"

        return render_template(
            "view_update.html",
            username=username,
            dlpd_3bln_html=dlpd_3bln_html,
            naik_html_v2=naik_html_v2,
            turun_html_v2=turun_html_v2,
            div_html_v2=div_html_v2,
            aman_html_v2=aman_html_v2,
            jam_nyala_html=jam_nyala_html,
            active_tab=active_tab,
            selected_kelompok=selected_kelompok,
            jam_nyala_min=jam_nyala_min,
            jam_nyala_max=jam_nyala_max
        )

    except Exception as e:
        print(f"Error in view_update: {str(e)}")
        flash("Pilih hari baca", "danger")
        return render_template("view_update.html", dlpd_3bln_html="<p>Error loading data</p>")


@app.route("/generate_nomet", methods=['GET'])
@login_required
def generate_nomet():
    print(f"[DEBUG] Starting /generate_nomet - Method: {request.method}, User: {session.get('username')}")

    username = session.get('username', '').lower()
    ulp_param = request.args.get("ulp", "").lower()
    start_kdkelompok = request.args.get("start_kdkelompok")
    end_kdkelompok = request.args.get("end_kdkelompok")

    print(f"[DEBUG] Params - username: {username}, ulp_param: {ulp_param}, KDKELOMPOK: {start_kdkelompok}–{end_kdkelompok}")

    try:
        start_kdkelompok = int(start_kdkelompok)
        end_kdkelompok = int(end_kdkelompok)
    except (TypeError, ValueError):
        flash("Rentang KDKELOMPOK harus berupa angka.", "danger")
        return redirect(url_for('admin_v2'))

    ulp_to_billing = {
        'ulp_skw': 'billing_skw2',
        'ulp_pmk': 'billing_pmk',
        'ulp_sbs': 'billing_sbs',
        'ulp_bkg': 'billing_bkg',
        'ulp_skr': 'billing_skr',
        'ulp_sdr': 'billing_sdr',
    }

    user_access = {
        'tel_skw': ['ulp_skw'],
        'tel_pmk': ['ulp_pmk'],
        'tel_sbs': ['ulp_sbs'],
        'tel_bkg': ['ulp_bkg'],
        'tel_skr': ['ulp_skr'],
        'tel_sdr': ['ulp_sdr'],
        'tel_up3': list(ulp_to_billing.keys())
    }

    allowed_ulp_keys = user_access.get(username, [])
    if not allowed_ulp_keys:
        flash("Anda tidak memiliki akses ke data ini.", "danger")
        return redirect(url_for('admin_v2'))

    if username == 'tel_up3':
        if not ulp_param:
            flash("User UP3 harus memilih ULP yang valid.", "danger")
            return redirect(url_for('admin_v2'))
        ulp_key = f"ulp_{ulp_param}"
        if ulp_key not in allowed_ulp_keys:
            flash("ULP yang dipilih tidak valid.", "danger")
            return redirect(url_for('admin_v2'))
    else:
        ulp_key = allowed_ulp_keys[0] if not ulp_param else f"ulp_{ulp_param}"
        if ulp_param and ulp_key not in allowed_ulp_keys:
            flash("Akses tidak diizinkan untuk ULP ini.", "danger")
            return redirect(url_for('admin_v2'))

    billing_table = ulp_to_billing.get(ulp_key)
    nomet_table = f"nomet_{ulp_key.split('_')[1]}"
    print(f"[DEBUG] Tables - billing_table: {billing_table}, nomet_table: {nomet_table}")

    max_retries = 3
    retry_delay = 5
    total_updated = 0

    for attempt in range(max_retries):
        try:
            with engine.begin() as conn:
                print(f"[DEBUG] Attempt {attempt + 1}: starting transaction")

                print("[DEBUG] Dropping and creating temp_nomet_cleaned...")
                conn.execute(text("DROP TEMPORARY TABLE IF EXISTS temp_nomet_cleaned;"))
                conn.execute(text(f"""
                    CREATE TEMPORARY TABLE temp_nomet_cleaned (
                        idpel_clean VARCHAR(255),
                        no_meter VARCHAR(255),
                        INDEX idx_idpel (idpel_clean)
                    ) AS
                    SELECT LOWER(TRIM(idpel)) AS idpel_clean, MAX(no_meter) AS no_meter
                    FROM {nomet_table}
                    WHERE no_meter IS NOT NULL AND TRIM(no_meter) != ''
                    GROUP BY idpel_clean;
                """))

                print("[DEBUG] Dropping and creating temp_batch_update...")
                conn.execute(text("DROP TEMPORARY TABLE IF EXISTS temp_batch_update;"))
                conn.execute(
                    text(f"""
                        CREATE TEMPORARY TABLE temp_batch_update (
                            id INT PRIMARY KEY AUTO_INCREMENT,
                            billing_id VARCHAR(255),
                            INDEX idx_billing_id (billing_id)
                        ) AS
                        SELECT b.IDPEL AS billing_id
                        FROM {billing_table} AS b
                        JOIN temp_nomet_cleaned AS n
                            ON LOWER(TRIM(b.IDPEL)) = n.idpel_clean
                        WHERE (b.NOMET IS NULL OR TRIM(b.NOMET) = '' OR b.NOMET IN ('-', '0', 'kosong'))
                            AND b.KDKELOMPOK BETWEEN :start_kdkelompok AND :end_kdkelompok;
                    """),
                    {
                        "start_kdkelompok": start_kdkelompok,
                        "end_kdkelompok": end_kdkelompok
                    }
                )

                batch_size = 10000
                offset = 0
                while True:
                    batch_ids = conn.execute(text(f"""
                        SELECT billing_id FROM temp_batch_update
                        ORDER BY id
                        LIMIT {batch_size} OFFSET {offset};
                    """)).fetchall()

                    if not batch_ids:
                        print("[DEBUG] No more batch IDs found. Breaking.")
                        break

                    id_list = [str(row[0]) for row in batch_ids]
                    print(f"[DEBUG] Processing batch {offset} – size: {len(id_list)}")

                    result = conn.execute(
                        text(f"""
                            UPDATE {billing_table} AS b
                            JOIN temp_nomet_cleaned AS n
                                ON LOWER(TRIM(b.IDPEL)) = n.idpel_clean
                            SET b.NOMET = n.no_meter
                            WHERE b.IDPEL IN :id_list;
                        """),
                        {"id_list": id_list}
                    )

                    updated = result.rowcount
                    total_updated += updated
                    print(f"[DEBUG] Updated {updated} rows in batch {offset}")
                    offset += batch_size
                    time.sleep(0.1)

                print("[DEBUG] Checking duplicate IDPEL entries in nomet...")
                duplikat_result = conn.execute(text(f"""
                    SELECT LOWER(TRIM(idpel)) AS idpel, COUNT(*) AS jumlah
                    FROM {nomet_table}
                    WHERE no_meter IS NOT NULL AND TRIM(no_meter) != ''
                    GROUP BY idpel
                    HAVING COUNT(*) > 1
                    ORDER BY jumlah DESC
                    LIMIT 20;
                """)).fetchall()

                if total_updated > 0:
                    flash(f"Baris No Meter ter-update (KDKELOMPOK {start_kdkelompok}-{end_kdkelompok}): {total_updated}", "success")
                else:
                    flash("Tidak ada data yang diupdate.", "info")

                if duplikat_result:
                    print("⚠️ IDPEL dengan No Meter ganda ditemukan:")
               
                else:
                    flash("Tidak ditemukan IDPEL ganda di tabel nomor meter.", "info")

                print("[DEBUG] Finished all updates successfully.")
                break

        except OperationalError as err:
            if "Lock wait timeout" in str(err) and attempt < max_retries - 1:
                print(f"[WARNING] Lock timeout on attempt {attempt + 1}, retrying in {retry_delay}s")
                time.sleep(retry_delay)
                continue
            print(f"[ERROR] OperationalError: {err}")
            flash("Terjadi kesalahan saat update NOMET: " + str(err), "danger")
            return redirect(url_for('index1_v2', ulp=ulp_param))
        except Exception as err:
            print(f"[ERROR] Exception occurred: {err}")
            flash("Terjadi kesalahan saat update NOMET.", "danger")
            return redirect(url_for('index1_v2', ulp=ulp_param))

    return redirect(url_for('index1_v2', ulp=ulp_param))




##########################
@app.route("/view_data1_v2", methods=["GET", "POST"])
@login_required
def view_data1_v2():
    try:
        active_tab = request.args.get('tab', 'dlpd_3bln')  # Default to dlpd_3bln
        username = session.get('username', '').lower()
        ulp_param = request.args.get('ulp', '').lower()

        ulp_to_billing = {
            'ulp_skw': 'billing_skw2',
            'ulp_pmk': 'billing_pmk',
            'ulp_sbs': 'billing_sbs',
            'ulp_bkg': 'billing_bkg',
            'ulp_skr': 'billing_skr',
            'ulp_sdr': 'billing_sdr',
        }

        user_access = {
            'tel_skw': ['ulp_skw'],
            'tel_pmk': ['ulp_pmk'],
            'tel_sbs': ['ulp_sbs'],
            'tel_bkg': ['ulp_bkg'],
            'tel_skr': ['ulp_skr'],
            'tel_sdr': ['ulp_sdr'],
            'tel_up3': list(ulp_to_billing.keys())
        }

        # Validate access
        if username == 'tel_up3':
            ulp_key = f"ulp_{ulp_param}"
            if ulp_key not in user_access['tel_up3']:
                flash("ULP tidak valid untuk user UP3.", "danger")
                return redirect(url_for('admin_v2'))
            billing_table = ulp_to_billing[ulp_key]
        else:
            allowed_ulp = user_access.get(username)
            if not allowed_ulp or f"ulp_{ulp_param}" not in allowed_ulp and ulp_param:
                flash("Anda tidak diperbolehkan melihat ULP lain.", "danger")
                return redirect(url_for('admin_v2'))
            billing_table = ulp_to_billing[allowed_ulp[0]]
        
        selected_kelompok = request.args.get("kdkelompok")
        blth = request.args.get("blth")

        # Base query hanya ambil bulan dan kelompok tertentu
        base_query = f"SELECT * FROM {billing_table} WHERE 1=1"

        if blth:
            base_query += f" AND BLTH = '{blth}'"
        else:
            # ambil BLTH terbaru jika user belum pilih
            latest_blth = pd.read_sql(f"SELECT MAX(BLTH) as latest FROM {billing_table}", engine).iloc[0]['latest']
            base_query += f" AND BLTH = '{latest_blth}'"

        if selected_kelompok:
            base_query += f" AND KDKELOMPOK = '{selected_kelompok}'"
        else:
            flash("Pilih KDKELOMPOK dahulu!", "danger")

        
        # selected_kelompok = request.args.get("kdkelompok")
        # # Base query with proper WHERE clause
        # base_query = f"SELECT * FROM {billing_table} WHERE 1=1"
        # if selected_kelompok:
        #     base_query += f" AND KDKELOMPOK = '{selected_kelompok}'"

        # Get data for each tab with proper WHERE clauses
        data_naik = pd.read_sql(text(base_query + " AND KET = 'NAIK'"), engine)
        data_turun = pd.read_sql(text(base_query + " AND KET = 'TURUN'"), engine)
        data_div = pd.read_sql(text(base_query + " AND KET = 'DIV/NA'"), engine)
        
        jam_nyala_min = request.args.get("jam_nyala_min", default=599, type=float)
        jam_nyala_max = request.args.get("jam_nyala_max", default=9999, type=float)

        # Corrected Jam Nyala query - using backticks for column with space
        jam_nyala_query = base_query + " AND `JAM NYALA` IS NOT NULL"
        if jam_nyala_min and jam_nyala_max:
            jam_nyala_query += f" AND `JAM NYALA` BETWEEN {jam_nyala_min} AND {jam_nyala_max}"
        
        data_jam_nyala = pd.read_sql(text(jam_nyala_query), engine)

        data_dlpd_3bln = pd.read_sql(
            text(base_query + " AND DLPD_3BLN = :condition"), 
            engine, params={'condition': 'Naik50% R3BLN'}
        )
        hasil_options = [
            "SESUAI", "TEMPER NYALA", "SALAH STAN", "SALAH FOTO", "FOTO BURAM",
            "ANOMALI PDL", "LEBIH TAGIH", "KURANG TAGIH", "BKN FOTO KWH",
            "BENCANA", "3BLN TANPA STAN", "BACA ULANG", "MASUK 720JN"
        ]

        
        def create_editable_df(df):
            if df.empty:
                return df
            df_display = df.copy()
            df_display['HASIL PEMERIKSAAN'] = [
                f'<select name="hasil_pemeriksaan_{row["IDPEL"]}" class="form-select">'
                '<option value="" selected hidden></option>' +
                ''.join([
                    f'<option value="{opt}" {"selected" if str(row.get("HASIL PEMERIKSAAN", "")) == opt else ""}>{opt}</option>'
                    for opt in hasil_options
                ]) +
                '</select>'
                for _, row in df.iterrows()
            ]

            df_display['STAN_VERIFIKASI'] = [
                f'<textarea name="stan_verifikasi_{row["IDPEL"]}" class="stan-verifikasi" rows="1" cols="15">'
                f'{"" if pd.isna(row.get("STAN_VERIFIKASI")) else escape(str(row.get("STAN_VERIFIKASI", "")))}</textarea>'
                for _, row in df.iterrows()
            ]
            df_display['TINDAK LANJUT'] = [
                f'<textarea name="tindak_lanjut_{row["IDPEL"]}" class="tindak-lanjut" rows="2" cols="15">'
                f'{"" if pd.isna(row.get("TINDAK LANJUT")) else escape(str(row.get("TINDAK LANJUT", "")))}</textarea>'
                for _, row in df.iterrows()
            ]
            return df_display
        
        flash("Pilih KDKELOMPOK Dahulu!", "danger")
        dlpd_3bln_html = create_editable_df(data_dlpd_3bln).to_html(
            classes="table table-striped", index=False, escape=False, na_rep=''
        ) if not data_dlpd_3bln.empty else "<p>Tidak ada data DLPD_3BLN</p>"

        naik_html_v2 = create_editable_df(data_naik).to_html(
            classes="table table-striped", index=False, escape=False, na_rep=''
        ) if not data_naik.empty else "<p>Tidak ada data sortir naik</p>"

        turun_html_v2 = create_editable_df(data_turun).to_html(
            classes="table table-striped", index=False, escape=False, na_rep=''
        ) if not data_turun.empty else "<p>Tidak ada data sortir turun</p>"

        div_html_v2 = create_editable_df(data_div).to_html(
            classes="table table-striped", index=False, escape=False, na_rep=''
        ) if not data_div.empty else "<p>Tidak ada data sortir DIV/NA</p>"

        jam_nyala_html = create_editable_df(data_jam_nyala).to_html(
            classes="table table-striped", index=False, escape=False, na_rep=''
        ) if not data_jam_nyala.empty else "<p>Tidak ada data sortir JN 600 up</p>"
        
        return render_template(
             "view_data1_v2.html",
            username=username,
            dlpd_3bln_html=dlpd_3bln_html,
            naik_html_v2=naik_html_v2,
            turun_html_v2=turun_html_v2,
            div_html_v2=div_html_v2,
            jam_nyala_html=jam_nyala_html,
            active_tab=active_tab,
            selected_kelompok=selected_kelompok,
            jam_nyala_min=jam_nyala_min,
            jam_nyala_max=jam_nyala_max
        )

    except Exception as e:
        print(f"Error in view_data1_v2: {str(e)}")
        flash("Pilih hari baca", "danger")
        return render_template("view_data1_v2.html", dlpd_3bln_html="<p>Error loading data</p>")
    



def get_tab_condition(tab_name):
    """Returns SQL condition for each tab type"""
    conditions = {
        'naik': "AND KET = 'NAIK'",
        'turun': "AND KET = 'TURUN'",
        'div': "AND KET = 'DIV/NA'",
        'dlpd': "AND DLPD_3BLN = 'Naik50% R3BLN'",
        'jam_nyala': "AND JAMNYALA600 = '600Up'"
    }
    return conditions.get(tab_name, "")

# @app.route("/update_pemeriksaan", methods=["POST"])
# @login_required
# def update_pemeriksaan():
#     try:
#         username = session.get('username', '').lower()
#         table_mapping = {
#             'tel_skw': 'billing_skw2',
#             'tel_pmk': 'billing_pmk',
#             'tel_sbs': 'billing_sbs',
#             'tel_bkg': 'billing_bkg',
#             'tel_skr': 'billing_skr',
#             'tel_sdr': 'billing_sdr',
#             'tel_up3': 'billing_up3',
#         }
#         billing_table = table_mapping.get(username)
        
#         if not billing_table:
#             return jsonify({"success": False, "message": "User tidak dikenali"}), 400

#         selected_kelompok = request.form.get("selected_kelompok")
        
#         # Get all form data
#         form_data = request.form.to_dict()
        
#         # Process each field
#         for key, value in form_data.items():
#             if key.startswith("hasil_pemeriksaan_"):
#                 idpel = key.replace("hasil_pemeriksaan_", "")
#                 tindak_lanjut = form_data.get(f"tindak_lanjut_{idpel}", "")
                
#                 # Update database
#                 update_query = text(f"""
#                     UPDATE {billing_table} 
#                     SET "HASIL PEMERIKSAAN" = :hasil, "TINDAK LANJUT" = :tindak 
#                     WHERE IDPEL = :idpel
#                     {"AND KDKELOMPOK IN (:kelompok)" if selected_kelompok else ""}
#                 """)
                
#                 params = {
#                     'hasil': value,
#                     'tindak': tindak_lanjut,
#                     'idpel': idpel
#                 }
#                 if selected_kelompok:
#                     params['kelompok'] = selected_kelompok
                
#                 engine.execute(update_query, params)
        
#         return jsonify({"success": True, "message": "Data berhasil diperbarui"})
        
#     except Exception as e:
#         print(f"Error updating data: {str(e)}")
#         return jsonify({"success": False, "message": str(e)}), 500

# app.route("/simpan_all_data", methods=["POST"])
# def simpan_all_data():
#     try:
#         # Menerima data JSON dari AJAX
#         data = request.get_json()
        
#         # Debugging: Print data yang diterima
#         print('Data yang diterima:', data)

#         # Lakukan sesuatu dengan data, misalnya simpan ke database
#         # Contoh:
#         # for key, value in data.items():
#         #     process_data(key, value)

#         return jsonify({"success": True})
#     except Exception as e:
#         return jsonify({"success": False, "message": str(e)})




# #SUDAH SAVE PAKAI FOREIGN KEY, INI DATA INDUKNYA

# @app.route('/simpan_aman', methods=['POST'])
# @login_required
# def simpan_aman():
#     try:
#         # Get all data from the database
#         current_data = pd.read_sql("SELECT * FROM billing_skw2", engine)
        
#         # Create a copy to store updates
#         updated_data = current_data.copy()
        
#         # Process form updates for all rows
#         for i in range(len(current_data)):
#             # Get values from form
#             hasil_pemeriksaan = request.form.get(f'hasil_pemeriksaan_{i}', '')
#             tindak_lanjut = request.form.get(f'tindak_lanjut_{i}', '')
            
#             # Update the DataFrame
#             updated_data.at[i, 'HASIL PEMERIKSAAN'] = hasil_pemeriksaan
#             updated_data.at[i, 'TINDAK LANJUT'] = tindak_lanjut
        
#         # Save back to database in a single transaction
#         with engine.begin() as connection:
#             # First delete all existing data
#             connection.execute(text("DELETE FROM billing_skw2"))
            
#             # Then insert the updated data
#             updated_data.to_sql(
#                 name='billing_skw2',
#                 con=connection,
#                 if_exists='append',
#                 index=False
#             )
        
#         flash("Data berhasil disimpan!", "success")
#     except Exception as e:
#         print(f"Error saving data: {str(e)}")
#         flash("Gagal menyimpan data", "error")
    
#     return redirect(url_for('view_data1_v2'))


@app.route('/simpan_dlpd', methods=['POST'])
@login_required
def simpan_dlpd():
    try:
        active_tab = request.form.get("active_tab", "dlpd_3bln")  # fallback ke tab ini kalau kosong
        username = session.get('username', '').lower()
        billing_table = TABLE_MAPPING.get(username)

        if not billing_table:
            flash("User tidak dikenali.", "danger")
            return redirect(url_for('view_data1_v2', tab=active_tab))

        selected_kelompok = request.form.get("kdkelompok")
        print(f"[DEBUG] POST kdkelompok (from form): {selected_kelompok}")

        form_data = request.form
        print(f"Form Data: {form_data}")
        selected_kelompok = form_data.get("kdkelompok")
        idpels = {field.replace('hasil_pemeriksaan_', '') for field in form_data if field.startswith('hasil_pemeriksaan_')}
        idpels.update({field.replace('tindak_lanjut_', '') for field in form_data if field.startswith('tindak_lanjut_')})
        idpels.update({field.replace('stan_verifikasi_', '') for field in form_data if field.startswith('stan_verifikasi_')})

        updated_count = 0
        with engine.begin() as connection:
            for idpel in idpels:
                hasil = form_data.get(f'hasil_pemeriksaan_{idpel}', '').strip()
                tindak = form_data.get(f'tindak_lanjut_{idpel}', '').strip()
                stan = form_data.get(f'stan_verifikasi_{idpel}', '').strip()

                if hasil or tindak:
                    query = f"""
                    UPDATE {billing_table}
                    SET `HASIL PEMERIKSAAN` = :hasil,
                        `TINDAK LANJUT` = :tindak,
                        STAN_VERIFIKASI = :stan
                    WHERE IDPEL = :id 
                      AND DLPD_3BLN = 'Naik50% R3BLN'
                      AND KDKELOMPOK = :kelompok
                     """
                    connection.execute(text(query), {
                        'hasil': hasil,
                        'tindak': tindak,
                        'stan': stan,
                        'id': idpel,
                        'kelompok': selected_kelompok
                    })
                    updated_count += 1
        flash(f"{updated_count} data berhasil diperbarui!", "success")

    except Exception as e:
        import logging
        logging.exception("Error in simpan_dlpd")
        flash(f"Gagal memperbarui data DLPD: {str(e)}", "danger")

    return redirect(url_for('view_data1_v2', tab=active_tab, kdkelompok=selected_kelompok))

def simpan_by_ket(filter_key):
    try:
        active_tab = request.form.get("active_tab", "dlpd_3bln")  # Get active tab from form
        username = session.get('username', '').lower()
        billing_table = TABLE_MAPPING.get(username)

        if not billing_table:
            flash("User tidak dikenali.", "danger")
            return redirect(url_for('view_data1_v2', tab=active_tab))

        selected_kelompok = request.form.get("kdkelompok")
        print(f"[DEBUG] POST kdkelompok (from form): {selected_kelompok}")

        form_data = request.form
        print(f"Form Data: {form_data}")
        selected_kelompok = form_data.get("kdkelompok")
        idpels = {field.replace('hasil_pemeriksaan_', '') for field in form_data if field.startswith('hasil_pemeriksaan_')}
        idpels.update({field.replace('tindak_lanjut_', '') for field in form_data if field.startswith('tindak_lanjut_')})
        idpels.update({field.replace('stan_verifikasi_', '') for field in form_data if field.startswith('stan_verifikasi_')})
        updated_count = 0
        with engine.begin() as connection:
            for idpel in idpels:
                hasil = form_data.get(f'hasil_pemeriksaan_{idpel}', '').strip()
                tindak = form_data.get(f'tindak_lanjut_{idpel}', '').strip()
                stan = form_data.get(f'stan_verifikasi_{idpel}', '').strip()
                
                if hasil or tindak:
                    if filter_key == "JAMNYALA600":
                        query = f"""
                            UPDATE {billing_table}
                            SET `HASIL PEMERIKSAAN` = :hasil,
                                `TINDAK LANJUT` = :tindak,
                                STAN_VERIFIKASI = :stan
                            WHERE IDPEL = :id AND JAMNYALA600 = '600Up'
                        """
                    else:
                        query = f"""
                            UPDATE {billing_table}
                            SET `HASIL PEMERIKSAAN` = :hasil,
                                `TINDAK LANJUT` = :tindak,
                                STAN_VERIFIKASI = :stan
                            WHERE IDPEL = :id AND KET = :ket
                        """
                    connection.execute(text(query), {
                        'hasil': hasil,
                        'tindak': tindak,
                        'stan':stan,
                        'id': idpel,
                        'ket': filter_key if filter_key != "JAMNYALA600" else None
                    })
                    updated_count += 1

        flash(f"{updated_count} data berhasil diperbarui untuk kategori {filter_key}!", "success")
    except Exception as e:
        import logging
        logging.exception(f"Error in simpan_by_ket ({filter_key})")
        flash(f"Gagal memperbarui data {filter_key}", "danger")

    
    return redirect(request.referrer or url_for('view_data1_v2'))


@app.route('/simpan_naik', methods=['POST'])
@login_required
def simpan_naik():
    return simpan_by_ket("NAIK")

@app.route('/simpan_turun', methods=['POST'])
@login_required
def simpan_turun():
    return simpan_by_ket("TURUN")

@app.route('/simpan_div', methods=['POST'])
@login_required
def simpan_div():
    return simpan_by_ket("DIV/NA")

@app.route('/simpan_aman', methods=['POST'])
@login_required
def simpan_aman():
    return simpan_by_ket("AMAN")

# @app.route('/simpan_jam_nyala', methods=['POST'])
# @login_required
# def simpan_jam_nyala():
#     return simpan_by_ket("JAMNYALA600")

@app.route('/simpan_jam_nyala', methods=['POST'])
@login_required
def simpan_jam_nyala():
    try:
        active_tab = request.form.get("active_tab", "jam_nyala")
        username = session.get('username', '').lower()
        billing_table = TABLE_MAPPING.get(username)

        if not billing_table:
            flash("User tidak dikenali.", "danger")
            return redirect(url_for('view_data1_v2', tab=active_tab))

        selected_kelompok = request.form.get("kdkelompok")
        jam_min = request.form.get("jam_nyala_min", 0)
        jam_max = request.form.get("jam_nyala_max", 9999)

        form_data = request.form
        idpels = {field.replace('hasil_pemeriksaan_', '') for field in form_data 
                 if field.startswith('hasil_pemeriksaan_')}
        idpels.update({field.replace('tindak_lanjut_', '') for field in form_data 
                      if field.startswith('tindak_lanjut_')})
        idpels.update({field.replace('stan_verifikasi_', '') for field in form_data 
                      if field.startswith('stan_verifikasi_')})
        
        updated_count = 0
        with engine.begin() as connection:
            for idpel in idpels:
                hasil = form_data.get(f'hasil_pemeriksaan_{idpel}', '').strip()
                tindak = form_data.get(f'tindak_lanjut_{idpel}', '').strip()
                stan = form_data.get(f'stan_verifikasi_{idpel}', '').strip()
                
                if hasil or tindak:
                    query = f"""
                        UPDATE {billing_table}
                        SET `HASIL PEMERIKSAAN` = :hasil,
                            `TINDAK LANJUT` = :tindak,
                            STAN_VERIFIKASI = :stan
                        WHERE IDPEL = :id 
                        AND KDKELOMPOK = :kelompok
                        AND `JAM NYALA` BETWEEN :min AND :max
                    """
                    connection.execute(text(query), {
                        'hasil': hasil,
                        'tindak': tindak,
                        'stan': stan,
                        'id': idpel,
                        'kelompok': selected_kelompok,
                        'min': jam_min,
                        'max': jam_max
                    })
                    updated_count += 1

        flash(f"{updated_count} data berhasil diperbarui untuk Jam Nyala ({jam_min}-{jam_max})!", "success")
    except Exception as e:
        import logging
        logging.exception(f"Error in simpan_jam_nyala: {str(e)}")
        flash("Gagal memperbarui data Jam Nyala", "danger")

    return redirect(url_for('view_data1_v2', tab=active_tab, kdkelompok=selected_kelompok))
#SAVE NYA MENCAR KETIKA SELESAI KLIK SIMPAN DATA
# @app.route('/simpan_dlpd', methods=['POST'])
# @login_required
# def simpan_dlpd():
#     try:
#         active_tab = request.form.get("active_tab", "dlpd_3bln")  # fallback ke tab ini kalau kosong
#         username = session.get('username', '').lower()
#         billing_table = TABLE_MAPPING.get(username)

#         if not billing_table:
#             flash("User tidak dikenali.", "danger")
#             return redirect(url_for('view_data1_v2'))

#         selected_kelompok = request.form.get("kdkelompok")
#         print(f"[DEBUG] POST kdkelompok (from form): {selected_kelompok}")

#         form_data = request.form
#         print(f"Form Data: {form_data}")
#         selected_kelompok = form_data.get("kdkelompok")
#         idpels = {field.replace('hasil_pemeriksaan_', '') for field in form_data if field.startswith('hasil_pemeriksaan_')}
#         idpels.update({field.replace('tindak_lanjut_', '') for field in form_data if field.startswith('tindak_lanjut_')})

#         updated_count = 0
#         with engine.begin() as connection:
#             for idpel in idpels:
#                 hasil = form_data.get(f'hasil_pemeriksaan_{idpel}', '').strip()
#                 tindak = form_data.get(f'tindak_lanjut_{idpel}', '').strip()

#                 # Hanya lanjutkan jika hasil atau tindak tidak kosong
#                 if hasil or tindak:
#                         query = f"""
#                         UPDATE {billing_table}
#                         SET `HASIL PEMERIKSAAN` = :hasil,
#                             `TINDAK LANJUT` = :tindak
#                         WHERE IDPEL = :id 
#                           AND DLPD_3BLN = 'Naik50% R3BLN'
#                           AND KDKELOMPOK = :kelompok
#                          """
#                         connection.execute(text(query), {
#                         'hasil': hasil,
#                         'tindak': tindak,
#                         'id': idpel,
#                         'kelompok': selected_kelompok
#                     })
#                 updated_count += 1
#         flash(f"{updated_count} data berhasil diperbarui!", "success")

#     except Exception as e:
#         import logging
#         logging.exception("Error in simpan_dlpd")
#         flash(f"Gagal memperbarui data DLPD: {str(e)}", "danger")

#     return redirect(url_for('view_data1_v2',tab=active_tab, kdkelompok=selected_kelompok))

# def simpan_by_ket(filter_key):
#     try:
#         username = session.get('username', '').lower()
#         billing_table = TABLE_MAPPING.get(username)

#         if not billing_table:
#             flash("User tidak dikenali.", "danger")
#             return redirect(url_for('view_data1_v2'))

#         selected_kelompok = request.form.get("kdkelompok")
#         print(f"[DEBUG] POST kdkelompok (from form): {selected_kelompok}")

#         form_data = request.form
#         print(f"Form Data: {form_data}")
#         selected_kelompok = form_data.get("kdkelompok")
#         idpels = {field.replace('hasil_pemeriksaan_', '') for field in form_data if field.startswith('hasil_pemeriksaan_')}
#         idpels.update({field.replace('tindak_lanjut_', '') for field in form_data if field.startswith('tindak_lanjut_')})
#         updated_count = 0
#         with engine.begin() as connection:
#             for idpel in idpels:
#                 hasil = form_data.get(f'hasil_pemeriksaan_{idpel}', '').strip()
#                 tindak = form_data.get(f'tindak_lanjut_{idpel}', '').strip()
                
#                 if hasil or tindak:
#                     if filter_key == "JAMNYALA600":
#                         query = f"""
#                             UPDATE {billing_table}
#                             SET `HASIL PEMERIKSAAN` = :hasil,
#                                 `TINDAK LANJUT` = :tindak
#                             WHERE IDPEL = :id AND JAMNYALA600 = '600Up'
#                         """
#                     else:
#                         query = f"""
#                             UPDATE {billing_table}
#                             SET `HASIL PEMERIKSAAN` = :hasil,
#                                 `TINDAK LANJUT` = :tindak
#                             WHERE IDPEL = :id AND KET = :ket
#                         """
#                     connection.execute(text(query), {
#                         'hasil': hasil,
#                         'tindak': tindak,
#                         'id': idpel,
#                         'ket': filter_key if filter_key != "JAMNYALA600" else None
#                     })
#                     updated_count += 1

#         flash(f"{updated_count} data berhasil diperbarui untuk kategori {filter_key}!", "success")
#     except Exception as e:
#         import logging
#         logging.exception(f"Error in simpan_by_ket ({filter_key})")
#         flash(f"Gagal memperbarui data {filter_key}", "danger")

#     return redirect(url_for('view_data1_v2', kdkelompok=selected_kelompok))

# @app.route('/simpan_naik', methods=['POST'])
# @login_required
# def simpan_naik():
#     return simpan_by_ket("NAIK")

# @app.route('/simpan_turun', methods=['POST'])
# @login_required
# def simpan_turun():
#     return simpan_by_ket("TURUN")

# @app.route('/simpan_div', methods=['POST'])
# @login_required
# def simpan_div():
#     return simpan_by_ket("DIV/NA")

# @app.route('/simpan_jam_nyala', methods=['POST'])
# @login_required
# def simpan_jam_nyala():
#     return simpan_by_ket("JAMNYALA600")


# # format lama=
# @app.route('/simpan_dlpd', methods=['POST'])
# @login_required
# def simpan_dlpd():
#     try:
#         username = session.get('username', '').lower()
#         billing_table = TABLE_MAPPING.get(username)

#         if not billing_table:
#             flash("User tidak dikenali.", "danger")
#             return redirect(url_for('view_data1_v2'))

#         selected_kelompok = request.form.get("kdkelompok")
#         print(f"[DEBUG] POST kdkelompok (from form): {selected_kelompok}")

#         form_data = request.form
#         print(f"Form Data: {form_data}")
#         idpels = {field.replace('hasil_pemeriksaan_', '') for field in form_data if field.startswith('hasil_pemeriksaan_')}
#         idpels.update({field.replace('tindak_lanjut_', '') for field in form_data if field.startswith('tindak_lanjut_')})

#         updated_count = 0
#         with engine.begin() as connection:
#             for idpel in idpels:
#                 hasil = form_data.get(f'hasil_pemeriksaan_{idpel}', '').strip()
#                 tindak = form_data.get(f'tindak_lanjut_{idpel}', '').strip()
                
#                 if hasil or tindak:
#                     query = f"""
#                         UPDATE {billing_table}
#                         SET `HASIL PEMERIKSAAN` = :hasil,
#                             `TINDAK LANJUT` = :tindak
#                         WHERE IDPEL = :id 
#                           AND DLPD_3BLN = 'Naik50% R3BLN'
#                           AND KDKELOMPOK = :kelompok
#                     """
#                     connection.execute(text(query), {
#                         'hasil': hasil,
#                         'tindak': tindak,
#                         'id': idpel,
#                         'kelompok': selected_kelompok
#                     })
#                     updated_count += 1

#         flash(f"{updated_count} data berhasil diperbarui!", "success")
#     except Exception as e:
#         import logging
#         logging.exception("Error in simpan_dlpd")
#         flash("Gagal memperbarui data DLPD", "danger")

#     return redirect(url_for('view_data1_v2', kdkelompok=selected_kelompok))


# @app.route('/simpan_dlpd', methods=['POST'])
# @login_required
# def simpan_dlpd():
#     try:
#         username = session.get('username', '').lower()
#         billing_table = TABLE_MAPPING.get(username)

#         if not billing_table:
#             flash("User tidak dikenali.", "danger")
#             return redirect(url_for('view_data1_v2'))
#         selected_kelompok = request.args.get("kdkelompok")
#         form_data = request.form
#         idpels = {field.replace('hasil_pemeriksaan_', '') for field in form_data if field.startswith('hasil_pemeriksaan_')}
#         idpels.update({field.replace('tindak_lanjut_', '') for field in form_data if field.startswith('tindak_lanjut_')})
#         updated_count = 0
#         with engine.begin() as connection:
#             for idpel in idpels:
#                 hasil = form_data.get(f'hasil_pemeriksaan_{idpel}', '').strip()
#                 tindak = form_data.get(f'tindak_lanjut_{idpel}', '').strip()
                
#                 if hasil or tindak:
#                     query = f"""
#                         UPDATE {billing_table}
#                         SET `HASIL PEMERIKSAAN` = :hasil,
#                             `TINDAK LANJUT` = :tindak
#                         WHERE IDPEL = :id AND DLPD_3BLN = 'Naik50% R3BLN' AND KDKELOMPOK: {selected_kelompok}
#                     """
#                     connection.execute(text(query), {'hasil': hasil, 'tindak': tindak, 'id': idpel})
#                     updated_count += 1

#         flash(f"{updated_count} data berhasil diperbarui!", "success")
#     except Exception as e:
#         logging.exception("Error in simpan_dlpd")
#         flash("Gagal memperbarui data DLPD", "danger")

#     return redirect(url_for('view_data1_v2'))


# @app.route('/simpan_dlpd', methods=['POST'])
# @login_required
# def simpan_dlpd():
#     try:
#         # Get all IDPELs from the form data
#         form_data = request.form
#         idpels = set()
        
#         # Extract all unique IDPELs from the form field names
#         for field_name in form_data.keys():
#             if field_name.startswith('hasil_pemeriksaan_'):
#                 idpel = field_name.replace('hasil_pemeriksaan_', '')
#                 idpels.add(idpel)
#             elif field_name.startswith('tindak_lanjut_'):
#                 idpel = field_name.replace('tindak_lanjut_', '')
#                 idpels.add(idpel)

#         with engine.begin() as connection:
#             for idpel in idpels:
#                 hasil = request.form.get(f'hasil_pemeriksaan_{idpel}', '').strip()
#                 tindak = request.form.get(f'tindak_lanjut_{idpel}', '').strip()
                
#                 if hasil or tindak:
#                     connection.execute(
#                         text("""
#                             UPDATE billing_sdr
#                             SET `HASIL PEMERIKSAAN` = :hasil,
#                                 `TINDAK LANJUT` = :tindak
#                             WHERE IDPEL = :id AND DLPD_3BLN = 'Naik50% R3BLN'
#                         """),
#                         {'hasil': hasil, 'tindak': tindak, 'id': idpel}
#                     )
#                     print(f"Updated IDPEL {idpel}")

#         flash("Data DLPD berhasil diperbarui!", "success")
        
#     except Exception as e:
#         print(f"Error in simpan_dlpd: {str(e)}")
#         flash("Gagal memperbarui data DLPD", "error")
    
#     return redirect(url_for('view_data1_v2'))

# @app.route('/simpan_naik', methods=['POST'])
# @login_required
# def simpan_naik():
#     try:
#         # Get all IDPELs from the form data
#         form_data = request.form
#         idpels = set()
        
#         # Extract all unique IDPELs from the form field names
#         for field_name in form_data.keys():
#             if field_name.startswith('hasil_pemeriksaan_'):
#                 idpel = field_name.replace('hasil_pemeriksaan_', '')
#                 idpels.add(idpel)
#             elif field_name.startswith('tindak_lanjut_'):
#                 idpel = field_name.replace('tindak_lanjut_', '')
#                 idpels.add(idpel)

#         with engine.begin() as connection:
#             for idpel in idpels:
#                 hasil = request.form.get(f'hasil_pemeriksaan_{idpel}', '').strip()
#                 tindak = request.form.get(f'tindak_lanjut_{idpel}', '').strip()
                
#                 if hasil or tindak:
#                     connection.execute(
#                         text("""
#                             UPDATE billing_skw2
#                             SET `HASIL PEMERIKSAAN` = :hasil,
#                                 `TINDAK LANJUT` = :tindak
#                             WHERE IDPEL = :id AND KET = 'NAIK'
#                         """),
#                         {'hasil': hasil, 'tindak': tindak, 'id': idpel}
#                     )
#                     print(f"Updated IDPEL {idpel}")

#         flash("Data NAIK berhasil diperbarui!", "success")
        
#     except Exception as e:
#         print(f"Error in simpan_NAIK: {str(e)}")
#         flash("Gagal memperbarui data NAIK", "error")
    
#     return redirect(url_for('view_data1_v2'))

# @app.route('/simpan_turun', methods=['POST'])
# @login_required
# def simpan_turun():
#     try:
#         # Get all IDPELs from the form data
#         form_data = request.form
#         idpels = set()
        
#         # Extract all unique IDPELs from the form field names
#         for field_name in form_data.keys():
#             if field_name.startswith('hasil_pemeriksaan_'):
#                 idpel = field_name.replace('hasil_pemeriksaan_', '')
#                 idpels.add(idpel)
#             elif field_name.startswith('tindak_lanjut_'):
#                 idpel = field_name.replace('tindak_lanjut_', '')
#                 idpels.add(idpel)

#         with engine.begin() as connection:
#             for idpel in idpels:
#                 hasil = request.form.get(f'hasil_pemeriksaan_{idpel}', '').strip()
#                 tindak = request.form.get(f'tindak_lanjut_{idpel}', '').strip()
                
#                 if hasil or tindak:
#                     connection.execute(
#                         text("""
#                             UPDATE billing_skw2
#                             SET `HASIL PEMERIKSAAN` = :hasil,
#                                 `TINDAK LANJUT` = :tindak
#                             WHERE IDPEL = :id AND KET = 'TURUN'
#                         """),
#                         {'hasil': hasil, 'tindak': tindak, 'id': idpel}
#                     )
#                     print(f"Updated IDPEL {idpel}")

#         flash("Data TURUN berhasil diperbarui!", "success")
        
#     except Exception as e:
#         print(f"Error in simpan_TURUN: {str(e)}")
#         flash("Gagal memperbarui data TURUN", "error")
    
#     return redirect(url_for('view_data1_v2'))

# @app.route('/simpan_div', methods=['POST'])
# @login_required
# def simpan_div():
#     try:
#         # Get all IDPELs from the form data
#         form_data = request.form
#         idpels = set()
        
#         # Extract all unique IDPELs from the form field names
#         for field_name in form_data.keys():
#             if field_name.startswith('hasil_pemeriksaan_'):
#                 idpel = field_name.replace('hasil_pemeriksaan_', '')
#                 idpels.add(idpel)
#             elif field_name.startswith('tindak_lanjut_'):
#                 idpel = field_name.replace('tindak_lanjut_', '')
#                 idpels.add(idpel)

#         with engine.begin() as connection:
#             for idpel in idpels:
#                 hasil = request.form.get(f'hasil_pemeriksaan_{idpel}', '').strip()
#                 tindak = request.form.get(f'tindak_lanjut_{idpel}', '').strip()
                
#                 if hasil or tindak:
#                     connection.execute(
#                         text("""
#                             UPDATE billing_skw2
#                             SET `HASIL PEMERIKSAAN` = :hasil,
#                                 `TINDAK LANJUT` = :tindak
#                             WHERE IDPEL = :id AND KET = 'DIV/NA'
#                         """),
#                         {'hasil': hasil, 'tindak': tindak, 'id': idpel}
#                     )
#                     print(f"Updated IDPEL {idpel}")

#         flash("Data DIV/NA berhasil diperbarui!", "success")
        
#     except Exception as e:
#         print(f"Error in simpan_DIV/NA: {str(e)}")
#         flash("Gagal memperbarui data DIV/NA", "error")
    
#     return redirect(url_for('view_data1_v2'))

# @app.route('/simpan_jam_nyala', methods=['POST'])
# @login_required
# def simpan_jam_nyala():
#     try:
#         # Get all IDPELs from the form data
#         form_data = request.form
#         idpels = set()
        
#         # Extract all unique IDPELs from the form field names
#         for field_name in form_data.keys():
#             if field_name.startswith('hasil_pemeriksaan_'):
#                 idpel = field_name.replace('hasil_pemeriksaan_', '')
#                 idpels.add(idpel)
#             elif field_name.startswith('tindak_lanjut_'):
#                 idpel = field_name.replace('tindak_lanjut_', '')
#                 idpels.add(idpel)

#         with engine.begin() as connection:
#             for idpel in idpels:
#                 hasil = request.form.get(f'hasil_pemeriksaan_{idpel}', '').strip()
#                 tindak = request.form.get(f'tindak_lanjut_{idpel}', '').strip()
                
#                 if hasil or tindak:
#                     connection.execute(
#                         text("""
#                             UPDATE billing_skw2
#                             SET `HASIL PEMERIKSAAN` = :hasil,
#                                 `TINDAK LANJUT` = :tindak
#                             WHERE IDPEL = :id AND JAMNYALA600 = '>600'
#                         """),
#                         {'hasil': hasil, 'tindak': tindak, 'id': idpel}
#                     )
#                     print(f"Updated IDPEL {idpel}")

#         flash("Data JAMNYALA600 berhasil diperbarui!", "success")
        
#     except Exception as e:
#         print(f"Error in simpan_JAMNYALA: {str(e)}")
#         flash("Gagal memperbarui data JAMNYALA", "error")
    
#     return redirect(url_for('view_data1_v2'))



@app.route('/1')

def main_dashboard():

    # Your main code to render data tables

    return render_template('view_data.html')



def get_db_connection():

    return mysql.connector.connect(**db_config)



@app.route('/download_data/<table>')

def download_data(table):

    # Query data from the specified table

    query = f"SELECT * FROM {table}"



    # Get data from the database

    data = pd.read_sql(query, engine)



    # Create Excel file in memory

    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:

        data.to_excel(writer, index=False, sheet_name=table)

        # No need to call save() or close() explicitly; using 'with' handles it

    output.seek(0)



    # Send file

    return send_file(

        output,

        as_attachment=True,

        download_name=f"{table}_data.xlsx",

        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    )




@app.route('/2')

def main_dashboard2():

    # Your main code to render data tables

    return render_template('view_data2.html')



def get_db_connection():

    return mysql.connector.connect(**db_config)



@app.route('/download_data2/<table>')

def download_data2(table):

    # Query data from the specified table

    query = f"SELECT * FROM {table}"



    # Get data from the database

    data = pd.read_sql(query, engine)



    # Create Excel file in memory

    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:

        data.to_excel(writer, index=False, sheet_name=table)

        # No need to call save() or close() explicitly; using 'with' handles it

    output.seek(0)



    # Send file

    return send_file(

        output,

        as_attachment=True,

        download_name=f"{table}_data.xlsx",

        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    )

     

@app.route('/3')

def main_dashboard3():

    # Your main code to render data tables

    return render_template('view_data3.html')



def get_db_connection():

    return mysql.connector.connect(**db_config)



@app.route('/download_data3/<table>')

def download_data3(table):

    # Query data from the specified table

    query = f"SELECT * FROM {table}"



    # Get data from the database

    data = pd.read_sql(query, engine)



    # Create Excel file in memory

    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:

        data.to_excel(writer, index=False, sheet_name=table)

        # No need to call save() or close() explicitly; using 'with' handles it

    output.seek(0)



    # Send file

    return send_file(

        output,

        as_attachment=True,

        download_name=f"{table}_data.xlsx",

        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    ) 



@app.route('/4')

def main_dashboard4():

    # Your main code to render data tables

    return render_template('view_data4.html')



def get_db_connection():

    return mysql.connector.connect(**db_config)



@app.route('/download_data4/<table>')

def download_data4(table):

    # Query data from the specified table

    query = f"SELECT * FROM {table}"



    # Get data from the database

    data = pd.read_sql(query, engine)



    # Create Excel file in memory

    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:

        data.to_excel(writer, index=False, sheet_name=table)

        # No need to call save() or close() explicitly; using 'with' handles it

    output.seek(0)



    # Send file

    return send_file(

        output,

        as_attachment=True,

        download_name=f"{table}_data.xlsx",

        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    ) 

 



@app.route('/5')

def main_dashboard5():

    # Your main code to render data tables

    return render_template('view_data5.html')



def get_db_connection():

    return mysql.connector.connect(**db_config)



@app.route('/download_data5/<table>')

def download_data5(table):

    # Query data from the specified table

    query = f"SELECT * FROM {table}"



    # Get data from the database

    data = pd.read_sql(query, engine)



    # Create Excel file in memory

    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:

        data.to_excel(writer, index=False, sheet_name=table)

        # No need to call save() or close() explicitly; using 'with' handles it

    output.seek(0)



    # Send file

    return send_file(

        output,

        as_attachment=True,

        download_name=f"{table}_data.xlsx",

        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    )  



@app.route('/6')

def main_dashboard6():

    # Your main code to render data tables

    return render_template('view_data6.html')



def get_db_connection():

    return mysql.connector.connect(**db_config)



@app.route('/download_data6/<table>')

def download_data6(table):

    # Query data from the specified table

    query = f"SELECT * FROM {table}"



    # Get data from the database

    data = pd.read_sql(query, engine)



    # Create Excel file in memory

    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:

        data.to_excel(writer, index=False, sheet_name=table)

        # No need to call save() or close() explicitly; using 'with' handles it

    output.seek(0)



    # Send file

    return send_file(

        output,

        as_attachment=True,

        download_name=f"{table}_data.xlsx",

        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    )  



# Set a secret key for session management (required for flash messages)

app.secret_key = '040104'  # Replace with a strong, unique key

@app.route("/download_excel", methods=["GET"])
def download_excel():
    try:
        username = session.get('username', '').lower()
        ulp_param = request.args.get('ulp', '').lower()
        tab_param = request.args.get('tab', '').lower()

        ulp_to_billing = {
            'ulp_skw': 'billing_skw2',
            'ulp_pmk': 'billing_pmk',
            'ulp_sbs': 'billing_sbs',
            'ulp_bkg': 'billing_bkg',
            'ulp_skr': 'billing_skr',
            'ulp_sdr': 'billing_sdr',
        }

        user_access = {
            'tel_skw': ['ulp_skw'],
            'tel_pmk': ['ulp_pmk'],
            'tel_sbs': ['ulp_sbs'],
            'tel_bkg': ['ulp_bkg'],
            'tel_skr': ['ulp_skr'],
            'tel_sdr': ['ulp_sdr'],
            'tel_up3': list(ulp_to_billing.keys())
        }

        # Tentukan tabel utama
        if username == 'tel_up3':
            ulp_key = f'ulp_{ulp_param}'
            if ulp_key not in user_access['tel_up3']:
                return jsonify({"error": "ULP tidak valid untuk user UP3"}), 403
            billing_table = ulp_to_billing[ulp_key]
        else:
            allowed_ulp = user_access.get(username)
            if not allowed_ulp:
                return jsonify({"error": "Akun tidak memiliki akses ULP"}), 403
            billing_table = ulp_to_billing[allowed_ulp[0]]

        # Tentukan filter berdasarkan tab
        if tab_param == "naik":
            ket_filter = "KET = 'NAIK'"
        elif tab_param == "turun":
            ket_filter = "KET = 'TURUN'"
        elif tab_param == "div":
            ket_filter = "KET IN ('DIV/NA', 'DIV')"
        elif tab_param == "dlpd_3bln":
            ket_filter = "KET = 'DLPD 3BLN'"
        elif tab_param == "jam_nyala":
            ket_filter = "KET = 'JAM NYALA'"
        else:
            return jsonify({"error": f"Tab '{tab_param}' tidak valid"}), 400

        with engine.connect() as connection:
            # Ambil BLTH terbaru
            result = connection.execute(text(f"SELECT MAX(BLTH) FROM {billing_table}"))
            latest_blth = result.scalar()

            if not latest_blth:
                return jsonify({"error": "Tidak ditemukan data BLTH"}), 404

            # Ambil data sesuai BLTH dan tab
            query = text(f"""
                SELECT * FROM {billing_table}
                WHERE BLTH = :blth AND {ket_filter}
            """)
            df = pd.read_sql(query, connection, params={"blth": latest_blth})

        if df.empty:
            return jsonify({"error": "Data untuk BLTH terbaru tidak ditemukan"}), 404

        # Buat file Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"Data_{latest_blth}"
        ws.append(df.columns.tolist())

        foto_columns = ["GRAFIK", "FOTO AKHIR", "FOTO LALU", "FOTO LALU2", "FOTO 3BLN"]
        foto_indexes = [i for i, col in enumerate(df.columns) if col.upper() in foto_columns]

        for i, row in df.iterrows():
            for j, value in enumerate(row):
                cell = ws.cell(row=i + 2, column=j + 1)
                if j in foto_indexes:
                    soup = BeautifulSoup(str(value), "html.parser")
                    a_tag = soup.find("a")
                    if a_tag and a_tag.has_attr("href"):
                        url = a_tag["href"]
                        cell.value = "LINK"
                        cell.hyperlink = url
                        cell.font = Font(color="0000EE", underline="single")
                    else:
                        cell.value = "TIDAK ADA LINK"
                else:
                    cell.value = str(value)

        # Kirim hasil file
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=f"{billing_table}_{tab_param}_{latest_blth}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print(f"Terjadi kesalahan: {str(e)}")
        return jsonify({"error": f"Gagal membuat file Excel: {str(e)}"}), 500





#DOWNLOAD EXCEL PER HB
@app.route("/download_excel_hb", methods=["GET"])
def download_excel_hb():
    try:
        username = session.get('username', '').lower()
        ulp_param = request.args.get('ulp', '').lower()

        ulp_to_billing = {
            'ulp_skw': 'billing_skw2',
            'ulp_pmk': 'billing_pmk',
            'ulp_sbs': 'billing_sbs',
            'ulp_bkg': 'billing_bkg',
            'ulp_skr': 'billing_skr',
            'ulp_sdr': 'billing_sdr',
        }

        user_access = {
            'tel_skw': ['ulp_skw'],
            'tel_pmk': ['ulp_pmk'],
            'tel_sbs': ['ulp_sbs'],
            'tel_bkg': ['ulp_bkg'],
            'tel_skr': ['ulp_skr'],
            'tel_sdr': ['ulp_sdr'],
            'tel_up3': list(ulp_to_billing.keys())
        }

        # Tentukan tabel berdasarkan role
        if username == 'tel_up3':
            ulp_key = f'ulp_{ulp_param}'
            if ulp_key not in user_access['tel_up3']:
                return jsonify({"error": "ULP tidak valid untuk user UP3"}), 403
            billing_table = ulp_to_billing[ulp_key]
        else:
            allowed_ulp = user_access.get(username)
            if not allowed_ulp:
                return jsonify({"error": "Akun tidak memiliki akses ULP"}), 403
            if ulp_param:
                return jsonify({"error": "Anda tidak diperbolehkan memilih ULP lain"}), 403
            billing_table = ulp_to_billing[allowed_ulp[0]]

        # # Gunakan request.args karena GET method
        # input1_hb = string(request.args.get('input1', 0))
        # input2_hb = string(request.args.get('input2', 0))

        # # Query data sesuai hari baca
        # query = text(f"""
        #     SELECT * FROM {billing_table}
        #     WHERE KDKELOMPOK BETWEEN :start AND :end
        # """)
        # df = pd.read_sql(query, engine, params={"start": input1_hb, "end": input2_hb})
        input1 = request.args.get('input1', '').strip().upper()
        input2 = request.args.get('input2', '').strip().upper()

        kelompok_order = ['1','2','3','4','5','6','7','8','P','A','I']

        if input1 not in kelompok_order or input2 not in kelompok_order:
            return "Input KDKELOMPOK tidak valid", 400

        # Ambil range dari posisi index
        idx1 = kelompok_order.index(input1)
        idx2 = kelompok_order.index(input2)

        if idx1 > idx2:
            idx1, idx2 = idx2, idx1  # tukar agar urutan benar

        kelompok_range = kelompok_order[idx1:idx2+1]

        placeholders = ','.join([':k'+str(i) for i in range(len(kelompok_range))])
        query = text(f"""
            SELECT * FROM {billing_table}
            WHERE KDKELOMPOK IN ({placeholders})
        """)

        params = {f'k{i}': v for i, v in enumerate(kelompok_range)}
        df = pd.read_sql(query, engine, params=params)

        if df.empty:
            return jsonify({"error": "Tabel kosong"}), 404

        # Tulis ke Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(df.columns.tolist())

        foto_columns = ["GRAFIK", "FOTO AKHIR", "FOTO LALU", "FOTO LALU2", "FOTO 3BLN"]
        foto_indexes = [i for i, col in enumerate(df.columns) if col.upper() in foto_columns]

        for i, row in df.iterrows():
            for j, value in enumerate(row):
                cell = ws.cell(row=i + 2, column=j + 1)
                if j in foto_indexes:
                    soup = BeautifulSoup(str(value), "html.parser")
                    a_tag = soup.find("a")
                    if a_tag and a_tag.has_attr("href"):
                        url = a_tag["href"]
                        cell.value = "LINK"
                        cell.hyperlink = url
                        cell.font = Font(color="0000EE", underline="single")
                        cell.style = "Hyperlink"
                    else:
                        cell.value = "TIDAK ADA LINK"
                else:
                    cell.value = str(value)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
        output,
        as_attachment=True,
        download_name=f"{billing_table}_hb_{input1}_to_{input2}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


    except Exception as e:
        print(f"Terjadi kesalahan: {str(e)}")
        return jsonify({"error": "Gagal membuat file Excel"}), 500


#DELETE TABLE CELAR TABLE HAPUS DATA BARU
@app.route("/clear_table", methods=["POST"])
def clear_table():
    table = request.args.get("table")
    if not table or not re.match(r"^billing_[a-z]+$", table):
        return jsonify({"message": "Tabel tidak valid"}), 400

    try:
        db.session.execute(text(f"DELETE FROM {table}"))
        db.session.commit()
        return jsonify({"message": f"Data dari {table} berhasil dihapus."})
    except Exception as e:
        return jsonify({"message": f"Gagal menghapus data: {str(e)}"}), 500
    


# @app.route("/download_excel", methods=["GET"])
# def download_excel():
#     try:
#         # Ambil nama tabel dari query parameter (misalnya, ?table=billing_skw2)
#         table = request.args.get('table')
#         if not table:
#             return jsonify({"error": "Nama tabel tidak ditemukan"}), 400

#         # Ambil data dari database
#         query = text(f"SELECT * FROM {table}")
#         df = pd.read_sql(query, engine)

#         if df.empty:
#             return jsonify({"error": "Tabel kosong"}), 404

#         # Inisialisasi workbook
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Sheet1"

#         # Tulis header
#         ws.append(df.columns.tolist())

#         # Daftar nama kolom yang mengandung link foto
#         foto_columns = ["FOTO AKHIR", "FOTO LALU", "FOTO LALU2"]
#         foto_indexes = [i for i, col in enumerate(df.columns) if col.upper() in foto_columns]

#         # Tulis data dan handle hyperlink
#         for i, row in df.iterrows():
#             for j, value in enumerate(row):
#                 cell = ws.cell(row=i + 2, column=j + 1)  # +2 karena header di baris 1

#                 if j in foto_indexes:
#                     soup = BeautifulSoup(str(value), "html.parser")
#                     a_tag = soup.find("a")
#                     if a_tag and a_tag.has_attr("href"):
#                         url = a_tag["href"]
#                         cell.value = "LINK FOTO"
#                         cell.hyperlink = url
#                         cell.font = Font(color="0000EE", underline="single")
#                         cell.style = "Hyperlink"
#                     else:
#                         cell.value = "TIDAK ADA LINK"
#                 else:
#                     cell.value = str(value)

#         # Simpan ke memory (BytesIO)
#         output = BytesIO()
#         wb.save(output)
#         output.seek(0)

#         return send_file(
#             output,
#             as_attachment=True,
#             download_name=f"{table}.xlsx",
#             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )

#     except Exception as e:
#         print(f"Terjadi kesalahan: {str(e)}")
#         return jsonify({"error": "Gagal membuat file Excel"}), 500


app.secret_key = '040104'  # Replace with a strong, unique key
@app.route("/delete_data_2", methods=["POST"])
def delete_data_2():
    try:
        table = request.json.get('table')  # Mengambil nama tabel dari request
        if not table:
            return jsonify({"error": "Nama tabel tidak ditemukan"}), 400
        # Membuat query DELETE menggunakan text
        delete_query = text(f"DELETE FROM {table}")  # Gunakan text() untuk query SQL
        with engine.connect() as connection:
            # Eksekusi query DELETE terlebih dahulu
            result = connection.execute(delete_query)
            # Periksa jika tidak ada data yang dihapus
            if result.rowcount == 0:
                return jsonify({"error": "Tidak ada data yang dihapus"}), 400
            # Query ALTER TABLE untuk reset auto increment (Hanya untuk MySQL atau MariaDB)
            try:
                alter_query = text(f"ALTER TABLE {table} AUTO_INCREMENT = 1")
                connection.execute(alter_query)  # Eksekusi ALTER TABLE
            except Exception as alter_error:
                # Jika ALTER gagal, tampilkan pesan bahwa reset auto increment gagal
                print(f"Gagal mereset auto increment: {alter_error}")
                return jsonify({"message": "Data berhasil dihapus, namun gagal mereset auto increment."}), 200
        # Jika berhasil menghapus data dan mereset auto increment, beri pesan sukses
        return jsonify({"message": "Data berhasil dihapus"}), 200
    except Exception as e:
        # Jika ada kesalahan yang terjadi, cetak kesalahan namun kembalikan error spesifik
        print(f"Terjadi kesalahan: {str(e)}")  # Log kesalahan
        return jsonify({"error": "Terjadi kesalahan saat menghapus data."}), 500

@app.route("/delete_data_by_range", methods=["POST"])
def delete_data_by_range():
    try:
        data = request.json
        table = data.get("table")
        start_kd = data.get("start_kd")
        end_kd = data.get("end_kd")

        if not table or start_kd is None or end_kd is None:
            return jsonify({"error": "Data tidak lengkap."}), 400

        # Debug log (sementara)
        print(f"Menghapus dari tabel {table} dengan KDKELOMPOK {start_kd} - {end_kd}")

        delete_query = text(f"""
            DELETE FROM `{table}`
            WHERE KDKELOMPOK BETWEEN :start_kd AND :end_kd
        """)

        with engine.begin() as connection:  # begin() akan auto commit
            result = connection.execute(delete_query, {"start_kd": start_kd, "end_kd": end_kd})
            if result.rowcount == 0:
                return jsonify({"error": "Tidak ada data yang cocok dengan rentang hari baca tersebut."}), 400

        return jsonify({"message": f"Data dengan hari baca {start_kd}-{end_kd} berhasil dihapus."}), 200
    except Exception as e:
        print(f"Terjadi kesalahan: {str(e)}")
        return jsonify({"error": "Terjadi kesalahan saat menghapus data."}), 500


# CEK ALAMAT (rizki)
# from fuzzywuzzy import fuzz
# from requests.adapters import HTTPAdapter
# from urllib3.util.retry import Retry

# ==============================================================================
# FUNGSI INTI
# ==============================================================================

# # Global job storage (use Redis in production)
# jobs = {}

# # Logging setup
# logging.basicConfig(level=logging.INFO)
# logger = logging.getLogger(__name__)

# # Function to create a session with retries
# def get_session_with_retries():
#     session = requests.Session()
#     retries = Retry(total=3, backoff_factor=1, status_forcelist=[403, 500, 502, 503, 504])
#     session.mount('https://', HTTPAdapter(max_retries=retries))
#     return session

# # Function to get address from coordinates using Nominatim
# def get_address_from_coordinates_nominatim(lat, lon, agenda_id, db_session):
#     url = "https://nominatim.openstreetmap.org/reverse"
#     params = {
#         'lat': lat,
#         'lon': lon,
#         'format': 'json'
#     }
#     headers = {
#         'User-Agent': 'AddressValidatorApp (your-email@example.com)'  # Replace with your actual email
#     }
#     try:
#         response = requests.get(url, params=params, headers=headers, timeout=10)
#         response.raise_for_status()
#         data = response.json()
#         address = data.get('display_name', 'Address not found')

#         # Remove specific regions and clean up
#         address_cleaned = address.replace("Kalimantan Barat", "").replace("Kalimantan", "").replace("Indonesia", "").strip()
#         address_cleaned = re.sub(r'\s+', ' ', address_cleaned)

#         # Clean alamat asli from DB
#         record = db_session.query(KppPksr).filter_by(no_agenda_pensem=agenda_id).first()
#         if not record:
#             return None, None, None
#         alamat_induk = clean_address(record.Alamat or "")

#         # Clean nominatim
#         address_nominatim_cleaned = clean_address(address_cleaned)

#         # Find common words
#         alamat_induk_words = set(alamat_induk.split())
#         nominatim_words = set(address_nominatim_cleaned.split())
#         common_nominatim_words = alamat_induk_words.intersection(nominatim_words)

#         # Calculate similarity score
#         similarity_nominatim_score = fuzz.ratio(alamat_induk, address_nominatim_cleaned)
#         threshold = 41
#         is_valid_nominatim = similarity_nominatim_score >= threshold or bool(common_nominatim_words)
#         status_tikor = 'ALAMAT VALID' if is_valid_nominatim else 'ALAMAT TIDAK VALID'

#         # Generate link
#         link_validasi = f"https://www.google.com/maps?q={lat},{lon}"

#         return address_cleaned, status_tikor, link_validasi
#     except Exception as e:
#         logger.error(f"Error in Nominatim for agenda {agenda_id}: {e}")
#         return f"Error: {e}", 'ERROR', None

# # Function to clean addresses
# def clean_address(address):
#     if isinstance(address, str):
#         address = re.sub(r'[^\w\s]', '', address)  # Remove non-word characters
#         address = address.lower()  # Convert to lowercase
#         address = re.sub(r'\s+', ' ', address)  # Remove extra spaces
#         return address.strip()
#     return ""








#IMPORT app_kdbaca
# kdbaca_bp = create_kdbaca_routes(login_required)
# app.register_blueprint(kdbaca_bp)




# if __name__ == "__main__":
#     print("Registered routes:")
#     for rule in app.url_map.iter_rules():
#         print(rule)
#     app.run(port=40000)

if __name__ == "__main__":
    print("Registered routes:")
    for rule in app.url_map.iter_rules():
        print(rule)
    with app.app_context():
        db.create_all()
    app.run(port=40000, debug=True)





# if __name__ == '__main__':

#     app.run(debug=True) 

