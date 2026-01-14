from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from sqlalchemy import create_engine, text
from sqlalchemy.exc import IntegrityError
import pandas as pd
import numpy as np
from datetime import datetime
from dateutil.relativedelta import relativedelta
import os
import logging
import sys
import traceback


from flask import send_file, jsonify, session, request
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font
from bs4 import BeautifulSoup
import io
from io import BytesIO
import pandas as pd
import re
import mysql.connector
from datetime import datetime

from flask import Flask, request, send_file, jsonify, session, render_template
from sqlalchemy import create_engine, text
from openpyxl.styles import Font, PatternFill


# app.py
from monitoring import monitoring_bp



def get_db_connection():
    return mysql.connector.connect(
        host='localhost',
        user='root',
        password='',
        database='billing_gabungan'  # sesuaikan dengan nama database kamu
    )

# =================== SETUP LOGGING ===================
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler('app_errors.log', encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.secret_key = 'your_super_secret_key_here'

app.register_blueprint(monitoring_bp)

# Upload folder
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Database connection
engine = create_engine("mysql+pymysql://root:@localhost/billing_gabungan", 
                      pool_pre_ping=True, 
                      pool_recycle=3600)

# =================== HELPER FUNCTIONS ===================
def normalize_idpel(idpel):
    """Normalize IDPEL to 12 digits"""
    try:
        return str(idpel).strip().zfill(12)
    except Exception as e:
        logger.error(f"Error normalizing IDPEL {idpel}: {e}")
        return str(idpel)

def normalize_blth(blth):
    """Normalize BLTH to YYYYMM format"""
    try:
        blth_str = str(blth).replace('-', '').replace('/', '').strip()
        # Handle different formats
        if len(blth_str) >= 6:
            return blth_str[:6]
        logger.warning(f"BLTH format invalid: {blth}")
        return blth_str
    except Exception as e:
        logger.error(f"Error normalizing BLTH {blth}: {e}")
        return str(blth)

def get_previous_blth(blth, months_back=1):
    """Get previous BLTH"""
    try:
        date = datetime.strptime(str(blth), '%Y%m')
        prev_date = date - relativedelta(months=months_back)
        return prev_date.strftime('%Y%m')
    except Exception as e:
        logger.error(f"Error getting previous BLTH for {blth}: {e}")
        return blth


# =================== HELPER: Normalisasi UNITUP ===================
def normalize_unitup(unitup):
    """Normalisasi UNITUP: hapus whitespace dan suffix _up3"""
    if unitup is None:
        return None
    return str(unitup).strip().replace('_up3', '')

def normalize_blth(blth):
    """Normalisasi BLTH: YYYYMM format"""
    if blth is None:
        return datetime.now().strftime('%Y%m')
    return str(blth).replace('-', '').strip()


def log_dataframe_info(df, name="DataFrame"):
    """Log DataFrame info for debugging"""
    logger.info(f"\n{'='*60}")
    logger.info(f"📊 {name} Info:")
    logger.info(f"{'='*60}")
    logger.info(f"Shape: {df.shape}")
    logger.info(f"Columns: {df.columns.tolist()}")
    logger.info(f"Dtypes:\n{df.dtypes}")
    logger.info(f"\nFirst 3 rows:\n{df.head(3)}")
    logger.info(f"\nNull values:\n{df.isnull().sum()}")
    logger.info(f"{'='*60}\n")
    
# =================== PROCESS DPM UPLOAD ===================
def process_dpm_upload(df_upload, blth_kini, unitup_session):
    """📥 Process uploaded DPM file and save to database"""
    try:
        username = session.get('username', '').lower()
        role = session.get('role', 'ULP')

        if not username:
            raise ValueError("Session tidak memiliki username aktif")

        logger.info(f"🚀 Processing DPM upload - BLTH: {blth_kini}, USER: {username} ({role})")
        
        # Standardize columns
        df_upload.columns = [c.strip().upper() for c in df_upload.columns]
        log_dataframe_info(df_upload, "DPM Raw Upload")
        
        # Required columns
        required_cols = ['IDPEL', 'LWBPPAKAI']
        missing = [col for col in required_cols if col not in df_upload.columns]
        if missing:
            raise ValueError(f"Kolom wajib tidak ditemukan: {missing}")
        
        # ✅ Handling UNITUP
        if role == 'UP3':
            # UP3 upload: UNITUP sudah diberi suffix _up3 di dashboard_ulp()
            if 'UNITUP' not in df_upload.columns:
                raise ValueError("❌ Upload oleh UP3 harus menyertakan kolom UNITUP!")
            if df_upload['UNITUP'].isnull().any():
                raise ValueError("❌ Ada baris dengan UNITUP kosong di Excel!")
            logger.info(f"✅ UP3 upload - UNITUP values: {df_upload['UNITUP'].unique().tolist()}")
        
        else:
            # ULP upload: gunakan unitup_session
            df_upload['UNITUP'] = unitup_session
            logger.info(f"✅ ULP upload - UNITUP from session: {unitup_session}")
        
        # Add metadata
        df_upload['BLTH'] = normalize_blth(blth_kini)
        df_upload['IDPEL'] = df_upload['IDPEL'].apply(normalize_idpel)
        df_upload['UPDATED_BY'] = username
        
        # Convert numeric columns
        numeric_cols = ['DAYA', 'SLALWBP', 'LWBPCABUT', 'LWBPPASANG', 'SAHLWBP', 'LWBPPAKAI']
        for col in numeric_cols:
            if col in df_upload.columns:
                df_upload[col] = pd.to_numeric(df_upload[col], errors='coerce').fillna(0).astype(int)
        
        # Select only columns yang ada di DB
        db_cols = ['BLTH', 'UNITUP', 'IDPEL', 'NAMA', 'TARIF', 'DAYA',
                   'SLALWBP', 'LWBPCABUT', 'LWBPPASANG', 'SAHLWBP', 
                   'LWBPPAKAI', 'DLPD', 'KDKELOMPOK', 'UPDATED_BY']
        df_final = df_upload[[c for c in db_cols if c in df_upload.columns]]
        
        # Log summary
        unitup_summary = df_final.groupby('UNITUP').size()
        logger.info("📊 Data summary per UNITUP:")
        for unit, count in unitup_summary.items():
            logger.info(f"   {unit}: {count} rows")
        
        # Save ke DB
        save_dpm_with_upsert(df_final, engine)
        logger.info(f"✅ DPM upload successful: {len(df_final)} rows by {username}")
        return len(df_final), None

    except Exception as e:
        error_msg = f"Error processing DPM: {str(e)}"
        logger.error(error_msg)
        logger.error(traceback.format_exc())
        return 0, error_msg



def save_dpm_with_upsert(df, engine):
    """💾 Save DPM data dengan INSERT ... ON DUPLICATE KEY UPDATE per user"""
    records = df.to_dict('records')
    
    with engine.begin() as conn:
        for record in records:
            # ✅ Validasi data
            if not record.get('UPDATED_BY'):
                logger.error(f"⚠️ UPDATED_BY kosong untuk IDPEL {record.get('IDPEL')}")
                continue
            
            if not record.get('UNITUP'):
                logger.error(f"⚠️ UNITUP kosong untuk IDPEL {record.get('IDPEL')}")
                continue
            
            cols = ', '.join(record.keys())
            placeholders = ', '.join([f":{k}" for k in record.keys()])
            
            # Primary key: BLTH, IDPEL, UPDATED_BY
            update_cols = [k for k in record.keys() 
                          if k not in ['BLTH', 'IDPEL', 'UPDATED_BY']]
            
            if update_cols:
                updates = ', '.join([f"{k}=VALUES({k})" for k in update_cols])
                
                sql = text(f"""
                    INSERT INTO dpm ({cols})
                    VALUES ({placeholders})
                    ON DUPLICATE KEY UPDATE 
                        {updates}
                """)
            else:
                sql = text(f"""
                    INSERT IGNORE INTO dpm ({cols})
                    VALUES ({placeholders})
                """)
            
            try:
                conn.execute(sql, record)
                logger.debug(f"✅ Saved: UNITUP={record['UNITUP']}, IDPEL={record['IDPEL']}, BY={record['UPDATED_BY']}")
            except Exception as e:
                logger.error(f"❌ Failed to save {record['IDPEL']}: {e}")
                continue


def process_dpm(df):
    """
    Proses dataframe untuk format DPM table
    Fungsi ini memvalidasi dan membersihkan data sebelum disimpan ke DPM
    """
    try:
        if df.empty:
            logger.warning("process_dpm: DataFrame kosong")
            return pd.DataFrame()
        
        # Pastikan kolom yang diperlukan ada
        required_cols = ['BLTH', 'IDPEL', 'LWBPPAKAI']
        for col in required_cols:
            if col not in df.columns:
                logger.error(f"process_dpm: Kolom {col} tidak ditemukan")
                df[col] = 0 if col == 'LWBPPAKAI' else ''
        
        # Buat copy untuk menghindari SettingWithCopyWarning
        df_clean = df[required_cols].copy()
        
        # Normalisasi IDPEL
        df_clean['IDPEL'] = df_clean['IDPEL'].astype(str).str.strip().str.lower()
        
        # Pastikan LWBPPAKAI numerik dan tidak negatif
        df_clean['LWBPPAKAI'] = pd.to_numeric(df_clean['LWBPPAKAI'], errors='coerce').fillna(0)
        df_clean['LWBPPAKAI'] = df_clean['LWBPPAKAI'].clip(lower=0)  # Tidak boleh negatif
        
        # Pastikan BLTH dalam format yang benar
        df_clean['BLTH'] = df_clean['BLTH'].astype(str)
        
        logger.info(f"process_dpm: Berhasil memproses {len(df_clean)} records")
        return df_clean
    
    except Exception as e:
        logger.error(f"Error in process_dpm: {str(e)}")
        logger.error(traceback.format_exc())
        return pd.DataFrame()


def update_dpm_table(df, dpm_table):
    """
    Update DPM table dengan data dari dataframe
    Menggunakan INSERT ... ON DUPLICATE KEY UPDATE untuk handle insert/update
    """
    try:
        if df.empty:
            logger.warning("update_dpm_table: DataFrame kosong")
            return {'inserted_or_updated': 0, 'message': 'DataFrame kosong'}
        
        inserted_or_updated = 0
        failed = 0
        
        # Validasi nama tabel untuk keamanan
        if not dpm_table or not isinstance(dpm_table, str):
            logger.error(f"update_dpm_table: Nama tabel tidak valid: {dpm_table}")
            return {'inserted_or_updated': 0, 'message': 'Nama tabel tidak valid'}
        
        with engine.begin() as conn:
            for _, row in df.iterrows():
                try:
                    sql = text(f"""
                        INSERT INTO {dpm_table} (BLTH, IDPEL, LWBPPAKAI)
                        VALUES (:blth, :idpel, :lwbppakai)
                        ON DUPLICATE KEY UPDATE 
                            LWBPPAKAI = VALUES(LWBPPAKAI)
                    """)
                    
                    conn.execute(sql, {
                        'blth': row['BLTH'],
                        'idpel': row['IDPEL'],
                        'lwbppakai': int(row['LWBPPAKAI'])
                    })
                    inserted_or_updated += 1
                    
                except Exception as e:
                    failed += 1
                    if failed <= 5:  # Log hanya 5 error pertama untuk menghindari spam
                        logger.error(f"Error updating IDPEL {row['IDPEL']}: {e}")
                    continue
        
        logger.info(f"update_dpm_table: Berhasil {inserted_or_updated}, Gagal {failed}")
        
        return {
            'inserted_or_updated': inserted_or_updated,
            'failed': failed,
            'message': f'Berhasil update {inserted_or_updated} records, gagal {failed} records'
        }
    
    except Exception as e:
        error_msg = f"Error in update_dpm_table: {str(e)}"
        logger.error(error_msg)
        logger.error(traceback.format_exc())
        return {'inserted_or_updated': 0, 'failed': len(df), 'message': error_msg}


# #dataframe utama
# def process_billing_advanced(blth_kini, unitup, engine):
#     """🔄 Process billing dengan perhitungan KWH antar bulan otomatis"""
#     try:
#         logger.info(f"🚀 Processing billing for UNITUP: {unitup}, BLTH: {blth_kini}")

#         # --- Normalisasi Periode ---
#         blth_kini = normalize_blth(blth_kini)
#         blth_lalu = get_previous_blth(blth_kini, 1)
#         blth_lalulalu = get_previous_blth(blth_kini, 2)
#         blth_lalu3 = get_previous_blth(blth_kini, 3)

#         # --- Ambil Data dari DPM ---
#         # ✅ Handle ULP filter (UNITUP spesifik)
#         if unitup:
#             query = text("""
#                 SELECT * FROM dpm 
#                 WHERE UNITUP = :unitup 
#                 AND BLTH IN (:kini, :lalu, :lalulalu, :lalu3)
#                 ORDER BY IDPEL
#             """)
#             df_all = pd.read_sql(query, engine, params={
#                 'unitup': unitup,
#                 'kini': blth_kini,
#                 'lalu': blth_lalu,
#                 'lalulalu': blth_lalulalu,
#                 'lalu3': blth_lalu3
#             })
#             logger.info(f"✅ ULP mode: Found {len(df_all)} records for UNITUP {unitup}")
        
#         # ✅ Fallback: ambil semua (bila diperlukan)
#         else:
#             query = text("""
#                 SELECT * FROM dpm 
#                 WHERE BLTH IN (:kini, :lalu, :lalulalu, :lalu3)
#                 ORDER BY UNITUP, IDPEL
#             """)
#             df_all = pd.read_sql(query, engine, params={
#                 'kini': blth_kini,
#                 'lalu': blth_lalu,
#                 'lalulalu': blth_lalulalu,
#                 'lalu3': blth_lalu3
#             })
#             logger.info(f"✅ All mode: Found {len(df_all)} records")

#         if df_all.empty:
#             return pd.DataFrame(), "Tidak ada data DPM untuk periode ini"

#         # ✅ BERSIHKAN karakter \n dari data source SEBELUM proses
#         text_cols_source = ['NAMA', 'TARIF', 'KDKELOMPOK', 'DLPD']
#         for col in text_cols_source:
#             if col in df_all.columns:
#                 df_all[col] = df_all[col].apply(
#                     lambda x: str(x).replace('\\n', ' ').replace('\n', ' ').strip() 
#                     if pd.notna(x) and x != '' else x
#                 )
#         logger.info("✅ Cleaned \\n from source data")

#         # --- Kolom Wajib ---
#         kolom_wajib = ['IDPEL','NAMA','TARIF','DAYA','KDKELOMPOK','SLALWBP','LWBPCABUT','LWBPPASANG','SAHLWBP','LWBPPAKAI','DLPD']
#         for kolom in kolom_wajib:
#             if kolom not in df_all.columns:
#                 df_all[kolom] = 0 if kolom != 'DLPD' else ''

#         # --- Pisah Data per Bulan ---
#         df_kini = df_all[df_all['BLTH'] == blth_kini].copy()
#         df_lalu = df_all[df_all['BLTH'] == blth_lalu][['IDPEL','LWBPPAKAI']].rename(columns={'LWBPPAKAI':'LWBPPAKAI_Y'})
#         df_lalulalu = df_all[df_all['BLTH'] == blth_lalulalu][['IDPEL','LWBPPAKAI']].rename(columns={'LWBPPAKAI':'LWBPPAKAI_X'})
#         df_lalu3 = df_all[df_all['BLTH'] == blth_lalu3][['IDPEL','LWBPPAKAI']].rename(columns={'LWBPPAKAI':'LWBPPAKAI_Z'})

#         # --- 🔹 PENGECEKAN DUPLIKASI ---
#         duplikat_lalu3 = df_lalu3['IDPEL'].duplicated().sum()
#         duplikat_lalulalu = df_lalulalu['IDPEL'].duplicated().sum()
#         duplikat_lalu = df_lalu['IDPEL'].duplicated().sum()
#         duplikat_kini = df_kini['IDPEL'].duplicated().sum()

#         logger.info(f"Duplikat di df_lalu3: {duplikat_lalu3}")
#         logger.info(f"Duplikat di df_lalulalu: {duplikat_lalulalu}")
#         logger.info(f"Duplikat di df_lalu: {duplikat_lalu}")
#         logger.info(f"Duplikat di df_kini: {duplikat_kini}")

#         if duplikat_lalu3 > 0:
#             flash(f"⚠️ Ditemukan {duplikat_lalu3} IDPEL duplikat di DPM N - 3.", "warning")
#         if duplikat_lalulalu > 0:
#             flash(f"⚠️ Ditemukan {duplikat_lalulalu} IDPEL duplikat di DPM N - 2.", "warning")
#         if duplikat_lalu > 0:
#             flash(f"⚠️ Ditemukan {duplikat_lalu} IDPEL duplikat di DPM N - 1.", "warning")
#         if duplikat_kini > 0:
#             flash(f"⚠️ Ditemukan {duplikat_kini} IDPEL duplikat di DPM bulan N.", "warning")

#         # --- Merge Semua Periode ---
#         df_merged = (
#             df_kini
#             .merge(df_lalu, on='IDPEL', how='left')
#             .merge(df_lalulalu, on='IDPEL', how='left')
#             .merge(df_lalu3, on='IDPEL', how='left')
#         )

#         logger.info(f"Duplikat setelah merge: {df_merged['IDPEL'].duplicated().sum()}")

#         # --- 🔹 Pastikan semua kolom input ada ---
#         for kol in ['SLALWBP', 'LWBPCABUT', 'SAHLWBP', 'LWBPPASANG', 'LWBPPAKAI']:
#             if kol not in df_merged.columns:
#                 df_merged[kol] = 0

#         # --- 🔹 Konversi ke numerik ---
#         for kol in ['SLALWBP', 'LWBPCABUT', 'SAHLWBP', 'LWBPPASANG', 'LWBPPAKAI']:
#             df_merged[kol] = pd.to_numeric(df_merged[kol], errors='coerce')

#         # --- 🔹 PERHITUNGAN LWBPPAKAI - Hanya jika kosong ---
#         lwbp_kosong = df_merged['LWBPPAKAI'].isna() | (df_merged['LWBPPAKAI'] == 0)
#         count_replaced = lwbp_kosong.sum()
#         logger.info(f"Jumlah data LWBPPAKAI yang dihitung ulang: {count_replaced}")

#         # Isi nilai hanya kalau kosong
#         df_merged['LWBPPAKAI'] = np.where(
#             lwbp_kosong,
#             (df_merged['LWBPCABUT'].fillna(0)
#             - df_merged['SLALWBP'].fillna(0)
#             + df_merged['SAHLWBP'].fillna(0)
#             - df_merged['LWBPPASANG'].fillna(0)),
#             df_merged['LWBPPAKAI']
#         )

#         # --- Siapkan Kolom KWH ---
#         df_merged['KWH_SEKARANG'] = df_merged['LWBPPAKAI'].fillna(0).astype(int)
#         df_merged['KWH_1_BULAN_LALU'] = df_merged['LWBPPAKAI_Y'].fillna(0).astype(int)
#         df_merged['KWH_2_BULAN_LALU'] = df_merged['LWBPPAKAI_X'].fillna(0).astype(int)

#         # --- Delta & Persen ---
#         delta = df_merged['KWH_SEKARANG'] - df_merged['KWH_1_BULAN_LALU']
#         with np.errstate(divide='ignore', invalid='ignore'):
#             percentage = (delta / df_merged['KWH_1_BULAN_LALU'].replace(0, np.nan)) * 100
#             percentage = np.nan_to_num(percentage, nan=0)

#         # --- Hitung Jam Nyala ---
#         daya_kw = df_merged['DAYA'].replace(0, np.nan) / 1000
#         jam_nyala = (df_merged['KWH_SEKARANG'] / daya_kw).replace([np.inf, -np.inf], 0).fillna(0)

#         # --- Hitung Rata-rata 3 Bulan ---
#         rerata = df_merged[['LWBPPAKAI_Y','LWBPPAKAI_X','LWBPPAKAI_Z']].fillna(0).mean(axis=1)

#         # --- 🔹 KLASIFIKASI DLPD_HITUNG ---
#         # Kondisi stan mundur
#         stan_mundur_condition = (
#             (df_merged['SAHLWBP'] < df_merged['SLALWBP']) &
#             (df_merged['LWBPCABUT'].fillna(0) == 0) &
#             (df_merged['LWBPPASANG'].fillna(0) == 0)
#         )

#         # Kondisi cek pecahan (ada nilai cabut atau pasang)
#         cek_pecahan_condition = (
#             (df_merged['LWBPCABUT'].fillna(0) != 0) |
#             (df_merged['LWBPPASANG'].fillna(0) != 0)
#         )

#         sortir_naik = 40
#         sortir_turun = 40
#         is_na = df_merged['LWBPPAKAI_Y'].isna() | (df_merged['LWBPPAKAI_Y'] == 0)
#         is_naik = (~is_na) & (percentage >= sortir_naik)
#         is_turun = (~is_na) & (percentage <= -sortir_turun)
        
#         # Daftar kondisi dan klasifikasi
#         conditions = [
#             (jam_nyala >= 720),
#             cek_pecahan_condition,
#             stan_mundur_condition,
#             (percentage > 50),
#             (is_na & (jam_nyala > 40)), 
#             (percentage < -50),
#             (df_merged['LWBPPAKAI'] == 0),
#             (jam_nyala > 0) & (jam_nyala < 40)
#         ]

#         choices = [
#             'JN>720',
#             'PECAHAN',
#             'STAN MUNDUR',
#             'NAIK>50%',
#             'DIV/NA',
#             'TURUN<50%',
#             'KWH NOL',
#             'JN<40'
#         ]
        
#         # Terapkan ke kolom baru
#         df_merged['DLPD_HITUNG'] = np.select(conditions, choices, default='')

#         # --- Hitung DLPD 3BLN ---
#         df_merged['DLPD_3BLN'] = np.where(
#             (df_merged['KWH_SEKARANG'] > 1.5 * rerata),
#             'Naik50% R3BLN',
#             'Turun=50% R3BLN'
#         )

#         # --- Klasifikasi KET ---
#         is_na = df_merged['KWH_1_BULAN_LALU'] == 0
#         is_naik = (~is_na) & (percentage >= 40)
#         is_turun = (~is_na) & (percentage <= -40)
#         ket = np.select([is_na, is_naik, is_turun], ['DIV/NA','NAIK','TURUN'], default='AMAN')

#         # --- Tambahkan JAMNYALA600 ---
#         df_merged['JAMNYALA600'] = np.select(
#             [jam_nyala > 600, jam_nyala <= 600],
#             ['600Up', '600Down'],
#             default='UNKNOWN'
#         )

#         # --- Ambil UNITUP dari df_merged ---
#         df_merged['UNITUP'] = df_merged['UNITUP'].fillna(unitup if unitup else 'UNKNOWN')

#         # ✅ BERSIHKAN IDPEL dari spasi/karakter aneh
#         df_merged['IDPEL'] = df_merged['IDPEL'].astype(str).str.strip().str.lower()

#         # --- Bangun DataFrame Akhir (LANGSUNG PAKAI UNDERSCORE) ---
#         kroscek = pd.DataFrame({
#             'BLTH': blth_kini,
#             'UNITUP': df_merged['UNITUP'],
#             'IDPEL': df_merged['IDPEL'],
#             'NAMA': df_merged['NAMA'],
#             'TARIF': df_merged['TARIF'],
#             'DAYA': df_merged['DAYA'].fillna(0).astype(int),
#             'KDKELOMPOK': df_merged['KDKELOMPOK'].fillna(''),
#             'SLALWBP': df_merged['SLALWBP'].fillna(0).astype(int),
#             'LWBPCABUT': df_merged['LWBPCABUT'].fillna(0).astype(int),
#             'SELISIH_STAN_BONGKAR': (df_merged['SLALWBP'] - df_merged['LWBPCABUT']).fillna(0).astype(int),
#             'LWBPPASANG': df_merged['LWBPPASANG'].fillna(0).astype(int),
#             'SAHLWBP': df_merged['SAHLWBP'].fillna(0).astype(int),
#             'KWH_SEKARANG': df_merged['KWH_SEKARANG'],
#             'KWH_1_BULAN_LALU': df_merged['KWH_1_BULAN_LALU'],
#             'KWH_2_BULAN_LALU': df_merged['KWH_2_BULAN_LALU'],
#             'DELTA_PEMKWH': delta.fillna(0).astype(int),
#             'PERSEN': pd.Series(percentage).round(1).astype(str) + '%',
#             'KET': ket,
#             'JAM_NYALA': jam_nyala.round(1),
#             'JAMNYALA600': df_merged['JAMNYALA600'],
#             'DLPD': df_merged['DLPD'].fillna(''),
#             'DLPD_HITUNG': df_merged['DLPD_HITUNG'],
#             'DLPD_3BLN': df_merged['DLPD_3BLN'],
#             'NOMORKWH': '',
#             'HASIL_PEMERIKSAAN': '',
#             'TINDAK_LANJUT': '',
#             'STAN_VERIFIKASI': ''
#         })

#         # --- Hapus duplikasi ---
#         kroscek = kroscek.drop_duplicates(subset='IDPEL', keep='first')
#         logger.info(f"✅ Removed duplicates, remaining: {len(kroscek)} records")

#         # ✅ BERSIHKAN semua kolom text dari karakter \n
#         text_columns = [
#             'NAMA', 'TARIF', 'KDKELOMPOK', 'DLPD', 'DLPD_HITUNG', 
#             'DLPD_3BLN', 'KET', 'JAMNYALA600', 'NOMORKWH', 
#             'HASIL_PEMERIKSAAN', 'TINDAK_LANJUT', 'STAN_VERIFIKASI'
#         ]

#         for col in text_columns:
#             if col in kroscek.columns:
#                 kroscek[col] = kroscek[col].apply(
#                     lambda x: str(x).replace('\\n', ' ').replace('\n', ' ').replace('\r', ' ').strip() 
#                     if pd.notna(x) and str(x) != 'nan' and str(x) != '' else ''
#                 )

#         logger.info(f"✅ Text columns cleaned from \\n, \\r characters")

#         # --- Foto & Grafik ---
#         path_foto1 = "https://portalapp.iconpln.co.id/acmt/DisplayBlobServlet1?idpel="
#         path_foto2 = "&blth="

#         # ✅ FOTO dengan button (SAFE VERSION)
#         kroscek['FOTO_AKHIR'] = kroscek['IDPEL'].apply(
#             lambda x: f"""<button class='btn btn-sm btn-success' 
#                 onclick="window.open('{path_foto1}{x}{path_foto2}{blth_kini}', '_blank', 
#                 'width=700,height=700,scrollbars=yes,toolbar=no'); return false;">
#                 <i class='bi bi-image'></i> Akhir
#             </button>"""
#         )

#         kroscek['FOTO_LALU'] = kroscek['IDPEL'].apply(
#             lambda x: f"""<button class='btn btn-sm btn-warning' 
#                 onclick="window.open('{path_foto1}{x}{path_foto2}{blth_lalu}', '_blank', 
#                 'width=700,height=700,scrollbars=yes,toolbar=no'); return false;">
#                 <i class='bi bi-image'></i> Lalu
#             </button>"""
#         )

#         kroscek['FOTO_LALU2'] = kroscek['IDPEL'].apply(
#             lambda x: f"""<button class='btn btn-sm btn-secondary' 
#                 onclick="window.open('{path_foto1}{x}{path_foto2}{blth_lalulalu}', '_blank', 
#                 'width=700,height=700,scrollbars=yes,toolbar=no'); return false;">
#                 <i class='bi bi-image'></i> Lalu2
#             </button>"""
#         )

#         kroscek['FOTO_3BLN'] = kroscek.apply(
#             lambda row: f"""<button class='btn btn-sm btn-primary foto-3bln-link-btn' 
#                 data-idpel="{row['IDPEL']}"
#                 data-sahlwbp="{row['SAHLWBP']}"
#                 onclick="open3Foto('{row['IDPEL']}', '{blth_kini}'); return false;">
#                 <i class='bi bi-images'></i> 3 Foto
#             </button>""",
#             axis=1
#         )
        

#         # Grafik link - Ubah jadi button dengan modal
#         kroscek['GRAFIK'] = kroscek.apply(
#             lambda row: f'''<button class="btn btn-sm btn-info grafik-btn" 
#                 data-idpel="{row["IDPEL"]}" 
#                 data-blth="{blth_kini}" 
#                 data-unitup="{row["UNITUP"]}">
#                 <i class="bi bi-graph-up"></i> Grafik
#             </button>''',
#             axis=1
# )

#         # ✅ FINAL VALIDATION: Pastikan tidak ada NaN yang jadi string 'nan'
#         kroscek = kroscek.replace('nan', '')
#         kroscek = kroscek.fillna('')

#         logger.info(f"✅ Billing processed successfully: {len(kroscek)} records")
        
#         return kroscek, None

#     except Exception as e:
#         error_msg = f"Error processing billing: {str(e)}"
#         logger.error(error_msg)
#         logger.error(traceback.format_exc())
#         return pd.DataFrame(), error_msg
def process_billing_advanced(blth_kini, unitup, engine):
    """🔄 Process billing dengan perhitungan KWH antar bulan otomatis"""
    try:
        logger.info(f"🚀 Processing billing for UNITUP: {unitup}, BLTH: {blth_kini}")

        # --- Normalisasi Periode ---
        blth_kini = normalize_blth(blth_kini)
        blth_lalu = get_previous_blth(blth_kini, 1)
        blth_lalulalu = get_previous_blth(blth_kini, 2)
        blth_lalu3 = get_previous_blth(blth_kini, 3)

        # --- Ambil Data dari DPM (TAMBAHKAN MARKING_KOREKSI) ---
        if unitup:
            query = text("""
                SELECT * FROM dpm 
                WHERE UNITUP = :unitup 
                AND BLTH IN (:kini, :lalu, :lalulalu, :lalu3)
                ORDER BY IDPEL
            """)
            df_all = pd.read_sql(query, engine, params={
                'unitup': unitup,
                'kini': blth_kini,
                'lalu': blth_lalu,
                'lalulalu': blth_lalulalu,
                'lalu3': blth_lalu3
            })
            logger.info(f"✅ ULP mode: Found {len(df_all)} records for UNITUP {unitup}")
        else:
            query = text("""
                SELECT * FROM dpm 
                WHERE BLTH IN (:kini, :lalu, :lalulalu, :lalu3)
                ORDER BY UNITUP, IDPEL
            """)
            df_all = pd.read_sql(query, engine, params={
                'kini': blth_kini,
                'lalu': blth_lalu,
                'lalulalu': blth_lalulalu,
                'lalu3': blth_lalu3
            })
            logger.info(f"✅ All mode: Found {len(df_all)} records")

        if df_all.empty:
            return pd.DataFrame(), "Tidak ada data DPM untuk periode ini"

        # ✅ BERSIHKAN karakter \n dari data source SEBELUM proses
        text_cols_source = ['NAMA', 'TARIF', 'KDKELOMPOK', 'DLPD']
        for col in text_cols_source:
            if col in df_all.columns:
                df_all[col] = df_all[col].apply(
                    lambda x: str(x).replace('\\n', ' ').replace('\n', ' ').strip() 
                    if pd.notna(x) and x != '' else x
                )
        logger.info("✅ Cleaned \\n from source data")

        # --- Kolom Wajib (TAMBAHKAN MARKING_KOREKSI) ---
        kolom_wajib = ['IDPEL','NAMA','TARIF','DAYA','KDKELOMPOK','SLALWBP','LWBPCABUT',
                       'LWBPPASANG','SAHLWBP','LWBPPAKAI','DLPD','MARKING_KOREKSI']
        for kolom in kolom_wajib:
            if kolom not in df_all.columns:
                if kolom == 'MARKING_KOREKSI':
                    df_all[kolom] = 0  # Default 0 untuk MARKING_KOREKSI
                elif kolom == 'DLPD':
                    df_all[kolom] = ''
                else:
                    df_all[kolom] = 0

        # --- Pisah Data per Bulan ---
        df_kini = df_all[df_all['BLTH'] == blth_kini].copy()
        df_lalu = df_all[df_all['BLTH'] == blth_lalu][['IDPEL','LWBPPAKAI']].rename(columns={'LWBPPAKAI':'LWBPPAKAI_Y'})
        df_lalulalu = df_all[df_all['BLTH'] == blth_lalulalu][['IDPEL','LWBPPAKAI']].rename(columns={'LWBPPAKAI':'LWBPPAKAI_X'})
        df_lalu3 = df_all[df_all['BLTH'] == blth_lalu3][['IDPEL','LWBPPAKAI']].rename(columns={'LWBPPAKAI':'LWBPPAKAI_Z'})

        # --- Merge Semua Periode ---
        df_merged = (
            df_kini
            .merge(df_lalu, on='IDPEL', how='left')
            .merge(df_lalulalu, on='IDPEL', how='left')
            .merge(df_lalu3, on='IDPEL', how='left')
        )

        logger.info(f"Duplikat setelah merge: {df_merged['IDPEL'].duplicated().sum()}")

        # --- Pastikan semua kolom input ada ---
        for kol in ['SLALWBP', 'LWBPCABUT', 'SAHLWBP', 'LWBPPASANG', 'LWBPPAKAI']:
            if kol not in df_merged.columns:
                df_merged[kol] = 0

        # --- Konversi ke numerik ---
        for kol in ['SLALWBP', 'LWBPCABUT', 'SAHLWBP', 'LWBPPASANG', 'LWBPPAKAI', 'MARKING_KOREKSI']:
            df_merged[kol] = pd.to_numeric(df_merged[kol], errors='coerce').fillna(0)

        # --- PERHITUNGAN LWBPPAKAI - Hanya jika kosong ---
        lwbp_kosong = df_merged['LWBPPAKAI'].isna() | (df_merged['LWBPPAKAI'] == 0)
        count_replaced = lwbp_kosong.sum()
        logger.info(f"Jumlah data LWBPPAKAI yang dihitung ulang: {count_replaced}")

        df_merged['LWBPPAKAI'] = np.where(
            lwbp_kosong,
            (df_merged['LWBPCABUT'].fillna(0)
            - df_merged['SLALWBP'].fillna(0)
            + df_merged['SAHLWBP'].fillna(0)
            - df_merged['LWBPPASANG'].fillna(0)),
            df_merged['LWBPPAKAI']
        )

        # --- Siapkan Kolom KWH ---
        df_merged['KWH_SEKARANG'] = df_merged['LWBPPAKAI'].fillna(0).astype(int)
        df_merged['KWH_1_BULAN_LALU'] = df_merged['LWBPPAKAI_Y'].fillna(0).astype(int)
        df_merged['KWH_2_BULAN_LALU'] = df_merged['LWBPPAKAI_X'].fillna(0).astype(int)

        # --- Delta & Persen ---
        delta = df_merged['KWH_SEKARANG'] - df_merged['KWH_1_BULAN_LALU']
        with np.errstate(divide='ignore', invalid='ignore'):
            percentage = (delta / df_merged['KWH_1_BULAN_LALU'].replace(0, np.nan)) * 100
            percentage = np.nan_to_num(percentage, nan=0)

        # --- Hitung Jam Nyala ---
        daya_kw = df_merged['DAYA'].replace(0, np.nan) / 1000
        jam_nyala = (df_merged['KWH_SEKARANG'] / daya_kw).replace([np.inf, -np.inf], 0).fillna(0)

        # --- Hitung Rata-rata 3 Bulan ---
        rerata = df_merged[['LWBPPAKAI_Y','LWBPPAKAI_X','LWBPPAKAI_Z']].fillna(0).mean(axis=1)

        # --- KLASIFIKASI DLPD_HITUNG ---
        stan_mundur_condition = (
            (df_merged['SAHLWBP'] < df_merged['SLALWBP']) &
            (df_merged['LWBPCABUT'].fillna(0) == 0) &
            (df_merged['LWBPPASANG'].fillna(0) == 0)
        )

        cek_pecahan_condition = (
            (df_merged['LWBPCABUT'].fillna(0) != 0) |
            (df_merged['LWBPPASANG'].fillna(0) != 0)
        )

        sortir_naik = 40
        sortir_turun = 40
        is_na = df_merged['LWBPPAKAI_Y'].isna() | (df_merged['LWBPPAKAI_Y'] == 0)
        is_naik = (~is_na) & (percentage >= sortir_naik)
        is_turun = (~is_na) & (percentage <= -sortir_turun)
        
        conditions = [
            (jam_nyala >= 720),
            cek_pecahan_condition,
            stan_mundur_condition,
            (percentage > 50),
            (is_na & (jam_nyala > 40)), 
            (percentage < -50),
            (df_merged['LWBPPAKAI'] == 0),
            (jam_nyala > 0) & (jam_nyala < 40)
        ]

        choices = [
            'JN>720',
            'PECAHAN',
            'STAN MUNDUR',
            'NAIK>50%',
            'DIV/NA',
            'TURUN<50%',
            'KWH NOL',
            'JN<40'
        ]
        
        df_merged['DLPD_HITUNG'] = np.select(conditions, choices, default='')

        # --- Hitung DLPD 3BLN ---
        df_merged['DLPD_3BLN'] = np.where(
            (df_merged['KWH_SEKARANG'] > 1.5 * rerata),
            'Naik50% R3BLN',
            'Turun=50% R3BLN'
        )

        # --- Klasifikasi KET ---
        is_na = df_merged['KWH_1_BULAN_LALU'] == 0
        is_naik = (~is_na) & (percentage >= 40)
        is_turun = (~is_na) & (percentage <= -40)
        ket = np.select([is_na, is_naik, is_turun], ['DIV/NA','NAIK','TURUN'], default='AMAN')

        # --- Tambahkan JAMNYALA600 ---
        df_merged['JAMNYALA600'] = np.select(
            [jam_nyala > 600, jam_nyala <= 600],
            ['600Up', '600Down'],
            default='UNKNOWN'
        )

        # --- Ambil UNITUP dari df_merged ---
        df_merged['UNITUP'] = df_merged['UNITUP'].fillna(unitup if unitup else 'UNKNOWN')

        # ✅ BERSIHKAN IDPEL dari spasi/karakter aneh
        df_merged['IDPEL'] = df_merged['IDPEL'].astype(str).str.strip().str.lower()

        # --- Bangun DataFrame Akhir (TAMBAHKAN MARKING_KOREKSI) ---
        kroscek = pd.DataFrame({
            'BLTH': blth_kini,
            'UNITUP': df_merged['UNITUP'],
            'IDPEL': df_merged['IDPEL'],
            'NAMA': df_merged['NAMA'],
            'TARIF': df_merged['TARIF'],
            'DAYA': df_merged['DAYA'].fillna(0).astype(int),
            'KDKELOMPOK': df_merged['KDKELOMPOK'].fillna(''),
            'SLALWBP': df_merged['SLALWBP'].fillna(0).astype(int),
            'LWBPCABUT': df_merged['LWBPCABUT'].fillna(0).astype(int),
            'SELISIH_STAN_BONGKAR': (df_merged['SLALWBP'] - df_merged['LWBPCABUT']).fillna(0).astype(int),
            'LWBPPASANG': df_merged['LWBPPASANG'].fillna(0).astype(int),
            'SAHLWBP': df_merged['SAHLWBP'].fillna(0).astype(int),
            'KWH_SEKARANG': df_merged['KWH_SEKARANG'],
            'KWH_1_BULAN_LALU': df_merged['KWH_1_BULAN_LALU'],
            'KWH_2_BULAN_LALU': df_merged['KWH_2_BULAN_LALU'],
            'DELTA_PEMKWH': delta.fillna(0).astype(int),
            'PERSEN': pd.Series(percentage).round(1).astype(str) + '%',
            'KET': ket,
            'JAM_NYALA': jam_nyala.round(1),
            'JAMNYALA600': df_merged['JAMNYALA600'],
            'DLPD': df_merged['DLPD'].fillna(''),
            'DLPD_HITUNG': df_merged['DLPD_HITUNG'],
            'DLPD_3BLN': df_merged['DLPD_3BLN'],
            'MARKING_KOREKSI': df_merged['MARKING_KOREKSI'].fillna(0).astype(int),  # ✅ TAMBAHAN
            'NOMORKWH': ''
            # 'HASIL_PEMERIKSAAN': '',
            # 'TINDAK_LANJUT': '',
            # 'STAN_VERIFIKASI': ''
        })

        # --- Hapus duplikasi ---
        kroscek = kroscek.drop_duplicates(subset='IDPEL', keep='first')
        logger.info(f"✅ Removed duplicates, remaining: {len(kroscek)} records")

        # ✅ BERSIHKAN semua kolom text dari karakter \n
        text_columns = [
            'NAMA', 'TARIF', 'KDKELOMPOK', 'DLPD', 'DLPD_HITUNG', 
            'DLPD_3BLN', 'KET', 'JAMNYALA600', 'NOMORKWH', 
            'HASIL_PEMERIKSAAN', 'TINDAK_LANJUT', 'STAN_VERIFIKASI'
        ]

        for col in text_columns:
            if col in kroscek.columns:
                kroscek[col] = kroscek[col].apply(
                    lambda x: str(x).replace('\\n', ' ').replace('\n', ' ').replace('\r', ' ').strip() 
                    if pd.notna(x) and str(x) != 'nan' and str(x) != '' else ''
                )

        logger.info(f"✅ Text columns cleaned from \\n, \\r characters")

        # --- Foto & Grafik ---
        path_foto1 = "https://portalapp.iconpln.co.id/acmt/DisplayBlobServlet1?idpel="
        path_foto2 = "&blth="

        kroscek['FOTO_AKHIR'] = kroscek['IDPEL'].apply(
            lambda x: f"""<button class='btn btn-sm btn-success' 
                onclick="window.open('{path_foto1}{x}{path_foto2}{blth_kini}', '_blank', 
                'width=700,height=700,scrollbars=yes,toolbar=no'); return false;">
                <i class='bi bi-image'></i> Akhir
            </button>"""
        )

        kroscek['FOTO_LALU'] = kroscek['IDPEL'].apply(
            lambda x: f"""<button class='btn btn-sm btn-warning' 
                onclick="window.open('{path_foto1}{x}{path_foto2}{blth_lalu}', '_blank', 
                'width=700,height=700,scrollbars=yes,toolbar=no'); return false;">
                <i class='bi bi-image'></i> Lalu
            </button>"""
        )

        kroscek['FOTO_LALU2'] = kroscek['IDPEL'].apply(
            lambda x: f"""<button class='btn btn-sm btn-secondary' 
                onclick="window.open('{path_foto1}{x}{path_foto2}{blth_lalulalu}', '_blank', 
                'width=700,height=700,scrollbars=yes,toolbar=no'); return false;">
                <i class='bi bi-image'></i> Lalu2
            </button>"""
        )

        # kroscek['FOTO_3BLN'] = kroscek.apply(
        #     lambda row: f"""<button class='btn btn-sm btn-primary foto-3bln-link-btn' 
        #         data-idpel="{row['IDPEL']}"
        #         data-sahlwbp="{row['SAHLWBP']}"
        #         onclick="open3Foto('{row['IDPEL']}', '{blth_kini}'); return false;">
        #         <i class='bi bi-images'></i> 3 Foto
        #     </button>""",
        #     axis=1
        # )
        
        kroscek['FOTO_3BLN'] = kroscek.apply(
            lambda row: f"""<button class='btn btn-sm btn-primary foto-3bln-link-btn' 
                data-idpel="{row['IDPEL']}"
                data-sahlwbp="{row['SAHLWBP']}"
                onclick="open3Foto('{row['IDPEL']}', '{blth_kini}'); return false;">
                {str(row['IDPEL'])[-5:] if len(str(row['IDPEL'])) >= 5 else str(row['IDPEL'])}
            </button>""",
            axis=1
        )

        kroscek['GRAFIK'] = kroscek.apply(
            lambda row: f'''<button class="btn btn-sm btn-info grafik-btn" 
                data-idpel="{row["IDPEL"]}" 
                data-blth="{blth_kini}" 
                data-unitup="{row["UNITUP"]}">
                <i class="bi bi-graph-up"></i> Grafik
            </button>''',
            axis=1
        )

        # ✅ FINAL VALIDATION: Pastikan tidak ada NaN yang jadi string 'nan'
        kroscek = kroscek.replace('nan', '')
        kroscek = kroscek.fillna('')

        logger.info(f"✅ Billing processed successfully: {len(kroscek)} records")
        
        return kroscek, None

    except Exception as e:
        error_msg = f"Error processing billing: {str(e)}"
        logger.error(error_msg)
        logger.error(traceback.format_exc())
        return pd.DataFrame(), error_msg


# =================== SAVE TO BILLING ===================
def save_to_billing_with_trigger(df, engine, username):
    """💾 Simpan ke billing table & auto-sync NOMORKWH dari DIL (support multi uploader)"""
    try:
        records = df.to_dict('records')
        success, failed = 0, 0

        with engine.begin() as conn:
            for record in records:
                try:
                    # Tambahkan kolom 'updated_by'
                    record['updated_by'] = username

                    # Bangun kolom & placeholder dinamis
                    cols = list(record.keys())
                    placeholders = ', '.join([f":{k}" for k in cols])

                    # Update kolom selain primary key
                    updates = ', '.join([
                        f"{k}=VALUES({k})"
                        for k in cols
                        if k not in ['BLTH', 'IDPEL', 'UPDATED_BY']
                    ])

                    sql = text(f"""
                        INSERT INTO billing ({', '.join(cols)})
                        VALUES ({placeholders})
                        ON DUPLICATE KEY UPDATE {updates}
                    """)
                    conn.execute(sql, record)
                    success += 1

                except IntegrityError as e:
                    failed += 1
                    logger.error(f"❌ Gagal simpan {record.get('IDPEL')} oleh {username}: {e}")
                    continue

        # =================== AUTO UPDATE NOMORKWH DARI DIL ===================
        try:
            with engine.begin() as conn:
                sql_update = text("""
                    UPDATE billing b
                    JOIN dil d ON b.idpel = d.idpel
                    SET b.nomorkwh = d.nomorkwh
                    WHERE (b.nomorkwh IS NULL OR b.nomorkwh = '' OR b.nomorkwh IN ('-', '0'));
                """)
                result = conn.execute(sql_update)
                logger.info(f"🔄 Auto-sync NOMORKWH dari DIL: {result.rowcount} baris diperbarui.")
        except Exception as sync_err:
            logger.warning(f"⚠️ Auto-sync NOMORKWH gagal: {sync_err}")

        return {
            'success': success,
            'failed': failed,
            'message': f"Berhasil: {success}, Gagal: {failed}"
        }

    except Exception as e:
        logger.error(f"🔥 Error save_to_billing_with_trigger: {str(e)}")
        logger.error(traceback.format_exc())
        return {'success': 0, 'failed': len(df), 'message': str(e)}



@app.route('/api/grafik/<idpel>', methods=['GET'])
def get_grafik_data(idpel):
    """
    📊 API untuk mendapatkan data grafik (JSON) - FIXED VERSION
    """
    try:
        # Normalize dan validasi input
        idpel = normalize_idpel(idpel)
        blth = request.args.get('blth', datetime.now().strftime('%Y%m'))
        ulp = request.args.get('ulp', '').strip()
        
        # Generate 6 bulan terakhir
        try:
            blth = normalize_blth(blth)
            months = [get_previous_blth(blth, i) for i in range(6)]
            months.reverse()
        except Exception as e:
            logger.error(f"Error generating months: {str(e)}")
            return jsonify({'error': f'Invalid BLTH format: {blth}'}), 400
        
        # Log untuk debugging
        logger.info(f"Fetching grafik for IDPEL={idpel}, BLTH={blth}, ULP={ulp}, Months={months}")
        
        # Query dengan filter UNITUP (PERBAIKAN: gunakan expandparams)
        if ulp:
            # Buat placeholder untuk IN clause
            placeholders = ','.join([f':month{i}' for i in range(len(months))])
            query_str = f"""
                SELECT 
                    d.blth,
                    d.lwbppakai,
                    d.daya,
                    COALESCE(b.jam_nyala, 0) as jam_nyala,
                    COALESCE(b.delta_pemkwh, 0) as delta_pemkwh
                FROM dpm d
                LEFT JOIN billing b 
                    ON d.idpel = b.idpel 
                    AND d.blth = b.blth
                    AND b.unitup = :unitup
                WHERE d.idpel = :idpel
                  AND d.blth IN ({placeholders})
                ORDER BY d.blth ASC
            """
            
            # Build params dictionary
            params = {'idpel': idpel, 'unitup': ulp}
            for i, month in enumerate(months):
                params[f'month{i}'] = month
            
            df = pd.read_sql(text(query_str), engine, params=params)
            
        else:
            # Tanpa filter ULP
            placeholders = ','.join([f':month{i}' for i in range(len(months))])
            query_str = f"""
                SELECT 
                    d.blth,
                    d.lwbppakai,
                    d.daya,
                    COALESCE(b.jam_nyala, 0) as jam_nyala,
                    COALESCE(b.delta_pemkwh, 0) as delta_pemkwh
                FROM dpm d
                LEFT JOIN billing b 
                    ON d.idpel = b.idpel 
                    AND d.blth = b.blth
                WHERE d.idpel = :idpel
                  AND d.blth IN ({placeholders})
                ORDER BY d.blth ASC
            """
            
            params = {'idpel': idpel}
            for i, month in enumerate(months):
                params[f'month{i}'] = month
            
            df = pd.read_sql(text(query_str), engine, params=params)
        
        # Log hasil query
        logger.info(f"Query returned {len(df)} rows")
        
        if df.empty:
            logger.warning(f"No data found for IDPEL={idpel}")
            return jsonify({
                'error': f'Data tidak ditemukan untuk IDPEL {idpel}',
                'detail': 'Tidak ada data DPM untuk 6 bulan terakhir'
            }), 404
        
        # Konversi ke numerik dengan error handling
        numeric_columns = ['lwbppakai', 'daya', 'jam_nyala', 'delta_pemkwh']
        for col in numeric_columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        
        # Hitung jam_nyala jika kosong (0 atau NULL)
        mask_empty_jn = (df['jam_nyala'] == 0) & (df['lwbppakai'] > 0) & (df['daya'] > 0)
        if mask_empty_jn.any():
            daya_kw = df.loc[mask_empty_jn, 'daya'] / 1000
            # Hindari division by zero
            daya_kw = daya_kw.replace(0, np.nan)
            jam_nyala_calculated = df.loc[mask_empty_jn, 'lwbppakai'] / daya_kw
            df.loc[mask_empty_jn, 'jam_nyala'] = jam_nyala_calculated.replace([np.inf, -np.inf], 0).fillna(0)
        
        # Hitung delta jika kosong
        mask_empty_delta = df['delta_pemkwh'] == 0
        if mask_empty_delta.any():
            df.loc[mask_empty_delta, 'delta_pemkwh'] = df['lwbppakai'].diff().fillna(0)
        
        # Format BLTH untuk display (YYYYMM -> MMM YYYY)
        def format_blth(blth_str):
            try:
                blth_str = str(blth_str)
                year = blth_str[:4]
                month = blth_str[4:6]
                month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'Mei', 'Jun', 
                               'Jul', 'Agu', 'Sep', 'Okt', 'Nov', 'Des']
                return f"{month_names[int(month)-1]} {year}"
            except:
                return str(blth_str)
        
        # Prepare response
        response_data = {
            'idpel': idpel,
            'unitup': ulp or 'ALL',
            'blth_requested': blth,
            'labels': [format_blth(b) for b in df['blth'].tolist()],
            'labels_raw': df['blth'].tolist(),
            'lwbppakai': df['lwbppakai'].astype(int).tolist(),
            'jam_nyala': df['jam_nyala'].round(2).tolist(),
            'delta': df['delta_pemkwh'].astype(int).tolist(),
            'daya': df['daya'].astype(int).tolist()
        }
        
        logger.info(f"Successfully returned grafik data: {len(response_data['labels'])} months")
        return jsonify(response_data)
        
    except Exception as e:
        logger.error(f"Error in get_grafik_data: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'error': 'Terjadi kesalahan saat memuat data grafik',
            'detail': str(e),
            'type': type(e).__name__
        }), 500
        
        
        
# @app.route('/dashboard_ulp', methods=['GET', 'POST'])
# def dashboard_ulp():
#     """📊 Dashboard ULP"""
#     if 'loggedin' not in session:
#         return redirect(url_for('login'))
    
#     unitup = session.get('unitup')
#     username = session.get('username')
#     role = session.get('role', 'ULP')
#     nama = session.get('nama_ulp')
    
#     # ✅ UP3 tidak perlu validasi UNITUP
#     if role != 'UP3' and not unitup:
#         flash('UNITUP tidak ditemukan di session', 'danger')
#         return redirect(url_for('login'))
    
#     blth_kini = request.form.get('blth', datetime.now().strftime('%Y%m'))
#     blth_kini = normalize_blth(blth_kini)
    
#     # =================== UPLOAD DPM ===================
#     if request.method == 'POST' and 'file_dpm' in request.files:
#         file = request.files['file_dpm']
        
#         if file.filename == '':
#             flash('Tidak ada file yang dipilih', 'warning')
#             return redirect(url_for('dashboard_ulp'))
        
#         try:
#             df_upload = pd.read_excel(file)
#             df_upload.columns = [c.strip().upper() for c in df_upload.columns]
            
#             # ✅ Jika UP3 upload, pastikan ada kolom UNITUP (TANPA suffix)
#             if role == 'UP3':
#                 if 'UNITUP' not in df_upload.columns:
#                     flash('❌ File Excel harus memiliki kolom UNITUP untuk upload oleh UP3!', 'danger')
#                     return redirect(url_for('dashboard_ulp'))
                
#                 # ✅ Normalisasi UNITUP tanpa suffix
#                 df_upload['UNITUP'] = df_upload['UNITUP'].astype(str).str.strip()
#                 unitup_for_upload = None  # Tidak pakai session
#                 logger.info(f"✅ UP3 upload detected. UNITUP: {df_upload['UNITUP'].unique()}")
            
#             else:
#                 # Untuk ULP biasa, pakai unitup dari session
#                 unitup_for_upload = unitup
#                 logger.info(f"✅ ULP upload detected. Using UNITUP from session: {unitup}")
            
#             # Proses upload
#             count, error = process_dpm_upload(df_upload, blth_kini, unitup_for_upload)
            
#             if error:
#                 flash(f'Gagal upload: {error}', 'danger')
#             else:
#                 uploader_label = f"oleh {username} (UP3)" if role == 'UP3' else f"untuk {unitup}"
#                 flash(f'✅ Berhasil upload {count} data DPM {uploader_label}', 'success')
            
#         except Exception as e:
#             logger.error(f"Upload error: {str(e)}")
#             logger.error(traceback.format_exc())
#             flash(f'Error processing file: {str(e)}', 'danger')
#             return redirect(url_for('dashboard_ulp'))

    
#     # =================== PROCESS BILLING ===================
#     if request.method == 'POST' and request.form.get('action') == 'process_billing':
#         # ✅ UP3 proses semua data atau filter by UNITUP jika diperlukan
#         if role == 'UP3':
#             # Jika UP3 ingin proses semua ULP, gunakan None atau '%'
#             df_billing, error = process_billing_advanced(blth_kini, None, engine)
#         else:
#             df_billing, error = process_billing_advanced(blth_kini, unitup, engine)
        
#         if error:
#             flash(f'Gagal proses billing: {error}', 'danger')
#         elif df_billing.empty:
#             flash('Tidak ada data untuk diproses', 'warning')
#         else:
#             result = save_to_billing_with_trigger(df_billing, engine, username)
#             flash(result['message'], 'success' if result['success'] > 0 else 'danger')
        
#     # =================== GET SUMMARY ===================
#     try:
#         if role == 'UP3':
#             # UP3: Tampilkan summary semua ULP
#             query = text("""
#                 SELECT 
#                     unitup,
#                     blth,
#                     COUNT(*) as total,
#                     SUM(CASE WHEN ket = 'NAIK' THEN 1 ELSE 0 END) as naik,
#                     SUM(CASE WHEN ket = 'TURUN' THEN 1 ELSE 0 END) as turun,
#                     SUM(CASE WHEN ket = 'DIV/NA' THEN 1 ELSE 0 END) as div_na,
#                     SUM(CASE WHEN ket = 'AMAN' THEN 1 ELSE 0 END) as aman
#                 FROM billing
#                 GROUP BY unitup, blth
#                 ORDER BY blth DESC, unitup ASC
#                 LIMIT 20
#             """)
#             df_summary = pd.read_sql(query, engine)
#         else:
#             # ULP: Tampilkan summary hanya unitup sendiri
#             query = text("""
#                 SELECT 
#                     blth,
#                     COUNT(*) as total,
#                     SUM(CASE WHEN ket = 'NAIK' THEN 1 ELSE 0 END) as naik,
#                     SUM(CASE WHEN ket = 'TURUN' THEN 1 ELSE 0 END) as turun,
#                     SUM(CASE WHEN ket = 'DIV/NA' THEN 1 ELSE 0 END) as div_na,
#                     SUM(CASE WHEN ket = 'AMAN' THEN 1 ELSE 0 END) as aman
#                 FROM billing
#                 WHERE unitup = :unitup
#                 GROUP BY blth
#                 ORDER BY blth DESC
#                 LIMIT 6
#             """)
#             df_summary = pd.read_sql(query, engine, params={'unitup': unitup})

#         # Konversi ke integer
#         numeric_cols = ['total', 'naik', 'turun', 'div_na', 'aman']
#         for col in numeric_cols:
#             if col in df_summary.columns:
#                 df_summary[col] = df_summary[col].fillna(0).astype(int)

#     except Exception as e:
#         logger.error(f"Error fetching summary: {str(e)}")
#         flash(f'Gagal membaca data: {str(e)}', 'danger')
#         df_summary = pd.DataFrame()
    
#     return render_template(
#         'dashboard_ulp.html',
#         nama=nama,
#         unitup=unitup if role != 'UP3' else 'UP3',
#         role=role,
#         summary=df_summary.to_dict('records') if not df_summary.empty else [],
#         blth_terakhir=blth_kini
#     )

# Tambahkan ini di file app.py, replace fungsi dashboard_ulp yang lama

# @app.route('/dashboard_ulp', methods=['GET', 'POST'])
# def dashboard_ulp():
#     """📊 Dashboard ULP"""
#     if 'loggedin' not in session:
#         return redirect(url_for('login'))
    
#     unitup = session.get('unitup')
#     username = session.get('username')
#     role = session.get('role', 'ULP')
#     nama = session.get('nama_ulp')
    
#     # ✅ UP3 tidak perlu validasi UNITUP
#     if role != 'UP3' and not unitup:
#         flash('UNITUP tidak ditemukan di session', 'danger')
#         return redirect(url_for('login'))
    
#     blth_kini = request.form.get('blth', datetime.now().strftime('%Y%m'))
#     blth_kini = normalize_blth(blth_kini)
    
#     # =================== UPLOAD DPM ===================
#     if request.method == 'POST' and 'file_dpm' in request.files:
#         file = request.files['file_dpm']
        
#         if file.filename == '':
#             flash('Tidak ada file yang dipilih', 'warning')
#             return redirect(url_for('dashboard_ulp'))
        
#         try:
#             df_upload = pd.read_excel(file)
#             df_upload.columns = [c.strip().upper() for c in df_upload.columns]
            
#             # ✅ Jika UP3 upload, pastikan ada kolom UNITUP (TANPA suffix)
#             if role == 'UP3':
#                 if 'UNITUP' not in df_upload.columns:
#                     flash('❌ File Excel harus memiliki kolom UNITUP untuk upload oleh UP3!', 'danger')
#                     return redirect(url_for('dashboard_ulp'))
                
#                 # ✅ Normalisasi UNITUP tanpa suffix
#                 df_upload['UNITUP'] = df_upload['UNITUP'].astype(str).str.strip()
#                 unitup_for_upload = None  # Tidak pakai session
#                 logger.info(f"✅ UP3 upload detected. UNITUP: {df_upload['UNITUP'].unique()}")
            
#             else:
#                 # Untuk ULP biasa, pakai unitup dari session
#                 unitup_for_upload = unitup
#                 logger.info(f"✅ ULP upload detected. Using UNITUP from session: {unitup}")
            
#             # Proses upload
#             count, error = process_dpm_upload(df_upload, blth_kini, unitup_for_upload)
            
#             if error:
#                 flash(f'Gagal upload: {error}', 'danger')
#             else:
#                 uploader_label = f"oleh {username} (UP3)" if role == 'UP3' else f"untuk {unitup}"
#                 flash(f'✅ Berhasil upload {count} data DPM {uploader_label}', 'success')
            
#         except Exception as e:
#             logger.error(f"Upload error: {str(e)}")
#             logger.error(traceback.format_exc())
#             flash(f'Error processing file: {str(e)}', 'danger')
#             return redirect(url_for('dashboard_ulp'))

#         # =================== UPDATE DPM KOREKSI (FIXED VERSION) ===================
#     if request.method == 'POST' and 'file_dpm_update' in request.files:
#         file = request.files['file_dpm_update']
        
#         if file.filename == '':
#             flash('Tidak ada file yang dipilih untuk update', 'warning')
#             return redirect(url_for('dashboard_ulp'))
        
#         try:
#             # 📥 STEP 1: Baca file Excel
#             logger.info(f"🔄 Starting DPM update by {username} ({role})")
#             df_update = pd.read_excel(file)
#             df_update.columns = [c.strip().upper() for c in df_update.columns]
#             logger.info(f"📊 File uploaded: {len(df_update)} rows")
#             logger.info(f"📋 Columns: {list(df_update.columns)}")
            
#             # ✅ STEP 2: Validasi kolom wajib
#             required_cols = ['BLTH', 'IDPEL', 'LWBPPAKAI', 'LWBPCABUT', 'LWBPPASANG', 'SAHLWBP']
#             missing_cols = [col for col in required_cols if col not in df_update.columns]
            
#             if missing_cols:
#                 logger.error(f"❌ Missing columns: {missing_cols}")
#                 flash(f'❌ Kolom tidak lengkap: {", ".join(missing_cols)}', 'danger')
#                 return redirect(url_for('dashboard_ulp'))
            
#             # 📝 STEP 3: Cek kolom UNITUP untuk UP3
#             if role == 'UP3':
#                 if 'UNITUP' not in df_update.columns:
#                     logger.error("❌ UP3 upload without UNITUP column")
#                     flash('❌ File Excel harus memiliki kolom UNITUP untuk upload oleh UP3!', 'danger')
#                     return redirect(url_for('dashboard_ulp'))
#                 df_update['UNITUP'] = df_update['UNITUP'].astype(str).str.strip()
#                 logger.info(f"✅ UP3 mode: UNITUP found: {df_update['UNITUP'].unique().tolist()}")
            
#             # 🔧 STEP 4: Cek format IDPEL di database terlebih dahulu
#             logger.info("🔍 Checking IDPEL format in database...")
#             sample_check = text("""
#                 SELECT idpel, LENGTH(idpel) as len, 
#                     CASE WHEN idpel = LOWER(idpel) THEN 'lowercase' ELSE 'mixed' END as case_type
#                 FROM dpm 
#                 WHERE unitup = :unitup 
#                 LIMIT 3
#             """)
            
#             with engine.connect() as conn:
#                 db_sample = conn.execute(sample_check, {
#                     'unitup': unitup if role == 'ULP' else df_update['UNITUP'].iloc[0]
#                 }).fetchall()
                
#                 if db_sample:
#                     logger.info(f"📋 Database IDPEL format samples:")
#                     for row in db_sample:
#                         logger.info(f"   IDPEL: {row.idpel}, Length: {row.len}, Case: {row.case_type}")
            
#             # 🔧 STEP 5: Normalisasi data sesuai format database
#             logger.info("🔧 Normalizing data...")
            
#             # Simpan IDPEL original untuk tracking
#             df_update['IDPEL_ORIGINAL'] = df_update['IDPEL'].astype(str).str.strip()
            
#             # Normalisasi IDPEL (sesuaikan dengan format di database)
#             df_update['IDPEL'] = df_update['IDPEL'].astype(str).str.strip().str.zfill(12).str.lower()
            
#             # Normalisasi BLTH
#             df_update['BLTH'] = df_update['BLTH'].apply(lambda x: normalize_blth(str(x)))
            
#             # Log sample normalisasi
#             sample = df_update[['IDPEL_ORIGINAL', 'IDPEL', 'BLTH']].head(3)
#             logger.info(f"📋 Sample after normalization:")
#             logger.info(f"\n{sample.to_string()}")
            
#             # Konversi kolom numerik
#             numeric_cols = ['LWBPPAKAI', 'LWBPCABUT', 'LWBPPASANG', 'SAHLWBP']
#             for col in numeric_cols:
#                 df_update[col] = pd.to_numeric(df_update[col], errors='coerce').fillna(0).astype(int)
            
#             logger.info(f"✅ Numeric conversion done")
#             logger.info(f"   Sample LWBPPAKAI: {df_update['LWBPPAKAI'].head(3).tolist()}")
#             logger.info(f"   Sample SAHLWBP: {df_update['SAHLWBP'].head(3).tolist()}")
            
#             # 💾 STEP 6: Update database
#             updated_count = 0
#             failed_count = 0
#             failed_records = []
            
#             logger.info("💾 Starting database update...")
#             logger.info(f"{'='*70}")
            
#             with engine.begin() as conn:
#                 for idx, row in df_update.iterrows():
#                     try:
#                         # Tentukan UNITUP
#                         target_unitup = row['UNITUP'] if role == 'UP3' else unitup
                        
#                         # 🔍 STEP 6.1: Cek data exists di database
#                         check_sql = text("""
#                             SELECT 
#                                 blth, idpel, unitup, nama, tarif, daya,
#                                 slalwbp, lwbp, lwbpcabut, lwbppasang, sahlwbp, lwbppakai
#                             FROM dpm 
#                             WHERE blth = :blth 
#                             AND idpel = :idpel 
#                             AND unitup = :unitup
#                         """)
                        
#                         existing = conn.execute(check_sql, {
#                             'blth': row['BLTH'],
#                             'idpel': row['IDPEL'],
#                             'unitup': target_unitup
#                         }).fetchone()
                        
#                         # Log detail untuk 5 record pertama
#                         if idx < 5:
#                             logger.info(f"\n🔍 Processing Row {idx + 1}:")
#                             logger.info(f"   Search params: BLTH={row['BLTH']}, IDPEL={row['IDPEL']}, UNITUP={target_unitup}")
#                             logger.info(f"   New values: LWBPPAKAI={row['LWBPPAKAI']}, LWBPCABUT={row['LWBPCABUT']}, "
#                                     f"LWBPPASANG={row['LWBPPASANG']}, SAHLWBP={row['SAHLWBP']}")
                        
#                         if not existing:
#                             # Data tidak ditemukan
#                             failed_count += 1
#                             failed_records.append({
#                                 'row': idx + 1,
#                                 'idpel_original': row['IDPEL_ORIGINAL'],
#                                 'idpel_normalized': row['IDPEL'],
#                                 'blth': row['BLTH'],
#                                 'unitup': target_unitup,
#                                 'reason': 'Data tidak ditemukan di database'
#                             })
                            
#                             if idx < 5:
#                                 logger.warning(f"   ❌ NOT FOUND in database")
                            
#                             continue
                        
#                         # Log data yang ditemukan
#                         if idx < 5:
#                             logger.info(f"   ✅ FOUND in database:")
#                             logger.info(f"      Nama: {existing.nama}")
#                             logger.info(f"      Tarif/Daya: {existing.tarif}/{existing.daya}")
#                             logger.info(f"      Current values: LWBPPAKAI={existing.lwbppakai}, SAHLWBP={existing.sahlwbp}")
                        
#                         # 🔧 STEP 6.2: Update data
#                         update_sql = text("""
#                             UPDATE dpm 
#                             SET 
#                                 LWBPPAKAI = :lwbppakai,
#                                 LWBPCABUT = :lwbpcabut,
#                                 LWBPPASANG = :lwbppasang,
#                                 SAHLWBP = :sahlwbp,
#                                 UPDATED_BY = :username
#                             WHERE BLTH = :blth 
#                             AND IDPEL = :idpel
#                             AND UNITUP = :unitup
#                         """)
                        
#                         result = conn.execute(update_sql, {
#                             'lwbppakai': int(row['LWBPPAKAI']),
#                             'lwbpcabut': int(row['LWBPCABUT']),
#                             'lwbppasang': int(row['LWBPPASANG']),
#                             'sahlwbp': int(row['SAHLWBP']),
#                             'username': username,
#                             'blth': row['BLTH'],
#                             'idpel': row['IDPEL'],
#                             'unitup': target_unitup
#                         })
                        
#                         if result.rowcount > 0:
#                             updated_count += 1
#                             if idx < 5:
#                                 logger.info(f"   ✅ UPDATE SUCCESS (affected {result.rowcount} row)")
#                         else:
#                             failed_count += 1
#                             failed_records.append({
#                                 'row': idx + 1,
#                                 'idpel_original': row['IDPEL_ORIGINAL'],
#                                 'idpel_normalized': row['IDPEL'],
#                                 'blth': row['BLTH'],
#                                 'unitup': target_unitup,
#                                 'reason': 'Update affected 0 rows (unexpected)'
#                             })
#                             if idx < 5:
#                                 logger.warning(f"   ⚠️ UPDATE FAILED: affected 0 rows")
                        
#                         # 🔍 STEP 6.3: Verify update (untuk 3 record pertama)
#                         if idx < 3 and result.rowcount > 0:
#                             verify = conn.execute(check_sql, {
#                                 'blth': row['BLTH'],
#                                 'idpel': row['IDPEL'],
#                                 'unitup': target_unitup
#                             }).fetchone()
                            
#                             if verify:
#                                 logger.info(f"   ✓ Verified: LWBPPAKAI={verify.lwbppakai}, SAHLWBP={verify.sahlwbp}")
                        
#                     except Exception as e:
#                         failed_count += 1
#                         failed_records.append({
#                             'row': idx + 1,
#                             'idpel_original': row.get('IDPEL_ORIGINAL', 'unknown'),
#                             'idpel_normalized': row.get('IDPEL', 'unknown'),
#                             'blth': row.get('BLTH', 'unknown'),
#                             'unitup': target_unitup if 'target_unitup' in locals() else 'unknown',
#                             'reason': f'Error: {str(e)}'
#                         })
#                         logger.error(f"   ❌ Exception on row {idx + 1}: {str(e)}")
#                         continue
            
#             # 📊 STEP 7: Summary
#             logger.info(f"\n{'='*70}")
#             logger.info(f"📊 UPDATE SUMMARY:")
#             logger.info(f"   📥 Total rows in file: {len(df_update)}")
#             logger.info(f"   ✅ Successfully updated: {updated_count}")
#             logger.info(f"   ❌ Failed: {failed_count}")
#             logger.info(f"   📈 Success rate: {(updated_count/len(df_update)*100):.1f}%")
            
#             if failed_records:
#                 logger.warning(f"\n⚠️ FAILED RECORDS (showing first 10):")
#                 for rec in failed_records[:10]:
#                     logger.warning(f"   Row {rec['row']}: IDPEL={rec['idpel_original']} → {rec['idpel_normalized']}")
#                     logger.warning(f"            BLTH={rec['blth']}, UNITUP={rec['unitup']}")
#                     logger.warning(f"            Reason: {rec['reason']}")
            
#             logger.info(f"{'='*70}\n")
            
#             # Flash messages
#             if updated_count > 0:
#                 flash(f'✅ Berhasil update {updated_count} dari {len(df_update)} data DPM', 'success')
            
#             if failed_count > 0:
#                 flash(f'⚠️ Gagal update {failed_count} data. Cek log untuk detail.', 'warning')
                
#                 # Save failed records to file
#                 if failed_records:
#                     df_failed = pd.DataFrame(failed_records)
#                     timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#                     failed_file = f'logs/failed_dpm_update_{timestamp}.xlsx'
                    
#                     # Ensure logs directory exists
#                     import os
#                     os.makedirs('logs', exist_ok=True)
                    
#                     df_failed.to_excel(failed_file, index=False)
#                     logger.info(f"📄 Failed records exported to: {failed_file}")
#                     flash(f'📄 Log kegagalan disimpan di: {failed_file}', 'info')
            
#             if updated_count == 0 and failed_count == len(df_update):
#                 flash('❌ Tidak ada data yang berhasil diupdate! Periksa format IDPEL dan BLTH.', 'danger')
            
#         except Exception as e:
#             logger.error(f"❌ CRITICAL ERROR in DPM update: {str(e)}")
#             logger.error(traceback.format_exc())
#             flash(f'❌ Error processing update: {str(e)}', 'danger')
#             return redirect(url_for('dashboard_ulp'))
    
#     # =================== PROCESS BILLING ===================
#     if request.method == 'POST' and request.form.get('action') == 'process_billing':
#         # ✅ UP3 proses semua data atau filter by UNITUP jika diperlukan
#         if role == 'UP3':
#             df_billing, error = process_billing_advanced(blth_kini, None, engine)
#         else:
#             df_billing, error = process_billing_advanced(blth_kini, unitup, engine)
        
#         if error:
#             flash(f'Gagal proses billing: {error}', 'danger')
#         elif df_billing.empty:
#             flash('Tidak ada data untuk diproses', 'warning')
#         else:
#             result = save_to_billing_with_trigger(df_billing, engine, username)
#             flash(result['message'], 'success' if result['success'] > 0 else 'danger')
        
#     # =================== GET SUMMARY ===================
#     try:
#         if role == 'UP3':
#             query = text("""
#                 SELECT 
#                     unitup,
#                     blth,
#                     COUNT(*) as total,
#                     SUM(CASE WHEN ket = 'NAIK' THEN 1 ELSE 0 END) as naik,
#                     SUM(CASE WHEN ket = 'TURUN' THEN 1 ELSE 0 END) as turun,
#                     SUM(CASE WHEN ket = 'DIV/NA' THEN 1 ELSE 0 END) as div_na,
#                     SUM(CASE WHEN ket = 'AMAN' THEN 1 ELSE 0 END) as aman
#                 FROM billing
#                 GROUP BY unitup, blth
#                 ORDER BY blth DESC, unitup ASC
#                 LIMIT 20
#             """)
#             df_summary = pd.read_sql(query, engine)
#         else:
#             query = text("""
#                 SELECT 
#                     blth,
#                     COUNT(*) as total,
#                     SUM(CASE WHEN ket = 'NAIK' THEN 1 ELSE 0 END) as naik,
#                     SUM(CASE WHEN ket = 'TURUN' THEN 1 ELSE 0 END) as turun,
#                     SUM(CASE WHEN ket = 'DIV/NA' THEN 1 ELSE 0 END) as div_na,
#                     SUM(CASE WHEN ket = 'AMAN' THEN 1 ELSE 0 END) as aman
#                 FROM billing
#                 WHERE unitup = :unitup
#                 GROUP BY blth
#                 ORDER BY blth DESC
#                 LIMIT 6
#             """)
#             df_summary = pd.read_sql(query, engine, params={'unitup': unitup})

#         numeric_cols = ['total', 'naik', 'turun', 'div_na', 'aman']
#         for col in numeric_cols:
#             if col in df_summary.columns:
#                 df_summary[col] = df_summary[col].fillna(0).astype(int)

#     except Exception as e:
#         logger.error(f"Error fetching summary: {str(e)}")
#         flash(f'Gagal membaca data: {str(e)}', 'danger')
#         df_summary = pd.DataFrame()
    
#     return render_template(
#         'dashboard_ulp.html',
#         nama=nama,
#         unitup=unitup if role != 'UP3' else 'UP3',
#         role=role,
#         summary=df_summary.to_dict('records') if not df_summary.empty else [],
#         blth_terakhir=blth_kini
#     )
@app.route('/dashboard_ulp', methods=['GET', 'POST'])
def dashboard_ulp():
    """📊 Dashboard ULP"""
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    unitup = session.get('unitup')
    username = session.get('username')
    role = session.get('role', 'ULP')
    nama = session.get('nama_ulp')
    
    if role != 'UP3' and not unitup:
        flash('UNITUP tidak ditemukan di session', 'danger')
        return redirect(url_for('login'))
    
    blth_kini = request.form.get('blth', datetime.now().strftime('%Y%m'))
    blth_kini = normalize_blth(blth_kini)
    
    # ✅ DEBUG: Log semua request data
    if request.method == 'POST':
        logger.info(f"📥 POST Request by {username} ({role})")
        logger.info(f"   Form data keys: {list(request.form.keys())}")
        logger.info(f"   Files keys: {list(request.files.keys())}")
        form_type = request.form.get('form_type', 'UNKNOWN')
        logger.info(f"   Form type: {form_type}")
    
    # =================== UPLOAD DPM BARU ===================
    if request.method == 'POST' and request.form.get('form_type') == 'upload_dpm':
        logger.info("🔵 Processing UPLOAD DPM BARU")
        
        if 'file_dpm' not in request.files:
            flash('❌ File tidak ditemukan', 'danger')
            return redirect(url_for('dashboard_ulp'))
        
        file = request.files['file_dpm']
        
        if file.filename == '':
            flash('Tidak ada file yang dipilih', 'warning')
            return redirect(url_for('dashboard_ulp'))
        
        try:
            df_upload = pd.read_excel(file)
            df_upload.columns = [c.strip().upper() for c in df_upload.columns]
            
            if role == 'UP3':
                if 'UNITUP' not in df_upload.columns:
                    flash('❌ File Excel harus memiliki kolom UNITUP untuk upload oleh UP3!', 'danger')
                    return redirect(url_for('dashboard_ulp'))
                
                df_upload['UNITUP'] = df_upload['UNITUP'].astype(str).str.strip()
                unitup_for_upload = None
                logger.info(f"✅ UP3 upload. UNITUP: {df_upload['UNITUP'].unique()}")
            else:
                unitup_for_upload = unitup
                logger.info(f"✅ ULP upload. UNITUP: {unitup}")
            
            count, error = process_dpm_upload(df_upload, blth_kini, unitup_for_upload)
            
            if error:
                flash(f'Gagal upload: {error}', 'danger')
            else:
                flash(f'✅ Berhasil upload {count} data DPM', 'success')
            
        except Exception as e:
            logger.error(f"Upload error: {str(e)}")
            flash(f'Error: {str(e)}', 'danger')
        
        return redirect(url_for('dashboard_ulp'))
    
    # =================== UPDATE DPM KOREKSI ===================
    if request.method == 'POST' and request.form.get('form_type') == 'update_koreksi':
        logger.info("🟡 Processing UPDATE DPM KOREKSI")
        
        if 'file_dpm_koreksi' not in request.files:
            flash('❌ File update koreksi tidak ditemukan', 'danger')
            logger.error("❌ file_dpm_koreksi NOT in request.files")
            return redirect(url_for('dashboard_ulp'))
        
        file = request.files['file_dpm_koreksi']
        logger.info(f"📄 File received: {file.filename}")
        
        if file.filename == '':
            flash('Tidak ada file yang dipilih', 'warning')
            return redirect(url_for('dashboard_ulp'))
        
        try:
            logger.info(f"🔄 Starting DPM Koreksi by {username} ({role})")
            
            df_koreksi = pd.read_excel(file)
            df_koreksi.columns = [c.strip().upper() for c in df_koreksi.columns]
            logger.info(f"📊 File: {len(df_koreksi)} rows, Columns: {list(df_koreksi.columns)}")
            
            # Validasi kolom
            required_cols = ['BLTH', 'IDPEL', 'LWBPCABUT', 'LWBPPASANG', 'SAHLWBP']
            missing = [col for col in required_cols if col not in df_koreksi.columns]
            
            if missing:
                flash(f'❌ Kolom tidak lengkap: {", ".join(missing)}', 'danger')
                return redirect(url_for('dashboard_ulp'))
            
            # Handle UNITUP
            has_unitup = 'UNITUP' in df_koreksi.columns
            
            if role == 'UP3':
                if not has_unitup:
                    flash('❌ File Excel harus ada kolom UNITUP untuk UP3!', 'danger')
                    return redirect(url_for('dashboard_ulp'))
                df_koreksi['UNITUP'] = df_koreksi['UNITUP'].astype(str).str.strip()
                logger.info(f"✅ UP3 mode. UNITUP: {df_koreksi['UNITUP'].unique().tolist()}")
            else:
                df_koreksi['UNITUP'] = unitup
                logger.info(f"✅ ULP mode. UNITUP: {unitup}")
            
            # Normalisasi
            df_koreksi['IDPEL'] = df_koreksi['IDPEL'].astype(str).str.strip().str.zfill(12).str.lower()
            df_koreksi['BLTH'] = df_koreksi['BLTH'].apply(lambda x: normalize_blth(str(x)))
            
            for col in ['LWBPCABUT', 'LWBPPASANG', 'SAHLWBP']:
                df_koreksi[col] = pd.to_numeric(df_koreksi[col], errors='coerce').fillna(0).astype(int)
            
            # Update database
            updated = failed = 0
            
            with engine.begin() as conn:
                for idx, row in df_koreksi.iterrows():
                    try:
                        target_unitup = str(row['UNITUP']).strip()
                        
                        if not target_unitup:
                            failed += 1
                            continue
                        
                        # Check existing
                        check = conn.execute(text("""
                            SELECT slalwbp, lwbppakai, COALESCE(marking_koreksi, 0) as mk
                            FROM dpm 
                            WHERE blth = :b AND idpel = :i AND unitup = :u
                        """), {'b': row['BLTH'], 'i': row['IDPEL'], 'u': target_unitup}).fetchone()
                        
                        if not check:
                            failed += 1
                            if idx < 3:
                                logger.warning(f"Row {idx+1}: NOT FOUND - {row['IDPEL']}")
                            continue
                        
                        # Calculate
                        lwbp_new = int(row['LWBPCABUT']) - int(check.slalwbp or 0) + int(row['SAHLWBP']) - int(row['LWBPPASANG'])
                        mk_new = check.mk + 1
                        
                        # Update
                        result = conn.execute(text("""
                            UPDATE dpm 
                            SET LWBPCABUT=:lc, LWBPPASANG=:lp, SAHLWBP=:sh, 
                                LWBPPAKAI=:lw, MARKING_KOREKSI=:mk, UPDATED_BY=:u
                            WHERE BLTH=:b AND IDPEL=:i AND UNITUP=:un
                        """), {
                            'lc': int(row['LWBPCABUT']), 'lp': int(row['LWBPPASANG']),
                            'sh': int(row['SAHLWBP']), 'lw': lwbp_new, 'mk': mk_new,
                            'u': username, 'b': row['BLTH'], 'i': row['IDPEL'], 'un': target_unitup
                        })
                        
                        if result.rowcount > 0:
                            updated += 1
                        else:
                            failed += 1
                        
                    except Exception as e:
                        failed += 1
                        logger.error(f"Row {idx+1} error: {e}")
            
            logger.info(f"📊 Summary: {updated} success, {failed} failed")
            
            if updated > 0:
                flash(f'✅ Update {updated}/{len(df_koreksi)} data (MARKING_KOREKSI +1)', 'success')
                flash('💡 Proses ulang billing untuk refresh', 'info')
            
            if failed > 0:
                flash(f'⚠️ Gagal {failed} data', 'warning')
            
        except Exception as e:
            logger.error(f"❌ CRITICAL: {str(e)}")
            logger.error(traceback.format_exc())
            flash(f'❌ Error: {str(e)}', 'danger')
        
        return redirect(url_for('dashboard_ulp'))
    
    # =================== PROCESS BILLING ===================
    if request.method == 'POST' and request.form.get('action') == 'process_billing':
        logger.info("🟢 Processing BILLING")
        
        if role == 'UP3':
            df_billing, error = process_billing_advanced(blth_kini, None, engine)
        else:
            df_billing, error = process_billing_advanced(blth_kini, unitup, engine)
        
        if error:
            flash(f'Gagal: {error}', 'danger')
        elif df_billing.empty:
            flash('Tidak ada data', 'warning')
        else:
            result = save_to_billing_with_trigger(df_billing, engine, username)
            flash(result['message'], 'success' if result['success'] > 0 else 'danger')
        
        return redirect(url_for('dashboard_ulp'))
    
    # =================== GET SUMMARY ===================
    try:
        if role == 'UP3':
            query = text("""
                SELECT unitup, blth, COUNT(*) as total,
                       SUM(CASE WHEN ket='NAIK' THEN 1 ELSE 0 END) as naik,
                       SUM(CASE WHEN ket='TURUN' THEN 1 ELSE 0 END) as turun,
                       SUM(CASE WHEN ket='DIV/NA' THEN 1 ELSE 0 END) as div_na,
                       SUM(CASE WHEN ket='AMAN' THEN 1 ELSE 0 END) as aman
                FROM billing
                GROUP BY unitup, blth
                ORDER BY blth DESC, unitup ASC
                LIMIT 20
            """)
            df_summary = pd.read_sql(query, engine)
        else:
            query = text("""
                SELECT blth, COUNT(*) as total,
                       SUM(CASE WHEN ket='NAIK' THEN 1 ELSE 0 END) as naik,
                       SUM(CASE WHEN ket='TURUN' THEN 1 ELSE 0 END) as turun,
                       SUM(CASE WHEN ket='DIV/NA' THEN 1 ELSE 0 END) as div_na,
                       SUM(CASE WHEN ket='AMAN' THEN 1 ELSE 0 END) as aman
                FROM billing
                WHERE unitup = :u
                GROUP BY blth
                ORDER BY blth DESC
                LIMIT 6
            """)
            df_summary = pd.read_sql(query, engine, params={'u': unitup})
        
        for col in ['total', 'naik', 'turun', 'div_na', 'aman']:
            if col in df_summary.columns:
                df_summary[col] = df_summary[col].fillna(0).astype(int)
    
    except Exception as e:
        logger.error(f"Summary error: {e}")
        df_summary = pd.DataFrame()
    
    return render_template(
        'dashboard_ulp.html',
        nama=nama,
        unitup=unitup if role != 'UP3' else 'UP3',
        role=role,
        summary=df_summary.to_dict('records') if not df_summary.empty else [],
        blth_terakhir=blth_kini
    )
# Import other routes from your original code...
# (view_billing, dashboard_up3, login, logout, etc.)

from flask import render_template, request, redirect, url_for, session, flash, jsonify
from sqlalchemy import text
import pandas as pd
from markupsafe import escape
import logging
import traceback

logger = logging.getLogger(__name__)

# =================== VIEW BILLING (Fixed & Secure) ===================
# =================== VIEW BILLING ===================
import math
from flask import request, render_template, redirect, url_for, session, flash
import pandas as pd
from sqlalchemy import text
import math  # ✅ Pastikan sudah di-import di atas file




@app.route('/view_billing', methods=['GET'])
def view_billing():
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    unitup = session.get('unitup')
    username = session.get('username')
    role = session.get('role', 'ULP')
    
    logger.info(f"📊 View billing: {username} (UNITUP: {unitup}, Role: {role})")
    
    # ===== PAGINATION PARAMETERS =====
    active_tab = request.args.get('tab', 'dlpd_3bln')
    selected_kelompok = request.args.get('kdkelompok', '')
    selected_blth = request.args.get('blth', '')
    jam_nyala_min = request.args.get('jam_nyala_min', type=float, default=0)
    jam_nyala_max = request.args.get('jam_nyala_max', type=float, default=9999)
    unitup_filter = request.args.get('unitup_filter', '')
    
    # ✅ AUTO-FILTER LOGIC
    auto_filter = request.args.get('auto_filter', '0')
    
    if auto_filter == '1' and unitup_filter:
        logger.info(f"🔍 AUTO-FILTER MODE: UNITUP={unitup_filter}")
        
        # Auto-select BLTH terbaru untuk UNITUP ini jika belum dipilih
        if not selected_blth:
            try:
                latest_query = text("""
                    SELECT MAX(blth) as latest 
                    FROM billing 
                    WHERE unitup = :unitup
                """)
                latest_result = pd.read_sql(latest_query, engine, params={'unitup': unitup_filter})
                
                if not latest_result.empty and latest_result.iloc[0]['latest']:
                    selected_blth = str(latest_result.iloc[0]['latest'])
                    logger.info(f"✅ Auto-selected BLTH: {selected_blth}")
            except Exception as e:
                logger.error(f"Error auto-selecting BLTH: {e}")
    
    # ✅ Pagination
    page = request.args.get('page', 1, type=int)
    rows_per_page = request.args.get('rows_per_page', 300, type=int)
    
    # DEBUG LOG
    logger.info(f"📊 Received rows_per_page: {rows_per_page}")
    logger.info(f"📍 Filters: BLTH={selected_blth}, UNITUP={unitup_filter}, TAB={active_tab}")
    
    # ✅ Validate rows_per_page
    ALLOWED_ROWS = [50, 100, 200, 300, 500]
    if rows_per_page not in ALLOWED_ROWS:
        logger.warning(f"⚠️ Invalid rows_per_page: {rows_per_page}, reset to 300")
        rows_per_page = 300
    else:
        logger.info(f"✅ Valid rows_per_page: {rows_per_page}")
    
    offset = (page - 1) * rows_per_page
    logger.info(f"📄 Page: {page}, Offset: {offset}, Limit: {rows_per_page}")
    
    # ===== BASE QUERY =====
    base_query = "SELECT * FROM billing WHERE 1=1"
    count_query = "SELECT COUNT(*) as total FROM billing WHERE 1=1"
    params = {}
    
    # ✅ PERBAIKAN: Prioritaskan unitup_filter (dari link) daripada session
    if unitup_filter:
        # Jika ada filter UNITUP dari parameter (UP3 mode)
        base_query += " AND unitup = :unitup"
        count_query += " AND unitup = :unitup"
        params['unitup'] = unitup_filter
        logger.info(f"🏢 Using unitup_filter: {unitup_filter}")
    elif role == 'ULP':
        # Jika ULP, gunakan unitup dari session
        base_query += " AND unitup = :unitup"
        count_query += " AND unitup = :unitup"
        params['unitup'] = unitup
        logger.info(f"🏢 Using session unitup: {unitup}")
    # elif role == 'UP3' tanpa filter = tampil semua
    
    # ===== AUTO-SELECT LATEST BLTH (jika belum ada) =====
    if selected_blth:
        base_query += " AND blth = :blth"
        count_query += " AND blth = :blth"
        params['blth'] = selected_blth
    else:
        try:
            latest_query = "SELECT MAX(blth) as latest FROM billing WHERE 1=1"
            latest_params = {}
            
            if params.get('unitup'):
                latest_query += " AND unitup = :unitup"
                latest_params['unitup'] = params['unitup']
            
            latest_result = pd.read_sql(text(latest_query), engine, params=latest_params)
            if not latest_result.empty and latest_result.iloc[0]['latest']:
                selected_blth = str(latest_result.iloc[0]['latest'])
                base_query += " AND blth = :blth"
                count_query += " AND blth = :blth"
                params['blth'] = selected_blth
                logger.info(f"✅ Auto-selected BLTH (fallback): {selected_blth}")
        except Exception as e:
            logger.error(f"Error getting latest BLTH: {e}")
    
    if selected_kelompok:
        base_query += " AND kdkelompok = :kdkelompok"
        count_query += " AND kdkelompok = :kdkelompok"
        params['kdkelompok'] = selected_kelompok
    
    try:
        dlpd_3bln_html = naik_html_v2 = turun_html_v2 = div_html_v2 = jam_nyala_html = ""
        total_rows = 0
        
        # ===== TAB DLPD 3 BULAN =====
        if active_tab == 'dlpd_3bln':
            count_q = count_query + " AND DLPD_3BLN = :dlpd_value"
            count_p = params.copy()
            count_p['dlpd_value'] = 'Naik50% R3BLN'
            total_rows = pd.read_sql(text(count_q), engine, params=count_p)['total'].iloc[0]
            
            query = base_query + " AND DLPD_3BLN = :dlpd_value ORDER BY idpel ASC LIMIT :limit OFFSET :offset"
            params['dlpd_value'] = 'Naik50% R3BLN'
            params['limit'] = rows_per_page
            params['offset'] = offset
            data = pd.read_sql(text(query), engine, params=params)
            data.drop(columns=['updated_by'], errors='ignore', inplace=True)
            dlpd_3bln_html = create_editable_table(data) if not data.empty else "<p>Tidak ada data DLPD 3BLN</p>"
        
        # ===== TAB NAIK =====
        elif active_tab == 'naik':
            count_q = count_query + " AND ket = :ket_value"
            count_p = params.copy()
            count_p['ket_value'] = 'NAIK'
            total_rows = pd.read_sql(text(count_q), engine, params=count_p)['total'].iloc[0]
            
            query = base_query + " AND ket = :ket_value ORDER BY idpel ASC LIMIT :limit OFFSET :offset"
            params['ket_value'] = 'NAIK'
            params['limit'] = rows_per_page
            params['offset'] = offset
            data = pd.read_sql(text(query), engine, params=params)
            data.drop(columns=['updated_by'], errors='ignore', inplace=True)
            naik_html_v2 = create_editable_table(data) if not data.empty else "<p>Tidak ada data sortir naik</p>"
        
        # ===== TAB TURUN =====
        elif active_tab == 'turun':
            count_q = count_query + " AND ket = :ket_value"
            count_p = params.copy()
            count_p['ket_value'] = 'TURUN'
            total_rows = pd.read_sql(text(count_q), engine, params=count_p)['total'].iloc[0]
            
            query = base_query + " AND ket = :ket_value ORDER BY idpel ASC LIMIT :limit OFFSET :offset"
            params['ket_value'] = 'TURUN'
            params['limit'] = rows_per_page
            params['offset'] = offset
            data = pd.read_sql(text(query), engine, params=params)
            data.drop(columns=['updated_by'], errors='ignore', inplace=True)
            turun_html_v2 = create_editable_table(data) if not data.empty else "<p>Tidak ada data sortir turun</p>"
        
        # ===== TAB DIV =====
        elif active_tab == 'div':
            count_q = count_query + " AND ket = :ket_value"
            count_p = params.copy()
            count_p['ket_value'] = 'DIV/NA'
            total_rows = pd.read_sql(text(count_q), engine, params=count_p)['total'].iloc[0]
            
            query = base_query + " AND ket = :ket_value ORDER BY idpel ASC LIMIT :limit OFFSET :offset"
            params['ket_value'] = 'DIV/NA'
            params['limit'] = rows_per_page
            params['offset'] = offset
            data = pd.read_sql(text(query), engine, params=params)
            data.drop(columns=['updated_by'], errors='ignore', inplace=True)
            div_html_v2 = create_editable_table(data) if not data.empty else "<p>Tidak ada data sortir DIV/NA</p>"
        
        # ===== TAB JAM NYALA =====
        elif active_tab == 'jam_nyala':
            count_q = count_query + " AND jam_nyala BETWEEN :min_jn AND :max_jn"
            count_p = params.copy()
            count_p['min_jn'] = jam_nyala_min
            count_p['max_jn'] = jam_nyala_max
            total_rows = pd.read_sql(text(count_q), engine, params=count_p)['total'].iloc[0]
            
            query = base_query + " AND jam_nyala BETWEEN :min_jn AND :max_jn ORDER BY jam_nyala DESC LIMIT :limit OFFSET :offset"
            params['min_jn'] = jam_nyala_min
            params['max_jn'] = jam_nyala_max
            params['limit'] = rows_per_page
            params['offset'] = offset
            data = pd.read_sql(text(query), engine, params=params)
            data.drop(columns=['updated_by'], errors='ignore', inplace=True)
            jam_nyala_html = create_editable_table(data) if not data.empty else "<p>Tidak ada data sortir JN</p>"
        
        # ===== GENERATE PAGINATION =====
        pagination_html = generate_pagination(page, total_rows, rows_per_page, request.args)
        
        # ===== BLTH Dropdown =====
        blth_query = "SELECT DISTINCT blth FROM billing"
        blth_params = {}
        
        if params.get('unitup'):
            blth_query += " WHERE unitup = :unitup"
            blth_params['unitup'] = params['unitup']
        
        blth_query += " ORDER BY blth DESC"
        blth_list = pd.read_sql(text(blth_query), engine, params=blth_params)['blth'].tolist()
        
        # ===== UNITUP Dropdown (untuk UP3) =====
        unitup_list = []
        if role == 'UP3':
            unitup_list = pd.read_sql(
                text("SELECT DISTINCT unitup FROM billing ORDER BY unitup"),
                engine
            )['unitup'].tolist()
        
    except Exception as e:
        logger.error(f"❌ Error loading data: {str(e)}")
        logger.exception(e)
        flash(f'Gagal membaca data: {str(e)}', 'danger')
        dlpd_3bln_html = naik_html_v2 = turun_html_v2 = div_html_v2 = jam_nyala_html = "<p>Error loading data</p>"
        blth_list = []
        unitup_list = []
        total_rows = 0
        pagination_html = ""
    
    return render_template(
        'view_billing.html',
        username=username,
        role=role,
        unitup=unitup,
        active_tab=active_tab,
        selected_kelompok=selected_kelompok,
        selected_blth=selected_blth,
        jam_nyala_min=jam_nyala_min,
        jam_nyala_max=jam_nyala_max,
        dlpd_3bln_html=dlpd_3bln_html,
        naik_html_v2=naik_html_v2,
        turun_html_v2=turun_html_v2,
        div_html_v2=div_html_v2,
        jam_nyala_html=jam_nyala_html,
        blth_list=blth_list,
        unitup_list=unitup_list,
        unitup_filter=unitup_filter,
        total_rows=total_rows,
        rows_per_page=rows_per_page,
        current_page=page,
        pagination_html=pagination_html
    )

def generate_pagination(current_page, total_rows, rows_per_page, args):
    """Generate Bootstrap pagination HTML with preserved filters"""
    # ✅ FIX: Gunakan operator // untuk integer division
    total_pages = math.ceil(total_rows / rows_per_page) if total_rows > 0 else 1
    
    if total_pages <= 1:
        return ""
    
    # Build query params (preserve all filters)
    params = dict(args)
    
    html = '<nav class="mt-2"><ul class="pagination pagination-sm justify-content-center">'
    
    # Previous button
    if current_page > 1:
        params['page'] = current_page - 1
        query_str = '&'.join([f"{k}={v}" for k, v in params.items()])
        html += f'<li class="page-item"><a class="page-link" href="?{query_str}" onclick="showLoading()">«</a></li>'
    else:
        html += '<li class="page-item disabled"><span class="page-link">«</span></li>'
    
    # Page numbers (show max 7 pages)
    start_page = max(1, current_page - 3)
    end_page = min(total_pages, current_page + 3)
    
    # First page
    if start_page > 1:
        params['page'] = 1
        query_str = '&'.join([f"{k}={v}" for k, v in params.items()])
        html += f'<li class="page-item"><a class="page-link" href="?{query_str}" onclick="showLoading()">1</a></li>'
        if start_page > 2:
            html += '<li class="page-item disabled"><span class="page-link">...</span></li>'
    
    # Page numbers
    for p in range(start_page, end_page + 1):
        params['page'] = p
        query_str = '&'.join([f"{k}={v}" for k, v in params.items()])
        active = 'active' if p == current_page else ''
        html += f'<li class="page-item {active}"><a class="page-link" href="?{query_str}" onclick="showLoading()">{p}</a></li>'
    
    # Last page
    if end_page < total_pages:
        if end_page < total_pages - 1:
            html += '<li class="page-item disabled"><span class="page-link">...</span></li>'
        params['page'] = total_pages
        query_str = '&'.join([f"{k}={v}" for k, v in params.items()])
        html += f'<li class="page-item"><a class="page-link" href="?{query_str}" onclick="showLoading()">{total_pages}</a></li>'
    
    # Next button
    if current_page < total_pages:
        params['page'] = current_page + 1
        query_str = '&'.join([f"{k}={v}" for k, v in params.items()])
        html += f'<li class="page-item"><a class="page-link" href="?{query_str}" onclick="showLoading()">»</a></li>'
    else:
        html += '<li class="page-item disabled"><span class="page-link">»</span></li>'
    
    html += '</ul></nav>'
    
    # ✅ FIX: Add info dengan format yang lebih baik
    start_row = (current_page - 1) * rows_per_page + 1
    end_row = min(current_page * rows_per_page, total_rows)
    html = f'<div class="text-center small text-muted mb-2">Showing {start_row:,}-{end_row:,} of {total_rows:,} rows (Page {current_page}/{total_pages})</div>' + html
    
    return html


# =================== CREATE EDITABLE TABLE - FIXED VERSION ===================
def create_editable_table(df):
    """
    📝 Convert DataFrame to editable HTML table with dropdown & textarea
    ✅ FIXED: Bersihkan \n dari semua kolom text sebelum render
    """
    if df.empty:
        return "<p class='text-center text-muted py-4'>Tidak ada data</p>"
    
    try:
        df_display = df.copy()
        
        # ✅ STEP 1: BERSIHKAN \n dari semua kolom text SEBELUM proses apapun
        text_columns = ['NAMA', 'TARIF', 'KDKELOMPOK', 'DLPD', 'DLPD_HITUNG', 
                        'DLPD_3BLN', 'KET', 'JAMNYALA600', 'NOMORKWH']
        
        for col in text_columns:
            if col in df_display.columns:
                df_display[col] = df_display[col].apply(
                    lambda x: str(x).replace('\\n', ' ').replace('\n', ' ').replace('\r', ' ').strip() 
                    if pd.notna(x) and str(x) not in ['nan', 'None', ''] else ''
                )
        
        logger.info("✅ Cleaned \\n from text columns before rendering")
        
        # ✅ STEP 2: Hasil Pemeriksaan options
        hasil_options = [
            "SESUAI", "TEMPER NYALA", "SALAH STAN", "SALAH FOTO", "FOTO BURAM",
            "ANOMALI PDL", "LEBIH TAGIH", "KURANG TAGIH", "BKN FOTO KWH",
            "BENCANA", "3BLN TANPA STAN", "BACA ULANG", "MASUK 720JN"
        ]
        
        # ✅ STEP 3: Create dropdown for HASIL PEMERIKSAAN
        hasil_dropdowns = []
        for _, row in df_display.iterrows():
            current_value = str(row.get('HASIL_PEMERIKSAAN', ''))
            # ⚠️ Bersihkan nilai current dari \n
            current_value = current_value.replace('\\n', ' ').replace('\n', ' ').strip()
            
            options_html = '<option value="" selected hidden>-- Pilih --</option>'
            for opt in hasil_options:
                selected = 'selected' if current_value == opt else ''
                options_html += f'<option value="{opt}" {selected}>{opt}</option>'
            
            dropdown = f'''
                <select name="hasil_pemeriksaan_{row["IDPEL"]}" 
                        class="form-select form-select-sm"
                        data-idpel="{row["IDPEL"]}"
                        data-blth="{row["BLTH"]}"
                        data-column="hasil_pemeriksaan">
                    {options_html}
                </select>
            '''
            hasil_dropdowns.append(dropdown)
        
        df_display['HASIL_PEMERIKSAAN'] = hasil_dropdowns
        
        # ✅ STEP 4: Create textarea for STAN VERIFIKASI
        stan_textareas = []
        for _, row in df_display.iterrows():
            value = row.get('STAN_VERIFIKASI', '')
            if pd.isna(value):
                value = ''
            else:
                # ⚠️ Bersihkan dari \n
                value = str(value).replace('\\n', ' ').replace('\n', ' ').strip()
            
            textarea = f'''
                <textarea name="stan_verifikasi_{row["IDPEL"]}" 
                          class="form-control form-control-sm stan-verifikasi"
                          rows="1"
                          data-idpel="{row["IDPEL"]}"
                          data-blth="{row["BLTH"]}"
                          data-column="stan_verifikasi">{escape(value)}</textarea>
            '''
            stan_textareas.append(textarea)
        
        df_display['STAN_VERIFIKASI'] = stan_textareas
        
        # ✅ STEP 5: Create textarea for TINDAK LANJUT
        tindak_textareas = []
        for _, row in df_display.iterrows():
            value = row.get('TINDAK_LANJUT', '')
            if pd.isna(value):
                value = ''
            else:
                # ⚠️ Bersihkan dari \n
                value = str(value).replace('\\n', ' ').replace('\n', ' ').strip()
            
            textarea = f'''
                <textarea name="tindak_lanjut_{row["IDPEL"]}" 
                          class="form-control form-control-sm tindak-lanjut"
                          rows="2"
                          data-idpel="{row["IDPEL"]}"
                          data-blth="{row["BLTH"]}"
                          data-column="tindak_lanjut">{escape(value)}</textarea>
            '''
            tindak_textareas.append(textarea)
        
        df_display['TINDAK_LANJUT'] = tindak_textareas
        
        # ✅ STEP 6: Bersihkan kolom FOTO dari \n (jika ada)
        foto_columns = ['FOTO_AKHIR', 'FOTO_LALU', 'FOTO_LALU2', 'FOTO_3BLN', 'GRAFIK']
        for col in foto_columns:
            if col in df_display.columns:
                df_display[col] = df_display[col].apply(
                    lambda x: str(x).replace('\n', '').replace('  ', ' ').strip() 
                    if pd.notna(x) else x
                )
        
        # ✅ STEP 7: Convert to HTML
        table_html = df_display.to_html(
            classes="table table-striped table-hover table-sm table-bordered",
            index=False,
            escape=False,  # Jangan escape HTML untuk kolom FOTO dan form elements
            na_rep='',
            table_id="billingTable"
        )
        
        # ✅ STEP 8: Post-process HTML - hapus \n yang tersisa
        table_html = table_html.replace('\\n', ' ').replace('\n\n', '\n')
        
        logger.info("✅ Table HTML generated successfully without \\n")
        
        return table_html
        
    except Exception as e:
        logger.error(f"Error creating table: {str(e)}")
        logger.error(traceback.format_exc())
        return f"<p class='text-danger'>Error: {str(e)}</p>"
# =================== SAVE DATA (Individual Auto-Save) ===================
@app.route('/update_data', methods=['POST'])
def update_data():
    """
    💾 Update single field (dipanggil via AJAX auto-save)
    """
    if 'loggedin' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        data = request.get_json()
        idpel = normalize_idpel(data.get('IDPEL'))
        column = data.get('column')
        value = data.get('value')
        username = session.get('username')
        
        # Map frontend column names to database columns
        column_map = {
            'HASIL PEMERIKSAAN': 'hasil_pemeriksaan',
            'TINDAK LANJUT': 'tindak_lanjut',
            'STAN_VERIFIKASI': 'stan_verifikasi'
            # 'KETERANGAN': 'keterangan'
        }
        
        db_column = column_map.get(column, column.lower().replace(' ', '_'))
        
        sql = text(f"""
            UPDATE billing 
            SET {db_column} = :value,
                updated_by = :username,
                updated_at = NOW()
            WHERE idpel = :idpel
        """)
        
        with engine.begin() as conn:
            result = conn.execute(sql, {
                'value': value,
                'username': username,
                'idpel': idpel
            })
        
        logger.info(f"✅ Updated {db_column} for IDPEL {idpel}")
        
        return jsonify({
            'success': True,
            'message': 'Data updated',
            'rowcount': result.rowcount
        })
        
    except Exception as e:
        logger.error(f"Error updating data: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500


# =================== BATCH SAVE (Form Submit) ===================
@app.route('/simpan_<tab_name>', methods=['POST'])
def save_tab_data(tab_name):
    """
    💾 Batch save untuk form submit (simpan_dlpd, simpan_naik, dll)
    """
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    try:
        username = session.get('username')
        form_data = request.form
        
        # Extract IDPEL dan values dari form
        updates = []
        for key, value in form_data.items():
            if key.startswith('hasil_pemeriksaan_'):
                idpel = key.replace('hasil_pemeriksaan_', '')
                stan = form_data.get(f'stan_verifikasi_{idpel}', '')
                tindak = form_data.get(f'tindak_lanjut_{idpel}', '')
                
                updates.append({
                    'idpel': normalize_idpel(idpel),
                    'hasil': value,
                    'stan': stan,
                    'tindak': tindak
                })
        
        # Batch update
        success = 0
        failed = 0
        
        with engine.begin() as conn:
            for item in updates:
                try:
                    sql = text("""
                        UPDATE billing 
                        SET hasil_pemeriksaan = :hasil,
                            stan_verifikasi = :stan,
                            tindak_lanjut = :tindak,
                            updated_by = :username,
                            updated_at = NOW()
                        WHERE idpel = :idpel
                    """)
                    
                    result = conn.execute(sql, {
                        'hasil': item['hasil'],
                        'stan': item['stan'],
                        'tindak': item['tindak'],
                        'username': username,
                        'idpel': item['idpel']
                    })
                    
                    if result.rowcount > 0:
                        success += 1
                    else:
                        failed += 1
                        
                except Exception as e:
                    logger.error(f"Failed to save {item['idpel']}: {e}")
                    failed += 1
                    continue
        
        flash(f'Berhasil menyimpan {success} data, Gagal: {failed}', 'success' if success > 0 else 'danger')
        
        # Redirect back dengan preserve filter
        kdkelompok = form_data.get('kdkelompok', '')
        active_tab = form_data.get('active_tab', tab_name)
        
        return redirect(url_for('view_billing', 
                               tab=active_tab, 
                               kdkelompok=kdkelompok))
        
    except Exception as e:
        logger.error(f"Error in batch save: {str(e)}")
        logger.error(traceback.format_exc())
        flash(f'Gagal menyimpan: {str(e)}', 'danger')
        return redirect(url_for('view_billing'))




# =================== DASHBOARD UP3 - Selector (Halaman Awal UP3) ===================
@app.route('/dashboard_up3', methods=['GET'])
def dashboard_up3():
    """📊 Dashboard UP3 - Card Selector untuk pilih ULP"""
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    if session.get('role') != 'UP3':
        flash('❌ Akses ditolak. Halaman ini khusus UP3.', 'danger')
        return redirect(url_for('dashboard_ulp'))
    
    username = session.get('username')
    nama = session.get('nama_ulp', 'Administrator UP3')
    
    # Get list of ULP dengan info last upload
    try:
        ulp_query = text("""
            SELECT 
                u.unitup, 
                u.nama_ulp,
                MAX(d.CREATED_AT) as last_upload
            FROM tb_user u
            LEFT JOIN dpm d ON u.unitup = d.UNITUP
            WHERE u.role = 'ULP'
            GROUP BY u.unitup, u.nama_ulp
            ORDER BY u.unitup
        """)
        ulp_list = pd.read_sql(ulp_query, engine).to_dict('records')
        
        # Format last_upload
        for ulp in ulp_list:
            if ulp['last_upload']:
                ulp['last_upload'] = ulp['last_upload'].strftime('%d-%m-%Y %H:%M')
                
    except Exception as e:
        logger.error(f"Error fetching ULP list: {e}")
        ulp_list = []
    
    # Get overview all ULP
    all_summary = []
    try:
        all_query = text("""
            SELECT 
                unitup,
                blth,
                COUNT(*) as total,
                SUM(CASE WHEN ket = 'NAIK' THEN 1 ELSE 0 END) as naik,
                SUM(CASE WHEN ket = 'TURUN' THEN 1 ELSE 0 END) as turun,
                SUM(CASE WHEN ket = 'DIV/NA' THEN 1 ELSE 0 END) as div_na,
                SUM(CASE WHEN ket = 'AMAN' THEN 1 ELSE 0 END) as aman
            FROM billing
            WHERE unitup NOT LIKE '%_up3%'
            GROUP BY unitup, blth
            ORDER BY blth DESC, unitup ASC
            LIMIT 30
        """)
        all_summary = pd.read_sql(all_query, engine).to_dict('records')
        
    except Exception as e:
        logger.error(f"Error fetching all summary: {e}")
    
    return render_template(
        'dashboard_up3_selector.html',
        nama=nama,
        username=username,
        ulp_list=ulp_list,
        all_summary=all_summary
    )
    
    
# =================== DASHBOARD UP3 - Manage ULP (Pakai dashboard_ulp.html) ===================
@app.route('/dashboard_up3/manage/<unitup>', methods=['GET', 'POST'])
def dashboard_up3_manage(unitup):
    """📊 UP3 mengelola ULP tertentu (pakai template dashboard_ulp.html)"""
    if 'loggedin' not in session or session.get('role') != 'UP3':
        flash('❌ Akses ditolak', 'danger')
        return redirect(url_for('login'))
    
    username = session.get('username')
    role = 'UP3'
    
    # Get nama ULP
    try:
        ulp_info = pd.read_sql(
            text("SELECT nama_ulp FROM tb_user WHERE unitup = :unitup"),
            engine,
            params={'unitup': unitup}
        )
        nama = ulp_info.iloc[0]['nama_ulp'] if not ulp_info.empty else f"ULP {unitup}"
    except:
        nama = f"ULP {unitup}"
    
    blth_kini = request.form.get('blth', datetime.now().strftime('%Y-%m'))
    blth_normalized = blth_kini.replace('-', '')
    
    # =================== UPLOAD DPM ===================
    if request.method == 'POST' and 'file_dpm' in request.files:
        file = request.files['file_dpm']
        
        if file.filename == '':
            flash('Tidak ada file yang dipilih', 'warning')
            return redirect(url_for('dashboard_up3_manage', unitup=unitup))
        
        try:
            df_upload = pd.read_excel(file)
            df_upload.columns = [c.strip().upper() for c in df_upload.columns]
            
            # ✅ Validasi kolom UNITUP
            if 'UNITUP' not in df_upload.columns:
                flash('❌ File Excel harus memiliki kolom UNITUP untuk upload oleh UP3!', 'danger')
                return redirect(url_for('dashboard_up3_manage', unitup=unitup))
            
            # ✅ Normalisasi UNITUP TANPA suffix _up3
            df_upload['UNITUP'] = df_upload['UNITUP'].astype(str).str.strip()
            
            # ✅ Filter hanya data untuk UNITUP yang sedang dikelola
            df_upload = df_upload[df_upload['UNITUP'] == unitup]
            
            if df_upload.empty:
                flash(f'⚠️ Tidak ada data untuk UNITUP {unitup} di file Excel', 'warning')
                return redirect(url_for('dashboard_up3_manage', unitup=unitup))
            
            logger.info(f"✅ UP3 upload for {unitup}. Total rows: {len(df_upload)}")
            
            # ✅ Pass None sebagai unitup_for_upload (data sudah ada UNITUP di df)
            count, error = process_dpm_upload(df_upload, blth_normalized, None)
            
            if error:
                flash(f'❌ Gagal upload: {error}', 'danger')
            else:
                flash(f'✅ Berhasil upload {count} data DPM untuk {unitup} oleh {username} (UP3)', 'success')
                
        except Exception as e:
            logger.error(f"Upload error: {str(e)}")
            logger.error(traceback.format_exc())
            flash(f'❌ Error: {str(e)}', 'danger')
        
        return redirect(url_for('dashboard_up3_manage', unitup=unitup))
    
    # =================== UPDATE DPM KOREKSI ===================
    if request.method == 'POST' and request.form.get('form_type') == 'update_koreksi':
        logger.info("🟡 UP3 Processing UPDATE DPM KOREKSI")
        
        if 'file_dpm_koreksi' not in request.files:
            flash('❌ File update koreksi tidak ditemukan', 'danger')
            return redirect(url_for('dashboard_up3_manage', unitup=unitup))
        
        file = request.files['file_dpm_koreksi']
        logger.info(f"📄 File received: {file.filename}")
        
        if file.filename == '':
            flash('Tidak ada file yang dipilih', 'warning')
            return redirect(url_for('dashboard_up3_manage', unitup=unitup))
        
        try:
            logger.info(f"🔄 Starting DPM Koreksi for {unitup} by {username} (UP3)")
            
            df_koreksi = pd.read_excel(file)
            df_koreksi.columns = [c.strip().upper() for c in df_koreksi.columns]
            logger.info(f"📊 File: {len(df_koreksi)} rows, Columns: {list(df_koreksi.columns)}")
            
            # Validasi kolom wajib
            required_cols = ['BLTH', 'IDPEL', 'LWBPCABUT', 'LWBPPASANG', 'SAHLWBP']
            missing = [col for col in required_cols if col not in df_koreksi.columns]
            
            if missing:
                flash(f'❌ Kolom tidak lengkap: {", ".join(missing)}', 'danger')
                return redirect(url_for('dashboard_up3_manage', unitup=unitup))
            
            # ✅ UP3 mengelola ULP tertentu: PAKSA UNITUP ke target ULP
            df_koreksi['UNITUP'] = unitup
            logger.info(f"✅ UP3 managing {unitup}. Forced UNITUP to: {unitup}")
            
            # Normalisasi data
            df_koreksi['IDPEL'] = df_koreksi['IDPEL'].astype(str).str.strip().str.zfill(12).str.lower()
            df_koreksi['BLTH'] = df_koreksi['BLTH'].apply(lambda x: normalize_blth(str(x)))
            
            for col in ['LWBPCABUT', 'LWBPPASANG', 'SAHLWBP']:
                df_koreksi[col] = pd.to_numeric(df_koreksi[col], errors='coerce').fillna(0).astype(int)
            
            # Update database
            updated = failed = 0
            
            with engine.begin() as conn:
                for idx, row in df_koreksi.iterrows():
                    try:
                        # Check existing data
                        check = conn.execute(text("""
                            SELECT slalwbp, lwbppakai, COALESCE(marking_koreksi, 0) as mk
                            FROM dpm 
                            WHERE blth = :b AND idpel = :i AND unitup = :u
                        """), {'b': row['BLTH'], 'i': row['IDPEL'], 'u': unitup}).fetchone()
                        
                        if not check:
                            failed += 1
                            if idx < 3:
                                logger.warning(f"Row {idx+1}: NOT FOUND - {row['IDPEL']} in UNITUP {unitup}")
                            continue
                        
                        # Calculate LWBPPAKAI
                        lwbp_new = int(row['LWBPCABUT']) - int(check.slalwbp or 0) + int(row['SAHLWBP']) - int(row['LWBPPASANG'])
                        mk_new = check.mk + 1
                        
                        # Update
                        result = conn.execute(text("""
                            UPDATE dpm 
                            SET LWBPCABUT=:lc, LWBPPASANG=:lp, SAHLWBP=:sh, 
                                LWBPPAKAI=:lw, MARKING_KOREKSI=:mk, UPDATED_BY=:u
                            WHERE BLTH=:b AND IDPEL=:i AND UNITUP=:un
                        """), {
                            'lc': int(row['LWBPCABUT']), 'lp': int(row['LWBPPASANG']),
                            'sh': int(row['SAHLWBP']), 'lw': lwbp_new, 'mk': mk_new,
                            'u': username, 'b': row['BLTH'], 'i': row['IDPEL'], 'un': unitup
                        })
                        
                        if result.rowcount > 0:
                            updated += 1
                        else:
                            failed += 1
                        
                    except Exception as e:
                        failed += 1
                        logger.error(f"Row {idx+1} error: {e}")
            
            logger.info(f"📊 Summary for {unitup}: {updated} success, {failed} failed")
            
            if updated > 0:
                flash(f'✅ Update {updated}/{len(df_koreksi)} data untuk {unitup} (MARKING_KOREKSI +1)', 'success')
                flash('💡 Proses ulang billing untuk refresh', 'info')
            
            if failed > 0:
                flash(f'⚠️ Gagal {failed} data', 'warning')
            
        except Exception as e:
            logger.error(f"❌ CRITICAL: {str(e)}")
            logger.error(traceback.format_exc())
            flash(f'❌ Error: {str(e)}', 'danger')
        
        return redirect(url_for('dashboard_up3_manage', unitup=unitup))
    
    # =================== PROCESS BILLING ===================
    if request.method == 'POST' and request.form.get('action') == 'process_billing':
        try:
            # ✅ Proses billing hanya untuk UNITUP ini (tanpa suffix)
            df_billing, error = process_billing_advanced(blth_normalized, unitup, engine)
            
            if error:
                flash(f'❌ Gagal proses billing: {error}', 'danger')
            elif df_billing.empty:
                flash(f'⚠️ Tidak ada data untuk diproses', 'warning')
            else:
                result = save_to_billing_with_trigger(df_billing, engine, username)
                flash(f'✅ {result["message"]} untuk {unitup}', 
                      'success' if result['success'] > 0 else 'danger')
                
        except Exception as e:
            logger.error(f"Billing error: {str(e)}")
            flash(f'❌ Error: {str(e)}', 'danger')
        
        return redirect(url_for('dashboard_up3_manage', unitup=unitup))
    
    # =================== GET SUMMARY ===================
    try:
        # ✅ Query tanpa suffix _up3
        query = text("""
            SELECT 
                blth,
                COUNT(*) as total,
                SUM(CASE WHEN ket = 'NAIK' THEN 1 ELSE 0 END) as naik,
                SUM(CASE WHEN ket = 'TURUN' THEN 1 ELSE 0 END) as turun,
                SUM(CASE WHEN ket = 'DIV/NA' THEN 1 ELSE 0 END) as div_na,
                SUM(CASE WHEN ket = 'AMAN' THEN 1 ELSE 0 END) as aman
            FROM billing
            WHERE unitup = :unitup
            GROUP BY blth
            ORDER BY blth DESC
            LIMIT 6
        """)
        df_summary = pd.read_sql(query, engine, params={'unitup': unitup})

        # Konversi ke integer
        for col in ['total', 'naik', 'turun', 'div_na', 'aman']:
            if col in df_summary.columns:
                df_summary[col] = df_summary[col].fillna(0).astype(int)

    except Exception as e:
        logger.error(f"Error fetching summary: {str(e)}")
        df_summary = pd.DataFrame()
    
    # Render template dashboard_ulp.html dengan data ULP target
    return render_template(
        'dashboard_ulp.html',
        nama=f"{nama} (Dikelola oleh {username})",
        unitup=unitup,
        role='UP3_MANAGE',
        summary=df_summary.to_dict('records') if not df_summary.empty else [],
        blth_terakhir=blth_normalized,
        back_url=url_for('dashboard_up3')
    )

# # =================== QUICK UPLOAD DIL ===================
# @app.route('/dashboard_up3/quick_upload', methods=['POST'])
# def dashboard_up3_quick_upload():
#     """📤 Quick upload DIL dari card navigation"""
#     if 'loggedin' not in session or session.get('role') != 'UP3':
#         return jsonify({'error': 'Unauthorized'}), 401
    
#     target_ulp = request.form.get('target_ulp')
    
#     if 'file_dil' not in request.files:
#         flash('❌ File tidak ditemukan', 'danger')
#         return redirect(url_for('dashboard_up3'))
    
#     file = request.files['file_dil']
    
#     if file.filename == '':
#         flash('❌ Tidak ada file yang dipilih', 'warning')
#         return redirect(url_for('dashboard_up3'))
    
#     try:
#         df = pd.read_excel(file)
#         df.columns = [c.strip().upper() for c in df.columns]

#         # Validasi
#         required = ['IDPEL', 'NOMORKWH']
#         for col in required:
#             if col not in df.columns:
#                 flash(f'❌ Kolom {col} tidak ditemukan', 'danger')
#                 return redirect(url_for('dashboard_up3'))

#         df['IDPEL'] = df['IDPEL'].apply(normalize_idpel)

#         upload_cols = ['IDPEL', 'NAMA', 'TARIF', 'DAYA', 'NOMORKWH', 'ALAMAT']
#         df_upload = df[[c for c in upload_cols if c in df.columns]]

#         df_upload.to_sql('dil', engine, if_exists='append', index=False)

#         flash(f'✅ Berhasil upload {len(df_upload)} data DIL untuk {target_ulp}', 'success')

#     except Exception as e:
#         flash(f'❌ Gagal upload DIL: {str(e)}', 'danger')

#     return redirect(url_for('dashboard_up3'))

# =================== UPDATE NOMOR KWH DARI DIL ===================
@app.route('/sync_nomorkwh', methods=['POST'])
def sync_nomorkwh():
    """
    🔄 Sync nomor KWH dari tabel DIL ke billing
    Bisa manual trigger atau auto via trigger
    """
    if 'loggedin' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    unitup = session.get('unitup')
    role = session.get('role')
    
    # Filter by UNITUP (kecuali UP3)
    unitup_filter = ""
    params = {}
    
    if role == 'ULP':
        unitup_filter = "AND b.unitup = :unitup"
        params['unitup'] = unitup
    
    try:
        with engine.begin() as conn:
            # Update nomor KWH dari DIL (ambil yang terbaru)
            sql = text(f"""
                UPDATE billing b
                JOIN (
                    SELECT d.idpel, d.nomorkwh, d.created_at
                    FROM dil d
                    JOIN (
                        SELECT idpel, MAX(created_at) as max_date
                        FROM dil
                        WHERE nomorkwh IS NOT NULL AND nomorkwh != ''
                        GROUP BY idpel
                    ) latest ON d.idpel = latest.idpel AND d.created_at = latest.max_date
                ) d ON b.idpel = d.idpel
                SET b.nomorkwh = d.nomorkwh
                WHERE (b.nomorkwh IS NULL OR b.nomorkwh = '' OR b.nomorkwh IN ('-', '0'))
                {unitup_filter}
            """)
            
            result = conn.execute(sql, params)
            updated_count = result.rowcount
            
        return jsonify({
            'success': True,
            'message': f'Berhasil sync {updated_count} nomor KWH',
            'updated': updated_count
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# =================== UPLOAD DIL (Data Induk Langganan) ===================
@app.route('/upload_dil', methods=['POST'])
def upload_dil():
    """
    📤 Upload data DIL (nomor KWH, dll)
    """
    if 'loggedin' not in session:
        return redirect(url_for('login'))

    if 'file_dil' not in request.files:
        flash('Tidak ada file yang diupload', 'warning')
        return redirect(request.referrer or url_for('dashboard_ulp'))

    file = request.files['file_dil']

    if file.filename == '':
        flash('Tidak ada file yang dipilih', 'warning')
        return redirect(request.referrer or url_for('dashboard_ulp'))

    try:
        df = pd.read_excel(file)
        df.columns = [c.strip().upper() for c in df.columns]

        # ✅ Kolom wajib
        required = ['IDPEL', 'NOMORKWH']
        for col in required:
            if col not in df.columns:
                flash(f'Kolom {col} tidak ditemukan', 'danger')
                return redirect(request.referrer or url_for('dashboard_ulp'))

        # ✅ Normalisasi IDPEL
        df['IDPEL'] = df['IDPEL'].apply(normalize_idpel)

        # ✅ Pilih hanya kolom yang benar-benar ada di tabel DIL
        upload_cols = ['IDPEL', 'NAMA', 'TARIF', 'DAYA', 'NOMORKWH', 'ALAMAT']
        df_upload = df[[c for c in upload_cols if c in df.columns]]

        # ✅ Simpan ke tabel DIL tanpa kolom UNITUP
        df_upload.to_sql('dil', engine, if_exists='append', index=False)

        flash(f'Berhasil upload {len(df_upload)} data DIL', 'success')

        # ✅ Auto-sync ke billing (tanpa filter unitup)
        try:
            with engine.begin() as conn:
                sql = text("""
                    UPDATE billing b
                    JOIN dil d ON b.idpel = d.idpel
                    SET b.nomorkwh = d.nomorkwh
                    WHERE (b.nomorkwh IS NULL OR b.nomorkwh = '')
                """)
                result = conn.execute(sql)
                flash(f'Auto-sync: {result.rowcount} nomor KWH diupdate', 'info')
        except Exception as sync_err:
            flash(f'Auto-sync gagal: {str(sync_err)}', 'warning')

    except Exception as e:
        flash(f'Gagal upload DIL: {str(e)}', 'danger')

    return redirect(request.referrer or url_for('dashboard_ulp'))


# =================== VIEW AUDIT LOG ===================
@app.route('/audit_log', methods=['GET'])
def view_audit_log():
    """
    📜 View audit log perubahan data
    """
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    # Filter
    idpel_filter = request.args.get('idpel', '')
    blth_filter = request.args.get('blth', '')
    days_back = request.args.get('days', type=int, default=7)
    
    query = """
        SELECT 
            id,
            table_name,
            idpel,
            blth,
            column_changed,
            old_value,
            new_value,
            changed_by,
            changed_at
        FROM audit_log
        WHERE changed_at >= DATE_SUB(NOW(), INTERVAL :days DAY)
    """
    params = {'days': days_back}
    
    if idpel_filter:
        query += " AND idpel = :idpel"
        params['idpel'] = normalize_idpel(idpel_filter)
    
    if blth_filter:
        query += " AND blth = :blth"
        params['blth'] = blth_filter
    
    query += " ORDER BY changed_at DESC LIMIT 500"
    
    try:
        df_log = pd.read_sql(text(query), engine, params=params)
    except Exception as e:
        flash(f'Gagal membaca audit log: {str(e)}', 'danger')
        df_log = pd.DataFrame()
    
    return render_template(
        'audit_log.html',
        logs=df_log.to_dict('records') if not df_log.empty else [],
        idpel_filter=idpel_filter,
        blth_filter=blth_filter,
        days_back=days_back
    )


@app.route('/grafik/<idpel>', methods=['GET'])
def view_grafik(idpel):
    """
    📈 Tampilkan grafik pemakaian 6 bulan terakhir
    """
    idpel = normalize_idpel(idpel)
    blth = request.args.get('blth', datetime.now().strftime('%Y%m'))
    ulp = request.args.get('ulp', '')

    try:
        blth = normalize_blth(blth)
        
        # Generate 12 bulan terakhir
        months = [get_previous_blth(blth, i) for i in range(12)]
        months.reverse()
        
        # Query dengan filter UNITUP
        if ulp:
            query = text("""
                SELECT 
                    d.blth,
                    d.lwbppakai,
                    d.daya,
                    b.jam_nyala,
                    b.delta_pemkwh
                FROM dpm d
                LEFT JOIN billing b 
                    ON d.idpel = b.idpel 
                    AND d.blth = b.blth
                    AND b.unitup = :unitup
                WHERE d.idpel = :idpel
                  AND d.blth IN :months
                ORDER BY d.blth ASC
            """)
            
            df = pd.read_sql(query, engine, params={
                'idpel': idpel, 
                'months': tuple(months),
                'unitup': ulp
            })
        else:
            query = text("""
                SELECT 
                    d.blth,
                    d.lwbppakai,
                    d.daya,
                    b.jam_nyala,
                    b.delta_pemkwh
                FROM dpm d
                LEFT JOIN billing b 
                    ON d.idpel = b.idpel 
                    AND d.blth = b.blth
                WHERE d.idpel = :idpel
                  AND d.blth IN :months
                ORDER BY d.blth ASC
            """)
            
            df = pd.read_sql(query, engine, params={
                'idpel': idpel, 
                'months': tuple(months)
            })
        
        if df.empty:
            return f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Data Tidak Ditemukan</title>
                <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
            </head>
            <body class="bg-light">
                <div class="container mt-5">
                    <div class="alert alert-warning">
                        <h4>⚠️ Data Tidak Ditemukan</h4>
                        <p>Data tidak ditemukan untuk IDPEL: <strong>{idpel}</strong></p>
                        <button onclick="window.close()" class="btn btn-secondary">Tutup</button>
                    </div>
                </div>
            </body>
            </html>
            """, 404

        # Konversi ke numerik
        df['lwbppakai'] = pd.to_numeric(df['lwbppakai'], errors='coerce').fillna(0)
        df['daya'] = pd.to_numeric(df['daya'], errors='coerce').fillna(0)
        df['jam_nyala'] = pd.to_numeric(df['jam_nyala'], errors='coerce').fillna(0)
        df['delta_pemkwh'] = pd.to_numeric(df['delta_pemkwh'], errors='coerce').fillna(0)
        
        # Jika jam_nyala kosong, hitung dari lwbppakai
        mask_empty = df['jam_nyala'] == 0
        if mask_empty.any():
            daya_kw = df.loc[mask_empty, 'daya'].replace(0, np.nan) / 1000
            df.loc[mask_empty, 'jam_nyala'] = (
                df.loc[mask_empty, 'lwbppakai'] / daya_kw
            ).replace([np.inf, -np.inf], 0).fillna(0)
        
        # Jika delta kosong, hitung dari selisih
        mask_empty_delta = df['delta_pemkwh'] == 0
        if mask_empty_delta.any():
            df.loc[mask_empty_delta, 'delta_pemkwh'] = df['lwbppakai'].diff().fillna(0)

        return render_template(
            'grafik.html',
            idpel=idpel,
            labels=df['blth'].tolist(),
            lwbppakai=df['lwbppakai'].astype(int).tolist(),
            jam_nyala=df['jam_nyala'].round(1).tolist(),
            delta=df['delta_pemkwh'].astype(int).tolist()
        )

    except Exception as e:
        error_msg = str(e)
        logger.error(f"Error grafik untuk IDPEL {idpel}: {error_msg}")
        logger.error(traceback.format_exc())
        
        return f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Error</title>
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
        </head>
        <body class="bg-light">
            <div class="container mt-5">
                <div class="alert alert-danger">
                    <h4>❌ Error</h4>
                    <p><strong>IDPEL:</strong> {idpel}</p>
                    <p><strong>Error:</strong> {error_msg}</p>
                    <hr>
                    <button onclick="window.history.back()" class="btn btn-secondary">Kembali</button>
                    <button onclick="window.close()" class="btn btn-outline-secondary">Tutup</button>
                </div>
            </div>
        </body>
        </html>
        """, 500



# =================== HELPER: CLEANUP OLD DATA ===================
def cleanup_old_dpm(months=6):
    """
    🧹 Hapus data DPM lebih dari X bulan
    Bisa dipanggil manual atau via scheduled job
    """
    try:
        cutoff_date = datetime.now() - relativedelta(months=months)
        cutoff_blth = cutoff_date.strftime('%Y%m')
        
        with engine.begin() as conn:
            result = conn.execute(
                text("DELETE FROM dpm WHERE CAST(blth AS UNSIGNED) < :cutoff"),
                {'cutoff': int(cutoff_blth)}
            )
            
        return result.rowcount
        
    except Exception as e:
        print(f"Error cleanup DPM: {e}")
        return 0


# # =================== LOGIN ROUTE ===================
# @app.route('/', methods=['GET', 'POST'])
# def login():
#     """
#     🔐 Halaman login - Route utama
#     """
#     # Jika sudah login, redirect ke dashboard
#     if 'loggedin' in session:
#         if session.get('role') == 'UP3':
#             return redirect(url_for('dashboard_up3'))
#         else:
#             return redirect(url_for('dashboard_ulp'))
    
#     if request.method == 'POST':
#         username = request.form.get('username', '').strip()
#         password = request.form.get('password', '')
        
#         if not username or not password:
#             flash('Username dan password harus diisi', 'danger')
#             return render_template('login.html')
        
#         try:
#             # Query user dari database
#             query = text("""
#                 SELECT id_user, username, password, unitup, nama_ulp, role
#                 FROM tb_user
#                 WHERE username = :username
#             """)
            
#             with engine.connect() as conn:
#                 result = conn.execute(query, {'username': username}).fetchone()
            
#             if result:
#                 # Verify password (SHA256)
#                 import hashlib
#                 hashed_input = hashlib.sha256(password.encode()).hexdigest()
                
#                 if hashed_input == result[2]:  # result[2] = password column
#                     # Set session
#                     session['loggedin'] = True
#                     session['id_user'] = result[0]
#                     session['username'] = result[1]
#                     session['unitup'] = result[3]
#                     session['nama_ulp'] = result[4]
#                     session['role'] = result[5] or 'ULP'
                    
#                     flash(f'Login berhasil! Selamat datang, {result[4]}', 'success')
                    
#                     # Redirect berdasarkan role
#                     if session['role'] == 'UP3':
#                         return redirect(url_for('dashboard_up3'))
#                     else:
#                         return redirect(url_for('dashboard_ulp'))
#                 else:
#                     flash('Password salah!', 'danger')
#             else:
#                 flash('Username tidak ditemukan!', 'danger')
                
#         except Exception as e:
#             flash(f'Error login: {str(e)}', 'danger')
#             print(f"Login error: {e}")
    
#     return render_template('login.html')

@app.route('/', methods=['GET', 'POST'])
def login():
    """
    🔐 Halaman login - Route utama
    """
    # Jika sudah login, redirect ke dashboard
    if 'loggedin' in session:
        if session.get('role') == 'UP3':
            return redirect(url_for('dashboard_up3'))
        else:
            return redirect(url_for('dashboard_ulp'))
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        if not username or not password:
            flash('Username dan password harus diisi', 'danger')
            return render_template('login.html')
        
        try:
            # Query user dari database
            query = text("""
                SELECT id_user, username, password, unitup, nama_ulp, role
                FROM tb_user
                WHERE username = :username
            """)
            
            with engine.connect() as conn:
                result = conn.execute(query, {'username': username}).fetchone()
            
            if result:
                # Verify password (SHA256)
                import hashlib
                hashed_input = hashlib.sha256(password.encode()).hexdigest()
                
                if hashed_input == result[2]:  # result[2] = password column
                    # ===== FIX: Set session dengan UNITUP =====
                    session['loggedin'] = True
                    session['id_user'] = result[0]
                    session['username'] = result[1]
                    session['unitup'] = str(result[3]) if result[3] else ''  # ← PENTING!
                    session['nama_ulp'] = result[4]
                    session['role'] = result[5] or 'ULP'
                    
                    # ===== DEBUG LOG =====
                    logger.info(f"✅ Login successful: {result[1]}")
                    logger.info(f"   UNITUP: {session['unitup']}")
                    logger.info(f"   Role: {session['role']}")
                    logger.info(f"   Session: {dict(session)}")
                    
                    # ===== VALIDASI UNITUP =====
                    if not session['unitup']:
                        logger.warning(f"⚠️ User {username} has empty UNITUP in database!")
                        flash('UNITUP tidak terdaftar. Hubungi administrator.', 'warning')
                    
                    flash(f'Login berhasil! Selamat datang, {result[4]}', 'success')
                    
                    # Redirect berdasarkan role
                    if session['role'] == 'UP3':
                        return redirect(url_for('dashboard_up3'))
                    else:
                        return redirect(url_for('dashboard_ulp'))
                else:
                    flash('Password salah!', 'danger')
            else:
                flash('Username tidak ditemukan!', 'danger')
                
        except Exception as e:
            flash(f'Error login: {str(e)}', 'danger')
            logger.error(f"Login error: {e}")
            import traceback
            logger.error(traceback.format_exc())
    
    return render_template('login.html')


# ===== TAMBAHAN: Route untuk fix session yang rusak =====
@app.route('/fix_session', methods=['GET'])
def fix_session():
    """
    🔧 Fix session UNITUP yang kosong
    """
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    username = session.get('username')
    
    try:
        # Re-fetch user data dari database
        query = text("""
            SELECT unitup, nama_ulp, role
            FROM tb_user
            WHERE username = :username
        """)
        
        with engine.connect() as conn:
            result = conn.execute(query, {'username': username}).fetchone()
        
        if result:
            # Update session
            session['unitup'] = str(result[0]) if result[0] else ''
            session['nama_ulp'] = result[1]
            session['role'] = result[2] or 'ULP'
            
            flash(f'✅ Session diperbaiki! UNITUP: {session["unitup"]}', 'success')
            logger.info(f"Session fixed for {username}: UNITUP={session['unitup']}")
        else:
            flash('User tidak ditemukan di database', 'danger')
        
    except Exception as e:
        flash(f'Gagal memperbaiki session: {str(e)}', 'danger')
        logger.error(f"Fix session error: {e}")
    
    return redirect(url_for('view_billing'))


# ===== TAMBAHAN: Debug session route =====
@app.route('/debug_session', methods=['GET'])
def debug_session():
    """
    🔍 Debug: Show current session
    """
    return jsonify({
        "session": dict(session),
        "keys": list(session.keys()),
        "unitup": session.get('unitup'),
        "unitup_type": type(session.get('unitup')).__name__,
        "unitup_empty": not session.get('unitup'),
        "username": session.get('username'),
        "role": session.get('role')
    })

@app.route('/logout')
def logout():
    """
    🚪 Logout user
    """
    session.clear()
    flash('Anda telah logout', 'success')
    return redirect(url_for('login'))


# =================== KELOLA USER (HANYA UP3) ===================
@app.route('/kelola_user', methods=['GET'])
def kelola_user():
    """
    👥 Kelola user - hanya untuk role UP3
    """
    if 'loggedin' not in session:
        return redirect(url_for('login'))
    
    if session.get('role') != 'UP3':
        flash('Akses ditolak. Hanya untuk user UP3.', 'danger')
        return redirect(url_for('dashboard_ulp'))
    
    try:
        query = text("""
            SELECT id_user, username, unitup, nama_ulp, role
            FROM tb_user
            ORDER BY unitup ASC
        """)
        
        users = pd.read_sql(query, engine)
        
    except Exception as e:
        flash(f'Gagal membaca data user: {str(e)}', 'danger')
        users = pd.DataFrame()
    
    return render_template(
        'kelola_user.html',
        users=users.to_dict('records') if not users.empty else []
    )


@app.route('/tambah_user', methods=['POST'])
def tambah_user():
    """
    ➕ Tambah user baru
    """
    if 'loggedin' not in session or session.get('role') != 'UP3':
        flash('Akses ditolak', 'danger')
        return redirect(url_for('login'))
    
    unitup = request.form.get('unitup', '').strip()
    nama_ulp = request.form.get('nama_ulp', '').strip()
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '')
    role = request.form.get('role', 'ULP')
    
    if not all([unitup, username, password]):
        flash('UNITUP, Username, dan Password harus diisi', 'danger')
        return redirect(url_for('kelola_user'))
    
    try:
        import hashlib
        hashed_pw = hashlib.sha256(password.encode()).hexdigest()
        
        with engine.begin() as conn:
            # Check if username already exists
            check = conn.execute(
                text("SELECT COUNT(*) FROM tb_user WHERE username = :username"),
                {'username': username}
            ).scalar()
            
            if check > 0:
                flash('Username sudah digunakan!', 'warning')
                return redirect(url_for('kelola_user'))
            
            # Check if unitup already exists (kecuali role UP3)
            if role == 'ULP':
                check_unitup = conn.execute(
                    text("SELECT COUNT(*) FROM tb_user WHERE unitup = :unitup"),
                    {'unitup': unitup}
                ).scalar()
                
                if check_unitup > 0:
                    flash('UNITUP sudah memiliki user!', 'warning')
                    return redirect(url_for('kelola_user'))
            
            # Insert user
            conn.execute(text("""
                INSERT INTO tb_user (unitup, nama_ulp, username, password, role)
                VALUES (:unitup, :nama_ulp, :username, :password, :role)
            """), {
                'unitup': unitup,
                'nama_ulp': nama_ulp,
                'username': username,
                'password': hashed_pw,
                'role': role
            })
            
        flash('User berhasil ditambahkan!', 'success')
        
    except Exception as e:
        flash(f'Gagal menambah user: {str(e)}', 'danger')
    
    return redirect(url_for('kelola_user'))


@app.route('/hapus_user/<int:id_user>')
def hapus_user(id_user):
    """
    🗑️ Hapus user
    """
    if 'loggedin' not in session or session.get('role') != 'UP3':
        flash('Akses ditolak', 'danger')
        return redirect(url_for('login'))
    
    try:
        with engine.begin() as conn:
            conn.execute(
                text("DELETE FROM tb_user WHERE id_user = :id_user"),
                {'id_user': id_user}
            )
        
        flash('User berhasil dihapus!', 'success')
        
    except Exception as e:
        flash(f'Gagal menghapus user: {str(e)}', 'danger')
    
    return redirect(url_for('kelola_user'))


@app.route("/download_excel")
def download_excel():
    try:
        # ===================== 🔹 Ambil Parameter dari Request & Session =====================
        username = session.get('username', '').lower()
        tab = request.args.get("tab", "")
        blth = request.args.get("blth", "")
        kdkelompok = request.args.get("kdkelompok", "")
        unitup_filter = request.args.get("unitup_filter", "")
        jam_nyala_min = request.args.get("jam_nyala_min", "")
        jam_nyala_max = request.args.get("jam_nyala_max", "")

        if not tab or not blth:
            return jsonify({"error": "Parameter tab dan blth harus diisi"}), 400

        # ===================== 🔹 Ambil Data User dari Database =====================
        with engine.connect() as conn:
            user_query = text("""
                SELECT unitup, role 
                FROM tb_user 
                WHERE LOWER(username) = :username
            """)
            user_result = conn.execute(user_query, {"username": username}).fetchone()

        if not user_result:
            return jsonify({"error": "User tidak ditemukan"}), 403

        user_unitup, user_role = user_result

        # ===================== 🔹 Bangun WHERE Clause =====================
        where_conditions = ["BLTH = :blth"]
        params = {"blth": blth}

        # Filter berdasarkan role
        if user_role == 'UP3':
            if unitup_filter:
                where_conditions.append("UNITUP = :unitup")
                params["unitup"] = unitup_filter
        elif user_role == 'ULP':
            if user_unitup:
                where_conditions.append("UNITUP = :unitup")
                params["unitup"] = user_unitup
            else:
                return jsonify({"error": "User tidak memiliki UNITUP"}), 403
        else:
            return jsonify({"error": "Role tidak valid"}), 403

        # Filter kelompok pelanggan
        if kdkelompok:
            where_conditions.append("KDKELOMPOK = :kdkelompok")
            params["kdkelompok"] = kdkelompok

        # Filter berdasarkan tab aktif
        if tab == "dlpd_3bln":
            where_conditions.append("DLPD_3BLN = 'Naik50% R3BLN'")
        elif tab == "naik":
            where_conditions.append("KET = 'NAIK'")
        elif tab == "turun":
            where_conditions.append("KET = 'TURUN'")
        elif tab == "div":
            where_conditions.append("(KET = 'DIV/NA' OR KET IS NULL)")
        elif tab == "jam_nyala":
            if jam_nyala_min:
                where_conditions.append("JAM_NYALA >= :jam_nyala_min")
                params["jam_nyala_min"] = float(jam_nyala_min)
            if jam_nyala_max:
                where_conditions.append("JAM_NYALA <= :jam_nyala_max")
                params["jam_nyala_max"] = float(jam_nyala_max)

        where_clause = " AND ".join(where_conditions)

        # ===================== 🔹 Query Data dari Tabel Billing =====================
        query = text(f"""
            SELECT * 
            FROM billing
            WHERE {where_clause}
            ORDER BY UNITUP, KDKELOMPOK, IDPEL
        """)
        df = pd.read_sql(query, engine, params=params)

        if df.empty:
            return jsonify({"error": "Tidak ada data untuk didownload"}), 404

        # ===================== 🔹 Hapus Kolom yang Tidak Diperlukan =====================
        columns_to_exclude = ['updated_by', 'created_at', 'id']
        df = df.drop(columns=[col for col in columns_to_exclude if col in df.columns], errors='ignore')

        # ===================== 🔹 Buat Workbook Excel =====================
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # Header (tebal & background)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        ws.append(df.columns.tolist())
        for col_num, _ in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        # Identifikasi kolom yang berisi foto/hyperlink
        foto_columns = ["GRAFIK", "FOTO_AKHIR", "FOTO_LALU", "FOTO_LALU2", "FOTO_3BLN"]
        foto_indexes = [i for i, col in enumerate(df.columns) if col.upper() in foto_columns]

        # ===================== 🔹 Isi Data ke Excel =====================
        for i, row in df.iterrows():
            for j, value in enumerate(row):
                cell = ws.cell(row=i + 2, column=j + 1)

                if j in foto_indexes:
                    value_str = str(value).strip() if value else ""

                    if not value_str:
                        cell.value = "TIDAK ADA FOTO"
                        cell.font = Font(color="FF0000", italic=True)
                        continue

                    # --- 🔍 Ekstrak URL dari <a> atau <button> ---
                    url = None

                    # Jika HTML mengandung 'window.open'
                    if "window.open" in value_str:
                        import re
                        match = re.search(r"window\.open\('([^']+)'", value_str)
                        if match:
                            url = match.group(1)

                    # Jika HTML <a href=...>
                    if not url and "<a" in value_str:
                        soup = BeautifulSoup(value_str, "html.parser")
                        a_tag = soup.find("a")
                        if a_tag and a_tag.has_attr("href"):
                            url = a_tag["href"]

                    # Jika string langsung URL
                    if not url and value_str.startswith("http"):
                        url = value_str

                    # --- 📎 Isi ke Excel ---
                    if url:
                        cell.value = "LINK FOTO"
                        cell.hyperlink = url
                        cell.font = Font(color="0000EE", underline="single")
                    else:
                        cell.value = "TIDAK ADA FOTO"
                        cell.font = Font(color="FF0000", italic=True)

                else:
                    # Kolom JAM_NYALA diformat agar desimal (bukan integer)
                    if df.columns[j].upper() == "JAM_NYALA":
                        try:
                            cell.value = float(value) if value is not None else None
                            cell.number_format = '0.00'
                        except:
                            cell.value = value
                    else:
                        cell.value = str(value) if value is not None else ""

        # ===================== 🔹 Auto Width untuk Kolom =====================
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        # ===================== 🔹 Simpan File Excel ke Memory =====================
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # ===================== 🔹 Nama File Dinamis =====================
        unitup_name = params.get('unitup', 'ALL_UP3')
        filename = f"billing_{unitup_name}_{tab}_{blth}"
        if kdkelompok:
            filename += f"_kel{kdkelompok}"
        filename += ".xlsx"

        # ===================== 🔹 Kirim File ke Browser =====================
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print(f"❌ Error download Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Gagal membuat file Excel: {str(e)}"}), 500
    


@app.route("/download_billing")
def download_billing():
    try:
        # ===================== 🔹 Ambil Parameter =====================
        blth = request.args.get("blth")
        if not blth:
            return "Parameter blth wajib", 400

        # ===================== 🔹 Query Database =====================
        query = text("""
            SELECT *
            FROM billing
            WHERE blth = :blth
            AND DLPD_HITUNG IS NOT NULL
            AND TRIM(DLPD_HITUNG) <> ''
            ORDER BY UNITUP, IDPEL
        """)



        df = pd.read_sql(query, engine, params={"blth": blth})

        if df.empty:
            return "Data tidak ditemukan", 404

        # ===================== 🔹 Hapus Kolom Tidak Perlu =====================
        columns_to_exclude = ['updated_by', 'created_at', 'id']
        df = df.drop(columns=[c for c in columns_to_exclude if c in df.columns], errors="ignore")

        # ===================== 🔹 Buat Workbook Excel =====================
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # Header Style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        ws.append(df.columns.tolist())
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        # ===================== 🔹 Identifikasi Kolom Foto =====================
        foto_columns = ["GRAFIK", "FOTO_AKHIR", "FOTO_LALU", "FOTO_LALU2", "FOTO_3BLN"]
        foto_indexes = [i for i, col in enumerate(df.columns) if col.upper() in foto_columns]

        # ===================== 🔹 Isi Data ke Excel =====================
        for i, row in df.iterrows():
            for j, value in enumerate(row):
                cell = ws.cell(row=i + 2, column=j + 1)

                if j in foto_indexes:
                    value_str = str(value).strip() if value else ""

                    if not value_str:
                        cell.value = "TIDAK ADA FOTO"
                        cell.font = Font(color="FF0000", italic=True)
                        continue

                    # 🔍 Ekstraksi URL
                    url = None

                    # window.open('url')
                    if "window.open" in value_str:
                        match = re.search(r"window\.open\('([^']+)'", value_str)
                        if match:
                            url = match.group(1)

                    # <a href="...">
                    if not url and "<a" in value_str:
                        soup = BeautifulSoup(value_str, "html.parser")
                        a_tag = soup.find("a")
                        if a_tag and a_tag.has_attr("href"):
                            url = a_tag["href"]

                    # URL langsung
                    if not url and value_str.startswith("http"):
                        url = value_str

                    # --- 📎 Isi ke Excel ---
                    if url:
                        cell.value = "LINK FOTO"
                        cell.hyperlink = url
                        cell.font = Font(color="0000EE", underline="single")
                    else:
                        cell.value = "TIDAK ADA FOTO"
                        cell.font = Font(color="FF0000", italic=True)

                else:
                    cell.value = value if value is not None else ""

        # ===================== 🔹 Auto Width Kolom =====================
        for col_cells in ws.columns:
            ws.column_dimensions[col_cells[0].column_letter].width = (
                max(len(str(cell.value)) if cell.value else 0 for cell in col_cells) + 2
            )

        # ===================== 🔹 Simpan ke Memory =====================
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # ===================== 🔹 Kirim File =====================
        return send_file(
            output,
            as_attachment=True,
            download_name=f"billing_gabung_{blth}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Gagal export Excel: {str(e)}"}), 500

    
# 📥 DOWNLOAD EXCEL PER HARI BACA - FIXED VERSION WITH MULTIPLE UNITUP FILTER
# ==========================
@app.route("/download_excel_hb", methods=["GET"])
def download_excel_hb():
    try:
        # ===================== 🔹 Ambil Parameter dari Request & Session =====================
        username = session.get('username', '').lower()
        input1 = request.args.get("input1", "").strip().upper()
        input2 = request.args.get("input2", "").strip().upper()
        blth = request.args.get("blth", "").strip()
        
        # ⭐ PERBAIKAN 1: Ambil unitup_filter dari URL - Support Multiple Values
        unitup_filter_raw = request.args.getlist("unitup_filter")  # Array dari checkbox
        if not unitup_filter_raw:
            # Fallback ke single value dari dropdown
            single_value = request.args.get("unitup_filter", "").strip()
            unitup_filter_raw = [single_value] if single_value else []
        
        # Hapus empty strings dan "Semua"
        unitup_filters = [u.strip() for u in unitup_filter_raw 
                         if u.strip() and u.strip().lower() not in ('semua', '(select all)')]
        
        # ⭐ PERBAIKAN 2: Ambil tab untuk filter KET jika diperlukan
        active_tab = request.args.get("tab", "").strip().lower()
        
        # ⭐ DEBUG: Log semua parameter yang diterima
        logger.info(f"📥 Download HB Request - user: {username}, blth: {blth}, tab: {active_tab}")
        logger.info(f"📥 Range: {input1} to {input2}")
        logger.info(f"📥 UNITUP filters: {unitup_filters}")

        if not input1 or not input2 or not blth:
            return jsonify({"error": "Parameter input1, input2, dan blth harus diisi"}), 400

        # ===================== 🔹 Ambil Data User dari Database =====================
        with engine.connect() as conn:
            user_query = text("""
                SELECT unitup, role 
                FROM tb_user 
                WHERE LOWER(username) = :username
            """)
            user_result = conn.execute(user_query, {"username": username}).fetchone()

        if not user_result:
            return jsonify({"error": "User tidak ditemukan"}), 403

        user_unitup, user_role = user_result

        # ===================== 🔹 Validasi Hari Baca =====================
        kelompok_order = ['1','2','3','4','5','6','7','8','P','A','I']

        if input1 not in kelompok_order or input2 not in kelompok_order:
            return jsonify({"error": "Input KDKELOMPOK tidak valid"}), 400

        idx1, idx2 = kelompok_order.index(input1), kelompok_order.index(input2)
        if idx1 > idx2:
            idx1, idx2 = idx2, idx1
        kelompok_range = kelompok_order[idx1:idx2 + 1]

        # ===================== 🔹 Bangun WHERE Clause =====================
        where_conditions = ["BLTH = :blth"]
        params = {"blth": blth}

        # ⭐ PERBAIKAN 3: Filter berdasarkan role dan unitup_filters (multiple)
        if user_role == "UP3":
            # UP3: Filter berdasarkan unitup_filters jika ada
            if unitup_filters:
                # ⭐ Support multiple UNITUP dari checkbox
                if len(unitup_filters) == 1:
                    # Single UNITUP
                    where_conditions.append("UNITUP = :unitup")
                    params["unitup"] = unitup_filters[0]
                    logger.info(f"✅ UP3 download dengan filter UNITUP: {unitup_filters[0]}")
                else:
                    # Multiple UNITUP
                    where_conditions.append("UNITUP IN :unitup_list")
                    params["unitup_list"] = tuple(unitup_filters)
                    logger.info(f"✅ UP3 download dengan filter UNITUP: {', '.join(unitup_filters)}")
            else:
                # Jika tidak ada filter, UP3 bisa lihat semua data
                logger.info(f"✅ UP3 download semua data (no UNITUP filter)")
            
            where_conditions.append("KDKELOMPOK IN :kelompok_range")
            params["kelompok_range"] = tuple(kelompok_range)
            
        elif user_role == "ULP":
            # ULP: Filter hanya data ULP sendiri
            if not user_unitup:
                return jsonify({"error": "User tidak memiliki UNITUP"}), 403
            
            where_conditions.append("UNITUP = :unitup")
            where_conditions.append("KDKELOMPOK IN :kelompok_range")
            params["unitup"] = user_unitup
            params["kelompok_range"] = tuple(kelompok_range)
            logger.info(f"✅ ULP download untuk UNITUP: {user_unitup}")
        else:
            return jsonify({"error": "Role tidak valid"}), 403

        # ⭐ PERBAIKAN 4: Tambahkan filter berdasarkan tab
        if active_tab:
            if active_tab in ('naik', 'turun', 'div'):
                # Filter berdasarkan kolom KET
                ket_mapping = {
                    'naik': 'NAIK',
                    'turun': 'TURUN',
                    'div': 'DIV/NA'
                }
                where_conditions.append("KET = :ket")
                params["ket"] = ket_mapping[active_tab]
                logger.info(f"✅ Filter KET: {ket_mapping[active_tab]}")
                
            elif active_tab == 'dlpd_3bln':
                # Filter berdasarkan kolom DLPD_3BLN - hanya yang mengandung "Naik"
                where_conditions.append("DLPD_3BLN LIKE :dlpd_pattern")
                params["dlpd_pattern"] = "%Naik%"
                logger.info(f"✅ Filter DLPD_3BLN: Hanya data yang mengandung 'Naik'")
                
            elif active_tab == 'jam_nyala':
                # Filter berdasarkan kolom JAM_NYALA (hanya yang punya nilai)
                where_conditions.append("JAM_NYALA IS NOT NULL")
                logger.info(f"✅ Filter JAM_NYALA: Data dengan JAM_NYALA tidak null")

        where_clause = " AND ".join(where_conditions)

        # ⭐ DEBUG: Cek dulu apakah data ada (tanpa filter KET)
        debug_conditions = [c for c in where_conditions if 'KET' not in c]
        debug_clause = " AND ".join(debug_conditions)
        debug_params = {k: v for k, v in params.items() if k != 'ket'}
        
        try:
            debug_query = text(f"""
                SELECT COUNT(*) as total, 
                       COUNT(DISTINCT UNITUP) as unitup_count
                FROM billing
                WHERE {debug_clause}
            """)
            
            debug_result = pd.read_sql(debug_query, engine, params=debug_params)
            logger.info(f"🔍 Debug count - Total: {debug_result.iloc[0]['total']}, "
                       f"UNITUP count: {debug_result.iloc[0]['unitup_count']}")
            
            # Ambil daftar UNITUP secara terpisah
            try:
                unitup_query = text(f"""
                    SELECT DISTINCT UNITUP 
                    FROM billing
                    WHERE {debug_clause}
                    LIMIT 10
                """)
                unitup_list = pd.read_sql(unitup_query, engine, params=debug_params)
                if not unitup_list.empty:
                    logger.info(f"🔍 Available UNITUPs: {', '.join(unitup_list['UNITUP'].tolist())}")
            except Exception as e:
                logger.warning(f"⚠️ Could not fetch UNITUP list: {e}")
            
            if debug_result.iloc[0]['total'] == 0:
                logger.error(f"❌ No data found even without KET filter!")
                return jsonify({"error": f"Tidak ada data untuk BLTH {blth} dengan filter yang dipilih. "
                                        f"Pastikan data sudah diupload untuk UNITUP dan KDKELOMPOK yang diminta."}), 404
        except Exception as e:
            logger.error(f"❌ Debug query error: {str(e)}")
            # Lanjutkan proses meskipun debug gagal

        # ===================== 🔹 Query Data =====================
        query = text(f"""
            SELECT * 
            FROM billing
            WHERE {where_clause}
            ORDER BY UNITUP, KDKELOMPOK, IDPEL
        """)
        
        logger.info(f"📊 Executing query with params: {params}")
        df = pd.read_sql(query, engine, params=params)

        if df.empty:
            logger.warning(f"⚠️ No data found for params: {params}")
            
            # ⭐ Cek apakah data ada jika filter KET dihilangkan
            if 'ket' in params:
                try:
                    no_ket_query = text(f"""
                        SELECT COUNT(*) as total,
                               COUNT(CASE WHEN KET = :ket THEN 1 END) as ket_count
                        FROM billing
                        WHERE {debug_clause}
                    """)
                    ket_check = pd.read_sql(no_ket_query, engine, params=params)
                    
                    # Ambil daftar KET yang tersedia
                    ket_list_query = text(f"""
                        SELECT DISTINCT KET, COUNT(*) as count
                        FROM billing
                        WHERE {debug_clause}
                        GROUP BY KET
                    """)
                    ket_list = pd.read_sql(ket_list_query, engine, params=debug_params)
                    available_kets = ', '.join([f"{row['KET']}({row['count']})" for _, row in ket_list.iterrows()])
                    
                    logger.info(f"🔍 KET Check - Total data: {ket_check.iloc[0]['total']}, "
                               f"Data with KET={params['ket']}: {ket_check.iloc[0]['ket_count']}, "
                               f"Available KETs: {available_kets}")
                    
                    return jsonify({
                        "error": f"Tidak ada data dengan KET={params['ket']} untuk filter yang dipilih. "
                                f"Total data tersedia: {ket_check.iloc[0]['total']}. "
                                f"KET tersedia: {available_kets}"
                    }), 404
                except Exception as e:
                    logger.error(f"❌ KET check error: {str(e)}")
            
            return jsonify({"error": "Tidak ada data untuk hari baca yang dipilih"}), 404

        logger.info(f"✅ Found {len(df)} rows")

        # ===================== 🔹 Hapus Kolom Tidak Diperlukan =====================
        columns_to_exclude = ['updated_by', 'created_at', 'id']
        df = df.drop(columns=[col for col in columns_to_exclude if col in df.columns], errors='ignore')

        # ===================== 🔹 Buat Workbook Excel =====================
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        ws.append(df.columns.tolist())
        for col_num, _ in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        # Kolom yang berisi link foto
        foto_columns = ["GRAFIK", "FOTO_AKHIR", "FOTO_LALU", "FOTO_LALU2", "FOTO_3BLN"]
        foto_indexes = [i for i, col in enumerate(df.columns) if col.upper() in foto_columns]

        # ===================== 🔹 Isi Data =====================
        for i, row in df.iterrows():
            for j, value in enumerate(row):
                cell = ws.cell(row=i + 2, column=j + 1)

                if j in foto_indexes:
                    value_str = str(value).strip() if value else ""
                    if not value_str:
                        cell.value = "TIDAK ADA FOTO"
                        cell.font = Font(color="FF0000", italic=True)
                        continue

                    # Cek apakah ada link
                    url = None
                    if "window.open" in value_str:
                        import re
                        match = re.search(r"window\.open\('([^']+)'", value_str)
                        if match:
                            url = match.group(1)
                    if not url and "<a" in value_str:
                        soup = BeautifulSoup(value_str, "html.parser")
                        a_tag = soup.find("a")
                        if a_tag and a_tag.has_attr("href"):
                            url = a_tag["href"]
                    if not url and value_str.startswith("http"):
                        url = value_str

                    if url:
                        cell.value = "LINK FOTO"
                        cell.hyperlink = url
                        cell.font = Font(color="0000EE", underline="single")
                    else:
                        cell.value = "TIDAK ADA FOTO"
                        cell.font = Font(color="FF0000", italic=True)
                else:
                    # Kolom JAM_NYALA diformat jadi desimal
                    if df.columns[j].upper() == "JAM_NYALA":
                        try:
                            cell.value = float(value) if value is not None else None
                            cell.number_format = '0.00'
                        except:
                            cell.value = value
                    else:
                        cell.value = str(value) if value is not None else ""

        # ===================== 🔹 Auto Width =====================
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        # ===================== 🔹 Simpan ke Memory =====================
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # ⭐ PERBAIKAN 5: Nama File Dinamis dengan info tab dan filter
        if user_role == "UP3":
            if unitup_filters:
                # Gunakan nama UNITUP yang dipilih
                if len(unitup_filters) == 1:
                    unitup_name = f"{unitup_filters[0]}_UP3"
                else:
                    # Multiple UNITUP - gabungkan nama (max 3)
                    if len(unitup_filters) <= 3:
                        unitup_name = "_".join(unitup_filters) + "_UP3"
                    else:
                        unitup_name = f"MULTI_{len(unitup_filters)}UNITUP_UP3"
            else:
                unitup_name = "ALL_UP3"
        else:
            unitup_name = user_unitup if user_unitup else "UNKNOWN"
        
        tab_suffix = f"_{active_tab}" if active_tab else ""
        filename = f"billing_{unitup_name}_hb_{input1}_to_{input2}_{blth}{tab_suffix}.xlsx"

        logger.info(f"✅ Generating file: {filename}")

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"❌ Error download Excel Hari Baca: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Gagal membuat file Excel: {str(e)}"}), 500
    
    
    
#############
@app.route('/simpan_dlpd', methods=['POST'])
# @login_required
def simpan_dlpd():
    try:
        active_tab = request.form.get("active_tab", "dlpd_3bln")  # fallback ke tab ini kalau kosong
        username = session.get('username', '').lower()
        billing_table = TABLE_MAPPING.get(username)

        if not billing_table:
            flash("User tidak dikenali.", "danger")
            return redirect(url_for('view_data1_v2', tab=active_tab))

        selected_kelompok = request.form.get("kdkelompok")
        print(f"[DEBUG] POST kdkelompok (from form): {selected_kelompok}")

        form_data = request.form
        print(f"Form Data: {form_data}")
        selected_kelompok = form_data.get("kdkelompok")
        idpels = {field.replace('hasil_pemeriksaan_', '') for field in form_data if field.startswith('hasil_pemeriksaan_')}
        idpels.update({field.replace('tindak_lanjut_', '') for field in form_data if field.startswith('tindak_lanjut_')})
        idpels.update({field.replace('stan_verifikasi_', '') for field in form_data if field.startswith('stan_verifikasi_')})

        updated_count = 0
        with engine.begin() as connection:
            for idpel in idpels:
                hasil = form_data.get(f'hasil_pemeriksaan_{idpel}', '').strip()
                tindak = form_data.get(f'tindak_lanjut_{idpel}', '').strip()
                stan = form_data.get(f'stan_verifikasi_{idpel}', '').strip()

                if hasil or tindak:
                    query = f"""
                    UPDATE {billing_table}
                    SET `HASIL PEMERIKSAAN` = :hasil,
                        `TINDAK LANJUT` = :tindak,
                        STAN_VERIFIKASI = :stan
                    WHERE IDPEL = :id 
                      AND DLPD_3BLN = 'Naik50% R3BLN'
                      AND KDKELOMPOK = :kelompok
                     """
                    connection.execute(text(query), {
                        'hasil': hasil,
                        'tindak': tindak,
                        'stan': stan,
                        'id': idpel,
                        'kelompok': selected_kelompok
                    })
                    updated_count += 1
        flash(f"{updated_count} data berhasil diperbarui!", "success")

    except Exception as e:
        import logging
        logging.exception("Error in simpan_dlpd")
        flash(f"Gagal memperbarui data DLPD: {str(e)}", "danger")

    return redirect(url_for('view_data1_v2', tab=active_tab, kdkelompok=selected_kelompok))

def simpan_by_ket(filter_key):
    try:
        active_tab = request.form.get("active_tab", "dlpd_3bln")  # Get active tab from form
        username = session.get('username', '').lower()
        billing_table = TABLE_MAPPING.get(username)

        if not billing_table:
            flash("User tidak dikenali.", "danger")
            return redirect(url_for('view_data1_v2', tab=active_tab))

        selected_kelompok = request.form.get("kdkelompok")
        print(f"[DEBUG] POST kdkelompok (from form): {selected_kelompok}")

        form_data = request.form
        print(f"Form Data: {form_data}")
        selected_kelompok = form_data.get("kdkelompok")
        idpels = {field.replace('hasil_pemeriksaan_', '') for field in form_data if field.startswith('hasil_pemeriksaan_')}
        idpels.update({field.replace('tindak_lanjut_', '') for field in form_data if field.startswith('tindak_lanjut_')})
        idpels.update({field.replace('stan_verifikasi_', '') for field in form_data if field.startswith('stan_verifikasi_')})
        updated_count = 0
        with engine.begin() as connection:
            for idpel in idpels:
                hasil = form_data.get(f'hasil_pemeriksaan_{idpel}', '').strip()
                tindak = form_data.get(f'tindak_lanjut_{idpel}', '').strip()
                stan = form_data.get(f'stan_verifikasi_{idpel}', '').strip()
                
                if hasil or tindak:
                    if filter_key == "JAMNYALA600":
                        query = f"""
                            UPDATE {billing_table}
                            SET `HASIL PEMERIKSAAN` = :hasil,
                                `TINDAK LANJUT` = :tindak,
                                STAN_VERIFIKASI = :stan
                            WHERE IDPEL = :id AND JAMNYALA600 = '600Up'
                        """
                    else:
                        query = f"""
                            UPDATE {billing_table}
                            SET `HASIL PEMERIKSAAN` = :hasil,
                                `TINDAK LANJUT` = :tindak,
                                STAN_VERIFIKASI = :stan
                            WHERE IDPEL = :id AND KET = :ket
                        """
                    connection.execute(text(query), {
                        'hasil': hasil,
                        'tindak': tindak,
                        'stan':stan,
                        'id': idpel,
                        'ket': filter_key if filter_key != "JAMNYALA600" else None
                    })
                    updated_count += 1

        flash(f"{updated_count} data berhasil diperbarui untuk kategori {filter_key}!", "success")
    except Exception as e:
        import logging
        logging.exception(f"Error in simpan_by_ket ({filter_key})")
        flash(f"Gagal memperbarui data {filter_key}", "danger")

    
    return redirect(request.referrer or url_for('view_data1_v2'))


@app.route('/simpan_naik', methods=['POST'])
# @login_required
def simpan_naik():
    return simpan_by_ket("NAIK")

@app.route('/simpan_turun', methods=['POST'])
# @login_required
def simpan_turun():
    return simpan_by_ket("TURUN")

@app.route('/simpan_div', methods=['POST'])
# @login_required
def simpan_div():
    return simpan_by_ket("DIV/NA")

@app.route('/simpan_aman', methods=['POST'])
# @login_required
def simpan_aman():
    return simpan_by_ket("AMAN")

# @app.route('/simpan_jam_nyala', methods=['POST'])
# @login_required
# def simpan_jam_nyala():
#     return simpan_by_ket("JAMNYALA600")

@app.route('/simpan_jam_nyala', methods=['POST'])
# @login_required
def simpan_jam_nyala():
    try:
        active_tab = request.form.get("active_tab", "jam_nyala")
        username = session.get('username', '').lower()
        billing_table = TABLE_MAPPING.get(username)

        if not billing_table:
            flash("User tidak dikenali.", "danger")
            return redirect(url_for('view_data1_v2', tab=active_tab))

        selected_kelompok = request.form.get("kdkelompok")
        jam_min = request.form.get("jam_nyala_min", 0)
        jam_max = request.form.get("jam_nyala_max", 9999)

        form_data = request.form
        idpels = {field.replace('hasil_pemeriksaan_', '') for field in form_data 
                 if field.startswith('hasil_pemeriksaan_')}
        idpels.update({field.replace('tindak_lanjut_', '') for field in form_data 
                      if field.startswith('tindak_lanjut_')})
        idpels.update({field.replace('stan_verifikasi_', '') for field in form_data 
                      if field.startswith('stan_verifikasi_')})
        
        updated_count = 0
        with engine.begin() as connection:
            for idpel in idpels:
                hasil = form_data.get(f'hasil_pemeriksaan_{idpel}', '').strip()
                tindak = form_data.get(f'tindak_lanjut_{idpel}', '').strip()
                stan = form_data.get(f'stan_verifikasi_{idpel}', '').strip()
                
                if hasil or tindak:
                    query = f"""
                        UPDATE {billing_table}
                        SET `HASIL PEMERIKSAAN` = :hasil,
                            `TINDAK LANJUT` = :tindak,
                            STAN_VERIFIKASI = :stan
                        WHERE IDPEL = :id 
                        AND KDKELOMPOK = :kelompok
                        AND `JAM NYALA` BETWEEN :min AND :max
                    """
                    connection.execute(text(query), {
                        'hasil': hasil,
                        'tindak': tindak,
                        'stan': stan,
                        'id': idpel,
                        'kelompok': selected_kelompok,
                        'min': jam_min,
                        'max': jam_max
                    })
                    updated_count += 1

        flash(f"{updated_count} data berhasil diperbarui untuk Jam Nyala ({jam_min}-{jam_max})!", "success")
    except Exception as e:
        import logging
        logging.exception(f"Error in simpan_jam_nyala: {str(e)}")
        flash("Gagal memperbarui data Jam Nyala", "danger")

    return redirect(url_for('view_data1_v2', tab=active_tab, kdkelompok=selected_kelompok))


# @app.route("/update_data", methods=["POST"])
# def update_data():
#     try:
#         data = request.get_json()
#         idpel = data.get("IDPEL")
#         column = data.get("column")
#         value = data.get("value")
#         table = data.get("table", "billing_result")  # default tabel

#         # Cek kolom valid
#         allowed_columns = ["HASIL_PEMERIKSAAN", "STAN_VERIFIKASI", "TINDAK_LANJUT"]
#         if column not in allowed_columns:
#             return jsonify({"status": "error", "message": "Kolom tidak diizinkan"}), 400

#         # Update database
#         with engine.begin() as conn:
#             sql = text(f"""
#                 UPDATE {table}
#                 SET `{column}` = :value
#                 WHERE IDPEL = :idpel
#             """)
#             conn.execute(sql, {"value": value, "idpel": idpel})

#         return jsonify({"status": "success", "message": "Data berhasil diperbarui"})
    
#     except Exception as e:
#         print(f"❌ Error update_data: {e}")
#         return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/update_verifikasi_single', methods=['POST'])
def update_verifikasi_single():
    try:
        # Ambil user & role dari session
        username = session.get('username', '').lower()
        role = session.get('role', '').upper()
        
        if not username:
            return jsonify({'success': False, 'message': 'Session expired, please login again'})

        # Ambil data dari frontend
        data = request.get_json()
        idpel = str(data.get('IDPEL', '')).strip().zfill(12)
        blth = str(data.get('BLTH', '')).strip()
        unitup = str(data.get('UNITUP', '')).strip()  # ✅ TERIMA UNITUP dari frontend
        column = data.get('column', '').strip()
        value = data.get('value', '').strip()

        logger.info(f"📥 Update: IDPEL={idpel}, BLTH={blth}, UNITUP={unitup}, {column}={value} by {username} ({role})")

        # Validasi input
        if not idpel or not blth or not column:
            return jsonify({'success': False, 'message': 'Missing required fields'})
        
        if not unitup:
            return jsonify({'success': False, 'message': 'UNITUP tidak ditemukan dari frontend'})

        # Validasi kolom
        allowed_columns = ['HASIL_PEMERIKSAAN', 'STAN_VERIFIKASI', 'TINDAK_LANJUT']
        if column not in allowed_columns:
            return jsonify({'success': False, 'message': 'Kolom tidak valid'})

        # ✅ Query UPDATE dengan filter UNITUP yang TEPAT
        if role == 'UP3':
            # UP3 bisa update data ULP manapun yang sedang dia lihat
            sql = text(f"""
                UPDATE billing 
                SET `{column}` = :value
                WHERE IDPEL = :idpel 
                  AND BLTH = :blth
                  AND UNITUP = :unitup
            """)
            params = {
                'value': value if value else None,
                'idpel': idpel,
                'blth': blth,
                'unitup': unitup
            }
            
        else:  # ULP
            sql = text(f"""
                UPDATE billing 
                SET `{column}` = :value, updated_by = :username
                WHERE IDPEL = :idpel 
                AND BLTH = :blth
                AND UNITUP = :unitup
            """)

            params = {
                'value': value if value else None,
                'idpel': idpel,
                'blth': blth,
                'unitup': unitup,
                'username': username
            }

        # Jalankan query
        with engine.begin() as conn:
            result = conn.execute(sql, params)
            rows_affected = result.rowcount

        logger.info(f"✅ Updated {rows_affected} rows for UNITUP={unitup}")

        if rows_affected > 0:
            return jsonify({
                'success': True,
                'message': f'✅ Berhasil update {column}',
                'updated': rows_affected,
                'unitup': unitup
            })
        else:
            return jsonify({
                'success': False,
                'message': f'❌ Tidak ada data yang diperbarui untuk UNITUP={unitup}'
            })

    except Exception as e:
        logger.error(f"❌ Error update_verifikasi_single: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': str(e)})


###Dashboard Monitoring


















# =================== MISSING IMPORT ===================
from flask import send_file


# =================== RUN APP ===================
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)



